"""Tests for the firm Excel export feature."""

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app import app, db
from models import Deal, DealCashflowEvent, Firm, FundMetadata


def test_export_firm_excel_returns_xlsx(client):
    """Authenticated user can download an Excel export for an accessible firm."""
    with app.app_context():
        firm = Firm.query.filter_by(slug="test-firm").first()
        deal = Deal(
            company_name="Acme Corp",
            fund_number="Fund I",
            sector="Tech",
            equity_invested=10.0,
            firm_id=firm.id,
        )
        db.session.add(deal)
        db.session.commit()
        firm_id = firm.id

    resp = client.get(f"/firms/{firm_id}/export-excel")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.content_type

    wb = load_workbook(BytesIO(resp.data))
    assert "Deals" in wb.sheetnames
    ws = wb["Deals"]
    headers = [c.value for c in ws[1]]
    assert "Company Name" in headers
    # Second row should contain our deal
    row2 = [c.value for c in ws[2]]
    assert "Acme Corp" in row2


def test_export_firm_excel_includes_optional_sheets(client):
    """Optional sheets appear only when supplemental data exists."""
    with app.app_context():
        firm = Firm.query.filter_by(slug="test-firm").first()
        deal = Deal(
            company_name="Beta Inc",
            fund_number="Fund II",
            equity_invested=5.0,
            firm_id=firm.id,
        )
        db.session.add(deal)
        db.session.flush()

        cf = DealCashflowEvent(
            deal_id=deal.id,
            event_date=date(2024, 1, 15),
            event_type="Capital Call",
            amount=1000000,
            firm_id=firm.id,
        )
        fm = FundMetadata(
            fund_number="Fund II",
            firm_id=firm.id,
            team_id=1,
            vintage_year=2020,
            strategy="Buyout",
        )
        db.session.add_all([cf, fm])
        db.session.commit()
        firm_id = firm.id

    resp = client.get(f"/firms/{firm_id}/export-excel")
    assert resp.status_code == 200

    wb = load_workbook(BytesIO(resp.data))
    assert "Cashflows" in wb.sheetnames
    assert "Fund Metadata" in wb.sheetnames


def test_export_firm_excel_empty_firm(client):
    """Exporting a firm with no deals returns a valid xlsx with header only."""
    with app.app_context():
        firm = Firm.query.filter_by(slug="test-firm").first()
        firm_id = firm.id

    resp = client.get(f"/firms/{firm_id}/export-excel")
    assert resp.status_code == 200

    wb = load_workbook(BytesIO(resp.data))
    ws = wb["Deals"]
    assert ws.max_row == 1  # header only


def test_export_firm_excel_inaccessible_firm(client):
    """Requesting export for a firm not accessible to the user's team redirects."""
    with app.app_context():
        other_firm = Firm(name="Other Firm", slug="other-firm")
        db.session.add(other_firm)
        db.session.commit()
        other_id = other_firm.id

    resp = client.get(f"/firms/{other_id}/export-excel")
    assert resp.status_code == 302  # redirect


def test_export_firm_excel_requires_login(anonymous_client):
    """Unauthenticated requests redirect to login."""
    resp = anonymous_client.get("/firms/1/export-excel")
    assert resp.status_code == 302
    assert "login" in resp.headers.get("Location", "").lower()
