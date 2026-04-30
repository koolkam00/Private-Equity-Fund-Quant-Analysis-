from datetime import date
from io import BytesIO
import re
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.exc import ProgrammingError

from models import (
    BenchmarkPoint,
    ChartBuilderTemplate,
    Deal,
    DealCashflowEvent,
    DealQuarterSnapshot,
    DealUnderwriteBaseline,
    Firm,
    FundMetadata,
    FundQuarterSnapshot,
    Team,
    TeamFirmAccess,
    TeamMembership,
    UploadIssue,
    db,
)


def _with_active_scope(deal):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    access = (
        TeamFirmAccess.query.filter_by(team_id=membership.team_id)
        .order_by(TeamFirmAccess.id.asc())
        .first()
    )
    assert access is not None
    deal.team_id = membership.team_id
    deal.firm_id = access.firm_id
    return deal


def _build_benchmark_workbook_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmarks"
    ws.append(["Asset Class", "Vintage Year", "Metric", "Quartile", "Value"])
    for row in rows:
        ws.append([row["asset_class"], row["vintage_year"], row["metric"], row["quartile"], row["value"]])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _sidebar_nav_html(response):
    html = response.data.decode("utf-8")
    match = re.search(r'<nav class="nav">(.*?)</nav>', html, re.DOTALL)
    assert match is not None
    return match.group(1)


def test_index_redirect(client):
    response = client.get("/")
    assert response.status_code == 302
    assert "/dashboard" in response.location


def test_dashboard_page(client):
    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"Portfolio Dashboard" in response.data
    assert b"Key Takeaway" in response.data
    assert b"Executive" in response.data
    assert b'id="bridge-lever-table-body"' in response.data
    assert b"Capital Value Loss Ratio" in response.data
    assert b"Download 4 Analysis PDFs" in response.data
    assert b"Upload Deals" in response.data
    assert b'id="firm-picker-trigger"' in response.data
    assert b'id="firm-picker-modal"' in response.data
    assert b'id="firm-picker-data-payload"' in response.data
    assert b"Cmd/Ctrl+K" in response.data
    assert b'global-firm-select' not in response.data


def test_readyz_reports_error_until_alembic_revision_exists(anonymous_client):
    response = anonymous_client.get("/readyz")
    assert response.status_code == 500
    body = response.get_json()
    assert body["status"] == "error"
    assert "alembic_version" in body["missing_tables"]

    with db.engine.begin() as conn:
        conn.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
        conn.execute(text("INSERT INTO alembic_version (version_num) VALUES ('4a62775748c8')"))

    healthy = anonymous_client.get("/readyz")
    assert healthy.status_code == 200
    body = healthy.get_json()
    assert body["status"] == "ok"
    assert body["revision"] == "4a62775748c8"


def test_dashboard_schema_failure_rolls_back_and_renders_fallback(client, monkeypatch):
    import legacy_app

    rollback_calls = {"count": 0}
    original_rollback = db.session.rollback

    def _spy_rollback():
        rollback_calls["count"] += 1
        return original_rollback()

    def _raise_programming_error(*args, **kwargs):
        raise ProgrammingError("SELECT * FROM benchmark_points", {}, Exception('relation "benchmark_points" does not exist'))

    monkeypatch.setattr(db.session, "rollback", _spy_rollback)
    monkeypatch.setattr(legacy_app, "_build_filtered_deals_context", _raise_programming_error)

    response = client.get("/dashboard")

    assert response.status_code == 503
    assert b"Portfolio Dashboard" in response.data
    assert b"Database schema is not ready." in response.data
    assert rollback_calls["count"] >= 1


def test_analysis_series_schema_failure_returns_json_503(client, monkeypatch):
    import legacy_app

    def _raise_programming_error(*args, **kwargs):
        raise ProgrammingError("SELECT * FROM benchmark_points", {}, Exception('relation "benchmark_points" does not exist'))

    monkeypatch.setattr(legacy_app, "_build_filtered_deals_context", _raise_programming_error)

    response = client.get("/api/analysis/benchmarking/series")

    assert response.status_code == 503
    body = response.get_json()
    assert body["error"] == "database_schema_not_ready"
    assert "db-upgrade" in body["message"]


def test_sidebar_primary_order_and_analysis_grouping(client):
    response = client.get("/dashboard")
    assert response.status_code == 200
    nav_html = _sidebar_nav_html(response)
    assert "Upload Deals" in nav_html
    assert "> Upload</a>" not in nav_html

    primary_labels = [
        "Dashboard",
        "Benchmarking Analysis",
        "Track Record",
        "Value Creation (EBITDA)",
        "Value Creation (Revenue)",
        "Value Creation (Add-Ons)",
        "Value Creation (Add-Ons Revenue)",
        "IC Memo",
        "Deals",
        "Download 4 Analysis PDFs",
        "Upload Deals",
    ]
    primary_positions = []
    for label in primary_labels:
        idx = nav_html.find(label)
        assert idx != -1, label
        primary_positions.append(idx)
    assert primary_positions == sorted(primary_positions)

    analysis_idx = nav_html.find('<div class="nav-section">Analysis</div>')
    assert analysis_idx != -1
    assert primary_positions[-1] < analysis_idx

    analysis_labels = [
        "NAV at Risk",
        "Public Market Comparison",
        "LP Due Diligence Memo",
        "Fund Liquidity",
        "Underwrite vs Outcome",
        "Valuation Quality",
        "Exit Readiness",
        "Stress Lab",
        "Deal Trajectory",
        "Chart Builder",
    ]
    for label in analysis_labels:
        idx = nav_html.find(label)
        assert idx != -1, label
        assert idx > analysis_idx, label

    removed_labels = [
        "LP Liquidity Quality",
        "Liquidity Forecast",
        "Manager Consistency",
        "Benchmark Confidence",
        "Reporting Quality",
        "Fee Drag",
    ]
    for label in removed_labels:
        assert label not in nav_html


def test_upload_page(client):
    response = client.get("/upload")
    assert response.status_code == 200
    assert b"Upload Deal Template" in response.data
    assert b"Download Current Deal Template" in response.data
    assert b"Benchmark File" in response.data
    assert b"Upload Benchmarks" in response.data
    assert b"No benchmark dataset is currently loaded for your team." in response.data
    assert b"Recent Uploads" in response.data
    assert b"Firm Currency" in response.data


def test_upload_page_lists_recent_batches_for_active_firm(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    access = TeamFirmAccess.query.filter_by(team_id=membership.team_id).first()
    assert access is not None

    deal = Deal(
        company_name="Upload History Co",
        fund_number="Fund Hist",
        team_id=membership.team_id,
        firm_id=access.firm_id,
        upload_batch="batchhist",
        equity_invested=10,
        realized_value=12,
        unrealized_value=0,
    )
    db.session.add(deal)
    db.session.commit()

    response = client.get("/upload")
    assert response.status_code == 200
    assert b"batchhist" in response.data
    assert b"Delete Upload" in response.data


def test_delete_upload_batch_removes_only_active_firm_records(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    access = TeamFirmAccess.query.filter_by(team_id=membership.team_id).first()
    assert access is not None
    active_firm_id = access.firm_id
    other_firm = Firm(name="Other Upload Firm", slug="other-upload-firm")
    db.session.add(other_firm)
    db.session.flush()

    deal = Deal(
        company_name="Delete Upload Co",
        fund_number="Fund Del",
        team_id=membership.team_id,
        firm_id=active_firm_id,
        upload_batch="deadbeef",
        equity_invested=20,
        realized_value=25,
        unrealized_value=0,
        investment_date=date(2020, 1, 1),
        exit_date=date(2023, 1, 1),
    )
    db.session.add(deal)
    db.session.flush()
    db.session.add_all(
        [
            DealCashflowEvent(
                deal_id=deal.id,
                event_date=date(2021, 1, 1),
                event_type="Distribution",
                amount=5,
                team_id=membership.team_id,
                firm_id=active_firm_id,
                upload_batch="deadbeef",
            ),
            DealQuarterSnapshot(
                deal_id=deal.id,
                quarter_end=date(2022, 12, 31),
                team_id=membership.team_id,
                firm_id=active_firm_id,
                upload_batch="deadbeef",
            ),
            DealUnderwriteBaseline(
                deal_id=deal.id,
                baseline_date=date(2020, 1, 1),
                team_id=membership.team_id,
                firm_id=active_firm_id,
                upload_batch="deadbeef",
            ),
            FundQuarterSnapshot(
                fund_number="Fund Del",
                quarter_end=date(2022, 12, 31),
                team_id=membership.team_id,
                firm_id=active_firm_id,
                upload_batch="deadbeef",
            ),
            UploadIssue(
                issue_report_id="issue-del-1",
                team_id=membership.team_id,
                firm_id=active_firm_id,
                upload_batch="deadbeef",
                file_type="deals",
                row_number=2,
                severity="warning",
                message="Synthetic issue",
            ),
        ]
    )

    foreign_deal = Deal(
        company_name="Other Firm Batch Co",
        fund_number="Fund X",
        team_id=membership.team_id,
        firm_id=other_firm.id,
        upload_batch="deadbeef",
        equity_invested=15,
        realized_value=16,
        unrealized_value=0,
    )
    db.session.add(foreign_deal)
    db.session.commit()

    response = client.post("/upload/batches/deadbeef/delete", follow_redirects=False)
    assert response.status_code == 302
    assert response.location.endswith("/upload")

    assert Deal.query.filter_by(firm_id=active_firm_id, upload_batch="deadbeef").count() == 0
    assert DealCashflowEvent.query.filter_by(firm_id=active_firm_id, upload_batch="deadbeef").count() == 0
    assert DealQuarterSnapshot.query.filter_by(firm_id=active_firm_id, upload_batch="deadbeef").count() == 0
    assert DealUnderwriteBaseline.query.filter_by(firm_id=active_firm_id, upload_batch="deadbeef").count() == 0
    assert FundQuarterSnapshot.query.filter_by(firm_id=active_firm_id, upload_batch="deadbeef").count() == 0
    assert UploadIssue.query.filter_by(firm_id=active_firm_id, upload_batch="deadbeef").count() == 0
    assert Deal.query.filter_by(firm_id=other_firm.id, upload_batch="deadbeef").count() == 1


def test_download_deal_template(client):
    response = client.get("/upload/deals/template")
    assert response.status_code == 200
    assert "attachment;" in response.headers.get("Content-Disposition", "")
    assert "PE_Fund_Data_Template.xlsx" in response.headers.get("Content-Disposition", "")
    wb = load_workbook(BytesIO(response.data), read_only=True)
    ws = wb["Deals"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Firm Currency" in headers
    assert "As Of Date" in headers
    for sheet_name in ["Cashflows", "Deal Quarterly", "Fund Quarterly", "Underwrite"]:
        sheet = wb[sheet_name]
        optional_headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert "Firm Name" in optional_headers


def test_download_benchmark_template(client):
    response = client.get("/upload/benchmarks/template")
    assert response.status_code == 200
    assert "attachment;" in response.headers.get("Content-Disposition", "")
    assert "PE_Benchmark_Template.xlsx" in response.headers.get("Content-Disposition", "")
    wb = load_workbook(BytesIO(response.data), read_only=True)
    ws = wb["Benchmarks"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["Asset Class", "Vintage Year", "Metric", "Quartile", "Value"]


def test_upload_benchmarks_route_success(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    rows = [
        {"asset_class": "Buyout", "vintage_year": 2019, "metric": "Net IRR", "quartile": "Lower Quartile", "value": 0.12},
        {"asset_class": "Buyout", "vintage_year": 2019, "metric": "Net IRR", "quartile": "Median", "value": 0.16},
        {"asset_class": "Buyout", "vintage_year": 2019, "metric": "Net IRR", "quartile": "Upper Quartile", "value": 0.2},
        {"asset_class": "Buyout", "vintage_year": 2019, "metric": "Net IRR", "quartile": "Top 5%", "value": 0.27},
    ]
    payload = {
        "file": (_build_benchmark_workbook_bytes(rows), "benchmarks.xlsx"),
    }
    response = client.post(
        "/upload/benchmarks",
        data=payload,
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b"Loaded 4 benchmark rows" in response.data
    assert b"Delete Benchmark Data" in response.data
    assert BenchmarkPoint.query.filter_by(team_id=membership.team_id).count() == 4


def test_delete_benchmark_dataset_is_team_scoped(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None

    other_team = Team(name="Other Bench Team", slug="other-bench-team")
    db.session.add(other_team)
    db.session.flush()

    db.session.add_all(
        [
            BenchmarkPoint(
                team_id=membership.team_id,
                asset_class="Buyout",
                vintage_year=2019,
                metric="net_irr",
                quartile="median",
                value=0.16,
                upload_batch="bench-a",
            ),
            BenchmarkPoint(
                team_id=other_team.id,
                asset_class="Growth",
                vintage_year=2020,
                metric="net_moic",
                quartile="median",
                value=1.7,
                upload_batch="bench-b",
            ),
        ]
    )
    db.session.commit()

    response = client.post("/upload/benchmarks/delete", follow_redirects=False)
    assert response.status_code == 302
    assert response.location.endswith("/upload")

    assert BenchmarkPoint.query.filter_by(team_id=membership.team_id).count() == 0
    assert BenchmarkPoint.query.filter_by(team_id=other_team.id).count() == 1


def test_dashboard_benchmark_selector_and_labels(client):
    deal = Deal(
        company_name="Bench Co",
        fund_number="Fund Bench",
        investment_date=date(2019, 1, 1),
        equity_invested=100,
        realized_value=170,
        unrealized_value=0,
        net_irr=0.24,
        net_moic=2.3,
        net_dpi=1.4,
    )
    db.session.add(_with_active_scope(deal))
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    db.session.add_all(
        [
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="lower_quartile", value=0.12),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="median", value=0.17),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="upper_quartile", value=0.21),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="top_5", value=0.3),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="lower_quartile", value=1.5),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="median", value=1.9),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="upper_quartile", value=2.2),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="top_5", value=2.8),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="lower_quartile", value=0.6),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="median", value=1.3),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="upper_quartile", value=1.5),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="top_5", value=1.8),
        ]
    )
    db.session.commit()

    response = client.get("/dashboard?benchmark_asset_class=Buyout")
    assert response.status_code == 200
    assert b"Benchmark Asset Class" in response.data
    assert b"Net IRR Benchmark" in response.data
    assert b"Net MOIC Benchmark" in response.data
    assert b"Net DPI Benchmark" in response.data
    assert b"1st Quartile" in response.data
    assert b"2nd Quartile" in response.data


def test_dashboard_benchmark_exact_vintage_match_required(client):
    deal = Deal(
        company_name="Bench No Match",
        fund_number="Fund Bench N/A",
        investment_date=date(2020, 1, 1),
        equity_invested=100,
        realized_value=170,
        unrealized_value=0,
        net_irr=0.24,
        net_moic=2.3,
        net_dpi=1.4,
    )
    db.session.add(_with_active_scope(deal))
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    db.session.add(
        BenchmarkPoint(
            team_id=membership.team_id,
            asset_class="Buyout",
            vintage_year=2019,
            metric="net_irr",
            quartile="lower_quartile",
            value=0.12,
        )
    )
    db.session.commit()

    response = client.get("/dashboard?benchmark_asset_class=Buyout")
    assert response.status_code == 200
    assert b"Net IRR Benchmark" in response.data
    assert b"benchmark-rank-na" in response.data


def test_dashboard_benchmark_selection_persists_in_session(client):
    deal = Deal(
        company_name="Bench Persist",
        fund_number="Fund Persist",
        investment_date=date(2019, 1, 1),
        equity_invested=100,
        realized_value=150,
        unrealized_value=0,
        net_irr=0.20,
        net_moic=2.0,
        net_dpi=1.2,
    )
    db.session.add(_with_active_scope(deal))
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    db.session.add_all(
        [
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="lower_quartile", value=0.12),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="median", value=0.16),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="upper_quartile", value=0.2),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="top_5", value=0.28),
        ]
    )
    db.session.commit()

    first = client.get("/dashboard?benchmark_asset_class=Buyout")
    assert first.status_code == 200

    second = client.get("/dashboard")
    assert second.status_code == 200
    html = second.data.decode("utf-8")
    assert 'name="benchmark_asset_class"' in html
    assert '<option value="Buyout" selected>' in html


def test_deals_page(client):
    response = client.get("/deals")
    assert response.status_code == 200
    assert b"Deals" in response.data


def test_deals_page_grouped_subtotals_and_detail_contract(client):
    deals = [
        Deal(
            company_name="Alpha",
            fund_number="Fund I",
            status="Fully Realized",
            geography="US",
            year_invested=2020,
            equity_invested=100,
            realized_value=170,
            unrealized_value=0,
            entry_revenue=50,
            entry_ebitda=10,
            entry_enterprise_value=120,
            entry_net_debt=30,
            exit_revenue=70,
            exit_ebitda=15,
            exit_enterprise_value=190,
            exit_net_debt=25,
            investment_date=date(2020, 1, 1),
            exit_date=date(2024, 1, 1),
        ),
        Deal(
            company_name="Beta",
            fund_number="Fund I",
            status="Unrealized",
            geography="US",
            year_invested=2021,
            equity_invested=60,
            realized_value=0,
            unrealized_value=85,
            entry_revenue=40,
            entry_ebitda=8,
            entry_enterprise_value=90,
            entry_net_debt=20,
            exit_revenue=60,
            exit_ebitda=12,
            exit_enterprise_value=130,
            exit_net_debt=18,
            investment_date=date(2021, 1, 1),
        ),
        Deal(
            company_name="Gamma",
            fund_number="Fund II",
            status="Partially Realized",
            geography="UK",
            year_invested=2019,
            equity_invested=80,
            realized_value=50,
            unrealized_value=35,
            entry_revenue=45,
            entry_ebitda=9,
            entry_enterprise_value=100,
            entry_net_debt=28,
            exit_revenue=58,
            exit_ebitda=11,
            exit_enterprise_value=125,
            exit_net_debt=23,
            investment_date=date(2019, 1, 1),
            exit_date=date(2023, 7, 1),
        ),
    ]
    db.session.add_all([_with_active_scope(d) for d in deals])
    db.session.commit()

    response = client.get("/deals")
    assert response.status_code == 200
    assert b"All Funds Summary" in response.data
    assert b"All 1 Fund I Fully Realized Investments" in response.data
    assert b"All 1 Fund I Unrealized Investments" in response.data
    assert b"All 2 Fund I Investments" in response.data
    assert response.data.count(b"<strong>All 1 Fund I Unrealized Investments</strong>") == 1
    assert response.data.count(b"<strong>All 1 Unrealized Investments</strong>") == 1
    assert b"EBITDA Margin" in response.data
    assert b"Revenue (USD $M)" in response.data
    assert b"EBITDA (USD $M)" in response.data
    assert b"TEV (USD $M)" in response.data
    assert b"Net Debt (USD $M)" in response.data
    assert b"Revenue CAGR" in response.data
    assert b"Revenue Cumulative" in response.data
    assert b"EBITDA CAGR" in response.data
    assert b"EBITDA Cumulative" in response.data
    assert b"<th>Sector</th>" not in response.data
    assert b"<th>Geo</th>" not in response.data
    assert b'class="deal-row"' in response.data
    assert b"rollup-expand-btn" in response.data
    assert b'id="detail-' in response.data
    assert b'id="rollup-detail-' in response.data
    assert b'id="deal-bridge-table-' in response.data
    assert b'id="rollup-bridge-table-' in response.data
    assert b'id="deals-rollup-details-payload"' in response.data
    assert b"Entry vs Exit Comparison" in response.data
    assert b"Rollup Bridge (Fund Pro-Rata)" in response.data
    assert b"Avg" in response.data
    assert b"Wtd" in response.data
    assert b"data-sort=" not in response.data


def test_deals_analysis_selection_recomputes_rollups_without_hiding_deals(client):
    alpha = Deal(
        company_name="Alpha Select",
        fund_number="Fund Select",
        status="Fully Realized",
        equity_invested=100,
        realized_value=180,
        unrealized_value=0,
        investment_date=date(2020, 1, 1),
        exit_date=date(2024, 1, 1),
    )
    beta = Deal(
        company_name="Beta Excluded",
        fund_number="Fund Select",
        status="Unrealized",
        equity_invested=50,
        realized_value=0,
        unrealized_value=60,
        investment_date=date(2021, 1, 1),
    )
    db.session.add_all([_with_active_scope(alpha), _with_active_scope(beta)])
    db.session.commit()

    response = client.get(f"/deals/analysis?excluded_deal={beta.id}")
    assert response.status_code == 200
    html = response.data.decode("utf-8")

    assert "Deals Analysis" in html
    assert "1 of 2 included" in html
    assert "Alpha Select" in html
    assert "Beta Excluded" in html
    assert f'data-deal-id="{beta.id}" data-deal-included="0"' in html
    assert "deal-include-checkbox" in html
    assert "All 1 Fund Select Fully Realized Investments" in html
    assert "All 1 Fund Select Investments" in html
    assert "All 2 Fund Select Investments" not in html

    normal_response = client.get("/deals")
    assert normal_response.status_code == 200
    assert b"All 2 Fund Select Investments" in normal_response.data


def test_track_record_page(client):
    response = client.get("/track-record")
    assert response.status_code == 200
    assert b"Deal Level Track Record" in response.data


def test_track_record_pdf_download(client):
    response = client.get("/track-record/pdf")
    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    assert "attachment;" in response.headers.get("Content-Disposition", "")
    assert "track_record_print_ready.pdf" in response.headers.get("Content-Disposition", "")


def test_ic_pdf_pack_download(client):
    response = client.get("/reports/ic-pdf-pack")
    assert response.status_code == 200
    assert response.mimetype == "application/zip"
    disposition = response.headers.get("Content-Disposition", "")
    assert "attachment;" in disposition
    assert ".zip" in disposition
    assert "Analysis PDF Pack" in disposition
    assert "As%20Of" not in disposition and "As Of" not in disposition
    assert re.search(r"Analysis PDF Pack \d{1,2}\.\d{1,2}\.\d{2}\.zip", disposition), disposition

    archive = ZipFile(BytesIO(response.data))
    names = sorted(archive.namelist())
    assert len(names) == 4
    assert all(name.endswith(".pdf") for name in names)

    expected_analysis_names = (
        "Deal Level Track Record",
        "Value Creation Analysis by EBITDA",
        "Value Creation Analysis by Revenue",
        "Benchmarking Analysis",
    )
    for analysis_name in expected_analysis_names:
        assert any(analysis_name in name for name in names)

    for name in names:
        assert "As Of " not in name
        assert re.search(r"\d{1,2}\.\d{1,2}\.\d{2}\.pdf$", name), name
        payload = archive.read(name)
        assert payload.startswith(b"%PDF"), name


def test_ic_pdf_pack_live_page_renders_export_links(client):
    response = client.get("/reports/ic-pdf-pack/live")
    assert response.status_code == 200
    assert b"Start 4-PDF Export" in response.data
    assert b"/track-record/pdf" in response.data
    assert b"/analysis/vca-ebitda" in response.data
    assert b"/analysis/vca-revenue" in response.data
    assert b"/analysis/benchmarking" in response.data
    assert b"autoprint=1" in response.data


def test_track_record_page_renders_template_columns_and_net_performance(client):
    deal = Deal(
        company_name="Track Co",
        fund_number="Fund IX",
        status="Unrealized",
        investment_date=date(2022, 1, 1),
        equity_invested=100,
        realized_value=20,
        unrealized_value=110,
        fund_size=500,
        net_irr=0.14,
        net_moic=1.7,
        net_dpi=0.8,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/track-record")
    assert response.status_code == 200
    assert b"% of Fund Size" in response.data
    assert b"Gross IRR" in response.data
    assert b"All Funds Summary" in response.data
    assert b"Net MOIC" in response.data
    assert b"DPI" in response.data


def test_track_record_page_repeats_column_header_for_each_fund(client):
    db.session.add_all(
        [
            _with_active_scope(
                Deal(
                    company_name="Header Alpha Co",
                    fund_number="Fund Header A",
                    status="Unrealized",
                    investment_date=date(2022, 1, 1),
                    equity_invested=100,
                    unrealized_value=120,
                )
            ),
            _with_active_scope(
                Deal(
                    company_name="Header Beta Co",
                    fund_number="Fund Header B",
                    status="Fully Realized",
                    investment_date=date(2020, 1, 1),
                    exit_date=date(2024, 1, 1),
                    equity_invested=100,
                    realized_value=150,
                )
            ),
        ]
    )
    db.session.commit()

    response = client.get("/track-record")

    assert response.status_code == 200
    html = response.data.decode("utf-8")
    assert html.count('class="tr-column-header"') == 2
    assert html.find("Fund Header A") < html.find("Header Alpha Co")
    assert html.find("Fund Header B") < html.find("Header Beta Co")


def test_track_record_page_sorts_funds_by_vintage_year(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    access = TeamFirmAccess.query.filter_by(team_id=membership.team_id).order_by(TeamFirmAccess.id.asc()).first()
    assert access is not None

    db.session.add_all(
        [
            FundMetadata(team_id=membership.team_id, firm_id=access.firm_id, fund_number="Fund Zeta", vintage_year=2018),
            FundMetadata(team_id=membership.team_id, firm_id=access.firm_id, fund_number="Fund Alpha", vintage_year=2021),
        ]
    )
    db.session.add_all(
        [
            _with_active_scope(
                Deal(
                    company_name="Vintage Zeta Co",
                    fund_number="Fund Zeta",
                    status="Fully Realized",
                    investment_date=date(2021, 1, 1),
                    equity_invested=100,
                    realized_value=150,
                    unrealized_value=0,
                )
            ),
            _with_active_scope(
                Deal(
                    company_name="Vintage Alpha Co",
                    fund_number="Fund Alpha",
                    status="Fully Realized",
                    investment_date=date(2019, 1, 1),
                    equity_invested=100,
                    realized_value=140,
                    unrealized_value=0,
                )
            ),
        ]
    )
    db.session.commit()

    response = client.get("/track-record")
    assert response.status_code == 200
    html = response.data.decode("utf-8")
    assert html.find("Fund Zeta") < html.find("Fund Alpha")


def test_track_record_page_marks_negative_values_red(client):
    deal = Deal(
        company_name="Negative Track Co",
        fund_number="Fund Negative",
        status="Fully Realized",
        investment_date=date(2020, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=80,
        unrealized_value=0,
        irr=-0.05,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/track-record")
    assert response.status_code == 200
    assert b'class="value-negative">-5.0%' in response.data


def test_ic_memo_page_renders_and_has_print_action(client):
    response = client.get("/ic-memo")
    assert response.status_code == 200
    assert b"IC Memo Presentation" in response.data
    assert b"Download as PDF" in response.data
    assert b"Page 1. Executive Summary" in response.data
    assert b"Capital Value Loss Ratio" in response.data


def test_ic_memo_path_fund_scope_overrides_query_fund_and_keeps_filters(client):
    d1 = Deal(
        company_name="Alpha Memo",
        fund_number="Fund I",
        status="Fully Realized",
        sector="Tech",
        geography="US",
        year_invested=2021,
        equity_invested=100,
        realized_value=150,
        unrealized_value=0,
    )
    d2 = Deal(
        company_name="Beta Memo",
        fund_number="Fund II",
        status="Fully Realized",
        sector="Tech",
        geography="US",
        year_invested=2021,
        equity_invested=100,
        realized_value=200,
        unrealized_value=0,
    )
    db.session.add_all([_with_active_scope(d1), _with_active_scope(d2)])
    db.session.commit()

    response = client.get("/ic-memo/Fund%20I?fund=Fund+II")
    assert response.status_code == 200
    assert b"Fund Scope: Fund I" in response.data
    assert b"Alpha Memo" in response.data
    assert b"Beta Memo" not in response.data

    response_filtered = client.get("/ic-memo/Fund%20I?status=Unrealized")
    assert response_filtered.status_code == 200
    assert b"No deals available for this IC memo filter set." in response_filtered.data


def test_methodology_page_renders_with_print_and_anchor_ids(client):
    response = client.get("/methodology")
    assert response.status_code == 200
    assert b"Calculation Methodology &amp; Audit" in response.data
    assert b"Download as PDF" in response.data
    assert b'id="metric-gross-moic"' in response.data
    assert b'id="metric-loss-ratio-capital"' in response.data
    assert b'id="metric-bridge-revenue"' in response.data
    assert b'id="metric-tvpi"' in response.data


def test_audit_alias_redirects_to_methodology(client):
    response = client.get("/audit")
    assert response.status_code == 302
    assert "/methodology" in response.location


def test_dashboard_has_methodology_jump_links(client):
    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"/methodology#metric-gross-moic" in response.data
    assert b"/methodology#metric-tev-ebitda" in response.data


def test_chart_builder_page_renders(client):
    response = client.get("/analysis/chart-builder")
    assert response.status_code == 200
    assert b"Chart Builder" in response.data
    assert b'id="chart-builder-root"' in response.data
    assert b'id="chart-builder-catalog-payload"' in response.data


def test_chart_builder_catalog_api(client):
    response = client.get("/api/chart-builder/catalog")
    assert response.status_code == 200
    payload = response.get_json()
    keys = [row["key"] for row in payload["sources"]]
    assert keys == ["deals", "deal_quarterly", "fund_quarterly", "cashflows", "underwrite", "benchmarks"]


def test_chart_builder_query_api_applies_global_and_local_filters(client):
    d1 = Deal(
        company_name="CB One",
        fund_number="Fund I",
        sector="Tech",
        status="Fully Realized",
        year_invested=2021,
        equity_invested=100,
        realized_value=150,
        unrealized_value=0,
    )
    d2 = Deal(
        company_name="CB Two",
        fund_number="Fund I",
        sector="Tech",
        status="Unrealized",
        year_invested=2021,
        equity_invested=100,
        realized_value=0,
        unrealized_value=110,
    )
    d3 = Deal(
        company_name="CB Three",
        fund_number="Fund II",
        sector="Health",
        status="Fully Realized",
        year_invested=2021,
        equity_invested=100,
        realized_value=130,
        unrealized_value=0,
    )
    db.session.add_all([_with_active_scope(d1), _with_active_scope(d2), _with_active_scope(d3)])
    db.session.commit()

    response = client.post(
        "/api/chart-builder/query?sector=Tech",
        json={
            "source": "deals",
            "chart_type": "bar",
            "x": {"field": "sector"},
            "y": [{"field": "company_name", "agg": "count"}],
            "filters": [{"field": "status", "op": "eq", "value": "Fully Realized"}],
        },
    )
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["labels"] == ["Tech"]
    assert payload["datasets"][0]["data"][0] == 1


def test_chart_builder_query_api_rejects_invalid_field(client):
    response = client.post(
        "/api/chart-builder/query",
        json={
            "source": "deals",
            "chart_type": "bar",
            "x": {"field": "fund_number"},
            "y": [{"field": "not_a_field", "agg": "sum"}],
        },
    )
    assert response.status_code == 400
    payload = response.get_json()
    assert "Unknown y field" in payload["error"]


def test_chart_builder_template_crud_is_team_scoped(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    other_team = Team(name="Template Other Team", slug="template-other-team")
    db.session.add(other_team)
    db.session.flush()
    db.session.add(
        ChartBuilderTemplate(
            team_id=other_team.id,
            name="Other Team Template",
            source="deals",
            config_json='{"config_version":1,"cards":[]}',
        )
    )
    db.session.commit()

    create_response = client.post(
        "/api/chart-builder/templates",
        json={
            "name": "Team Template A",
            "source": "deals",
            "config": {"config_version": 1, "cards": []},
        },
    )
    assert create_response.status_code == 201
    created = create_response.get_json()
    template_id = created["id"]

    list_response = client.get("/api/chart-builder/templates")
    assert list_response.status_code == 200
    listed_names = [row["name"] for row in list_response.get_json()["templates"]]
    assert "Team Template A" in listed_names
    assert "Other Team Template" not in listed_names

    update_response = client.put(
        f"/api/chart-builder/templates/{template_id}",
        json={"name": "Team Template B", "config": {"config_version": 1, "cards": [{"source": "deals"}]}},
    )
    assert update_response.status_code == 200
    assert update_response.get_json()["name"] == "Team Template B"

    delete_response = client.delete(f"/api/chart-builder/templates/{template_id}")
    assert delete_response.status_code == 200
    assert delete_response.get_json()["deleted"] is True


def test_deals_page_recovers_missing_deals_table(client):
    with client.application.app_context():
        db.drop_all()

    response = client.get("/deals")
    assert response.status_code == 302
    assert "/auth/login" in response.location


def test_dashboard_filter_context(client):
    d1 = Deal(
        company_name="Alpha",
        fund_number="Fund I",
        sector="Tech",
        geography="US",
        year_invested=2021,
        exit_type="Strategic Sale",
        lead_partner="Jane Doe",
        security_type="Common Equity",
        deal_type="Platform",
        entry_channel="Proprietary",
        equity_invested=100,
        realized_value=150,
        unrealized_value=0,
    )
    d2 = Deal(
        company_name="Beta",
        fund_number="Fund II",
        sector="Health",
        geography="UK",
        year_invested=2022,
        exit_type="Secondary Buyout",
        lead_partner="Alex Reed",
        security_type="Preferred Equity",
        deal_type="Add-on",
        entry_channel="Broad Auction",
        equity_invested=100,
        realized_value=90,
        unrealized_value=0,
    )
    db.session.add_all([_with_active_scope(d1), _with_active_scope(d2)])
    db.session.commit()

    response = client.get("/dashboard?fund=Fund+I&exit_type=Strategic+Sale")
    assert response.status_code == 200
    assert b"Fund I" in response.data
    assert b"All Exit Types" in response.data


def test_dashboard_shows_latest_visible_as_of_date(client):
    d1 = Deal(
        company_name="As Of One",
        fund_number="Fund I",
        year_invested=2021,
        as_of_date=date(2025, 12, 31),
        equity_invested=100,
        realized_value=120,
        unrealized_value=0,
    )
    d2 = Deal(
        company_name="As Of Two",
        fund_number="Fund II",
        year_invested=2022,
        as_of_date=date(2026, 1, 31),
        equity_invested=120,
        realized_value=130,
        unrealized_value=0,
    )
    db.session.add_all([_with_active_scope(d1), _with_active_scope(d2)])
    db.session.commit()

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"As of 01/31/26" in response.data


def test_analysis_non_vca_page_shows_as_of_metadata(client):
    deal = Deal(
        company_name="Analysis As Of Co",
        fund_number="Fund AO",
        status="Unrealized",
        year_invested=2023,
        as_of_date=date(2025, 11, 30),
        equity_invested=100,
        realized_value=0,
        unrealized_value=110,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/analysis/underwrite-outcome")
    assert response.status_code == 200
    assert b"As of 11/30/25" in response.data


def test_dashboard_fund_summary_table(client):
    d1 = Deal(
        company_name="Alpha",
        fund_number="Fund I",
        status="Unrealized",
        investment_date=date(2017, 1, 1),
        equity_invested=100,
        realized_value=0,
        unrealized_value=140,
        fund_size=500,
        net_irr=0.14,
        net_moic=1.8,
        net_dpi=0.4,
    )
    d2 = Deal(
        company_name="Beta",
        fund_number="Fund II",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        equity_invested=120,
        realized_value=210,
        unrealized_value=0,
        fund_size=600,
        net_irr=0.22,
        net_moic=2.1,
        net_dpi=1.1,
    )
    db.session.add_all([_with_active_scope(d1), _with_active_scope(d2)])
    db.session.commit()

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"Fund Summary" in response.data
    assert b"Vintage Year" in response.data
    assert b"Fund Size" in response.data
    assert b"Net IRR" in response.data
    assert b"Net MOIC" in response.data
    assert b"Net DPI" in response.data
    assert b"Fund I" in response.data
    assert b"Fund II" in response.data
    assert b"2017" in response.data
    assert b"2019" in response.data
    assert b"14.0%" in response.data
    assert b"1.80x" in response.data
    assert b"0.40x" in response.data


def test_dashboard_fund_summary_sorted_by_vintage_year(client):
    newer = Deal(
        company_name="NewCo",
        fund_number="Fund Newer",
        status="Unrealized",
        investment_date=date(2021, 1, 1),
        equity_invested=100,
        realized_value=0,
        unrealized_value=110,
        fund_size=600,
        net_irr=0.12,
        net_moic=1.4,
        net_dpi=0.2,
    )
    older = Deal(
        company_name="OldCo",
        fund_number="Fund Older",
        status="Unrealized",
        investment_date=date(2016, 1, 1),
        equity_invested=80,
        realized_value=0,
        unrealized_value=120,
        fund_size=500,
        net_irr=0.15,
        net_moic=1.6,
        net_dpi=0.3,
    )
    db.session.add_all([_with_active_scope(newer), _with_active_scope(older)])
    db.session.commit()

    response = client.get("/dashboard")
    assert response.status_code == 200
    html = response.data.decode("utf-8")
    match = re.search(r"<h3>Fund Summary</h3>.*?<tbody>(.*?)</tbody>", html, re.S)
    assert match is not None
    table_body = match.group(1)
    assert table_body.find("Fund Older") < table_body.find("Fund Newer")


def test_api_dashboard_series_qualitative_filters(client):
    d1 = Deal(
        company_name="Gamma",
        fund_number="Fund V",
        sector="Tech",
        geography="US",
        status="Fully Realized",
        year_invested=2020,
        exit_type="Strategic Sale",
        lead_partner="Jane Doe",
        security_type="Common Equity",
        deal_type="Platform",
        entry_channel="Proprietary",
        equity_invested=100,
        realized_value=200,
        unrealized_value=0,
    )
    d2 = Deal(
        company_name="Delta",
        fund_number="Fund V",
        sector="Tech",
        geography="US",
        status="Fully Realized",
        year_invested=2020,
        exit_type="Secondary Buyout",
        lead_partner="Alex Reed",
        security_type="Preferred Equity",
        deal_type="Add-on",
        entry_channel="Broad Auction",
        equity_invested=100,
        realized_value=120,
        unrealized_value=0,
    )
    db.session.add_all([_with_active_scope(d1), _with_active_scope(d2)])
    db.session.commit()

    response = client.get(
        "/api/dashboard/series?"
        "exit_type=Strategic+Sale&lead_partner=Jane+Doe&security_type=Common+Equity"
        "&deal_type=Platform&entry_channel=Proprietary"
    )
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["kpis"]["total_deals"] == 1


def test_api_dashboard_series_scales_monetary_fields_when_fx_active(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None

    firm = Firm(
        name="Scaled API Firm",
        slug="scaled-api-firm",
        base_currency="EUR",
        fx_rate_to_usd=1.2,
        fx_rate_date=date(2026, 2, 20),
        fx_rate_source="Frankfurter (ECB)",
        fx_last_status="ok",
    )
    db.session.add(firm)
    db.session.flush()
    db.session.add(TeamFirmAccess(team_id=membership.team_id, firm_id=firm.id))
    db.session.add(
        Deal(
            company_name="Scaled API Deal",
            fund_number="Fund API FX",
            team_id=membership.team_id,
            firm_id=firm.id,
            equity_invested=100,
            realized_value=150,
            unrealized_value=0,
        )
    )
    db.session.commit()

    with client.session_transaction() as sess:
        sess["active_firm_id"] = firm.id

    response = client.get("/api/dashboard/series")
    assert response.status_code == 200
    payload = response.get_json()
    # Values are already in USD from upload-time conversion; no display-time
    # scaling is applied, so stored values pass through unchanged.
    assert abs(payload["kpis"]["total_equity"] - 100.0) < 1e-9
    assert abs(payload["kpis"]["total_value"] - 150.0) < 1e-9


def test_api_dashboard_series_schema(client):
    deal = Deal(
        company_name="API Co",
        fund_number="Fund I",
        geography="US",
        year_invested=2020,
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=50,
        exit_revenue=60,
        entry_ebitda=10,
        exit_ebitda=12,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
        investment_date=date(2020, 1, 1),
        exit_date=date(2023, 1, 1),
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/api/dashboard/series")
    assert response.status_code == 200
    payload = response.get_json()
    for key in (
        "kpis",
        "loss_ratios",
        "moic_distribution",
        "entry_exit_summary",
        "bridge_aggregate",
        "vintage_series",
        "moic_hold_scatter",
        "value_creation_mix",
        "realized_unrealized_exposure",
        "loss_concentration_heatmap",
        "exit_type_performance",
        "lead_partner_scorecard",
    ):
        assert key in payload
    mix = payload["value_creation_mix"]
    assert mix["current"] == "fund"
    assert set(mix["series"].keys()) == {"fund", "sector", "exit_type"}
    assert "fallback_ready_count" in payload["bridge_aggregate"]


def test_api_deal_bridge_query_params(client):
    deal = Deal(
        company_name="Bridge Co",
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=50,
        exit_revenue=60,
        entry_ebitda=10,
        exit_ebitda=12,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get(f"/api/deals/{deal.id}/bridge?unit=pct&basis=fund")
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["deal_id"] == deal.id
    assert payload["model"] == "additive"
    assert payload["unit"] == "pct"
    assert payload["basis"] == "fund"
    assert payload["start_value"] == 0.0
    assert payload["end_value"] == 1.0
    assert payload["calculation_method"] == "ebitda_additive"
    assert payload["fallback_reason"] is None


def test_api_deal_bridge_scales_dollar_fields_under_fx_conversion(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None

    firm = Firm(
        name="Bridge FX Firm",
        slug="bridge-fx-firm",
        base_currency="EUR",
        fx_rate_to_usd=1.2,
        fx_rate_date=date(2026, 2, 20),
        fx_rate_source="Frankfurter (ECB)",
        fx_last_status="ok",
    )
    db.session.add(firm)
    db.session.flush()
    db.session.add(TeamFirmAccess(team_id=membership.team_id, firm_id=firm.id))

    deal = Deal(
        company_name="Bridge FX Co",
        team_id=membership.team_id,
        firm_id=firm.id,
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=50,
        exit_revenue=60,
        entry_ebitda=10,
        exit_ebitda=12,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
    )
    db.session.add(deal)
    db.session.commit()

    with client.session_transaction() as sess:
        sess["active_firm_id"] = firm.id

    dollar_resp = client.get(f"/api/deals/{deal.id}/bridge?unit=dollar&basis=fund")
    assert dollar_resp.status_code == 200
    dollar_payload = dollar_resp.get_json()
    # Values are already in USD from upload-time conversion; no display-time
    # scaling is applied.
    assert abs(dollar_payload["start_dollar"] - 100.0) < 1e-9
    assert abs(dollar_payload["equity_invested"] - 100.0) < 1e-9

    moic_resp = client.get(f"/api/deals/{deal.id}/bridge?unit=moic&basis=fund")
    assert moic_resp.status_code == 200
    moic_payload = moic_resp.get_json()
    assert moic_payload["start_value"] == 1.0
    assert abs(moic_payload["end_value"] - 1.4) < 1e-9


def test_api_deal_bridge_negative_ebitda_uses_fallback_method(client):
    deal = Deal(
        company_name="Bridge Fallback Co",
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=50,
        exit_revenue=60,
        entry_ebitda=-10,
        exit_ebitda=-8,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get(f"/api/deals/{deal.id}/bridge?unit=moic&basis=fund")
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["ready"] is True
    assert payload["calculation_method"] == "revenue_multiple_fallback"
    assert payload["fallback_reason"] == "negative_ebitda"


def test_api_deal_bridge_missing_revenue_uses_ebitda_fallback_method(client):
    deal = Deal(
        company_name="Bridge Missing Revenue Co",
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=None,
        exit_revenue=None,
        entry_ebitda=10,
        exit_ebitda=12,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get(f"/api/deals/{deal.id}/bridge?unit=moic&basis=fund")
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["ready"] is True
    assert payload["calculation_method"] == "ebitda_multiple_fallback"
    assert payload["fallback_reason"] == "missing_revenue"
    display_keys = [row.get("key") for row in payload.get("display_drivers") or []]
    assert display_keys[:3] == ["ebitda_growth", "multiple", "leverage"]
    assert "margin" not in display_keys


def test_api_deal_bridge_rejects_non_additive_model(client):
    deal = Deal(
        company_name="Bridge Co 2",
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=50,
        exit_revenue=60,
        entry_ebitda=10,
        exit_ebitda=12,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get(f"/api/deals/{deal.id}/bridge?model=multiplicative&unit=moic&basis=fund")
    assert response.status_code == 400


def test_api_deal_bridge_rejects_company_basis(client):
    deal = Deal(
        company_name="Bridge Co 3",
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=50,
        exit_revenue=60,
        entry_ebitda=10,
        exit_ebitda=12,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get(f"/api/deals/{deal.id}/bridge?basis=company&unit=moic")
    assert response.status_code == 400
    assert response.get_json()["error"] == "Only fund pro-rata basis is supported"


def test_api_deal_bridge_firm_scope_returns_404_for_other_firm_deal(client):
    other_firm = Firm(name="Other Firm", slug="other-firm")
    db.session.add(other_firm)
    db.session.flush()

    foreign_deal = Deal(
        company_name="Foreign Bridge Co",
        fund_number="Fund Other",
        firm_id=other_firm.id,
        equity_invested=100,
        realized_value=130,
        unrealized_value=10,
        entry_revenue=50,
        exit_revenue=60,
        entry_ebitda=10,
        exit_ebitda=12,
        entry_enterprise_value=100,
        exit_enterprise_value=130,
        entry_net_debt=30,
        exit_net_debt=20,
    )
    db.session.add(foreign_deal)
    db.session.commit()

    response = client.get(f"/api/deals/{foreign_deal.id}/bridge?unit=moic&basis=fund")
    assert response.status_code == 404


def test_manage_firms_page_and_scope_switch(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None

    firm_a = Firm(name="Firm Scope A", slug="firm-scope-a")
    firm_b = Firm(name="Firm Scope B", slug="firm-scope-b")
    firm_hidden = Firm(name="Firm Hidden", slug="firm-hidden")
    db.session.add_all([firm_a, firm_b])
    db.session.flush()
    db.session.add(firm_hidden)
    db.session.flush()

    db.session.add_all(
        [
            TeamFirmAccess(team_id=membership.team_id, firm_id=firm_a.id),
            TeamFirmAccess(team_id=membership.team_id, firm_id=firm_b.id),
        ]
    )

    db.session.add(
        Deal(
            company_name="Firm Scope Deal B",
            fund_number="Fund B",
            firm_id=firm_b.id,
            equity_invested=75,
            realized_value=95,
            unrealized_value=0,
        )
    )
    db.session.commit()

    page = client.get("/firms")
    assert page.status_code == 200
    assert b"Manage Firms" in page.data
    assert b"Firm Scope B" in page.data
    assert b"Firm Hidden" not in page.data
    assert b"Currency" in page.data

    select = client.post(f"/firms/{firm_b.id}/select", follow_redirects=False)
    assert select.status_code == 302

    with client.session_transaction() as sess:
        assert sess.get("active_firm_id") == firm_b.id

    scoped = client.get("/deals")
    assert scoped.status_code == 200
    assert b"Firm Scope Deal B" in scoped.data

    denied = client.post(f"/firms/{firm_hidden.id}/select", follow_redirects=False)
    assert denied.status_code == 302


def test_rendered_pages_expose_active_currency_metadata(client):
    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b'data-currency-code="USD"' in response.data


def test_dashboard_conversion_banner_and_usd_reporting_metadata(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None

    firm = Firm(
        name="EUR Reporting Firm",
        slug="eur-reporting-firm",
        base_currency="EUR",
        fx_rate_to_usd=1.10,
        fx_rate_date=date(2026, 2, 20),
        fx_rate_source="Frankfurter (ECB)",
        fx_last_status="ok",
    )
    db.session.add(firm)
    db.session.flush()
    db.session.add(TeamFirmAccess(team_id=membership.team_id, firm_id=firm.id))
    db.session.add(
        Deal(
            company_name="EUR Deal",
            fund_number="Fund FX",
            team_id=membership.team_id,
            firm_id=firm.id,
            equity_invested=100,
            realized_value=150,
            unrealized_value=0,
        )
    )
    db.session.commit()

    with client.session_transaction() as sess:
        sess["active_firm_id"] = firm.id

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"Converted from EUR to USD at 1.100000" in response.data
    assert b'data-currency-code="USD"' in response.data


def test_dashboard_native_currency_warning_when_fx_unavailable(client):
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None

    firm = Firm(
        name="EUR Native Firm",
        slug="eur-native-firm",
        base_currency="EUR",
        fx_rate_to_usd=None,
        fx_rate_date=None,
        fx_rate_source="Frankfurter (ECB)",
        fx_last_status="lookup_failed",
    )
    db.session.add(firm)
    db.session.flush()
    db.session.add(TeamFirmAccess(team_id=membership.team_id, firm_id=firm.id))
    db.session.add(
        Deal(
            company_name="Native EUR Deal",
            fund_number="Fund Native",
            team_id=membership.team_id,
            firm_id=firm.id,
            equity_invested=80,
            realized_value=100,
            unrealized_value=0,
        )
    )
    db.session.commit()

    with client.session_transaction() as sess:
        sess["active_firm_id"] = firm.id

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"FX unavailable; showing native EUR values." in response.data
    assert b'data-currency-code="EUR"' in response.data


def test_analysis_pages_render(client):
    pages = {
        "fund-liquidity": b"Fund Liquidity &amp; Performance Curve",
        "underwrite-outcome": b"Underwrite vs Outcome",
        "valuation-quality": b"Unrealized Valuation Quality",
        "exit-readiness": b"Exit Readiness &amp; Aging",
        "stress-lab": b"Concentration Stress Lab",
        "deal-trajectory": b"Deal Trajectory",
        "nav-at-risk": b"NAV at Risk",
        "public-market-comparison": b"Public Market Comparison",
        "lp-due-diligence-memo": b"LP Due Diligence Memo",
        "vca-ebitda": b"Value Creation Analysis - by EBITDA",
        "vca-revenue": b"Value Creation Analysis - by Revenue",
        "vca-addons": b"Value Creation Analysis - with Add-Ons",
        "vca-addons-revenue": b"Value Creation Analysis - with Add-Ons by Revenue",
        "benchmarking": b"Benchmarking Analysis (IC PDF)",
    }
    for page, marker in pages.items():
        response = client.get(f"/analysis/{page}")
        assert response.status_code == 200
        assert marker in response.data
        if page == "stress-lab":
            assert b"Current EBITDA" in response.data
            assert b"Stressed Implied IRR" in response.data
            assert b"Hold Period (Yrs)" in response.data
            assert b"Expected Hold (Yrs)" in response.data
            assert b"Delay (Years)" not in response.data
            assert b"Print / Save PDF" in response.data
        if page in {"vca-ebitda", "vca-revenue", "vca-addons", "vca-addons-revenue", "benchmarking"}:
            assert b"Download / Print PDF" in response.data
            assert b"Preview Print Layout" in response.data


def test_removed_analysis_pages_return_404(client):
    for page in (
        "lp-liquidity-quality",
        "liquidity-forecast",
        "manager-consistency",
        "benchmark-confidence",
        "reporting-quality",
        "fee-drag",
    ):
        response = client.get(f"/analysis/{page}")
        assert response.status_code == 404


def test_analysis_api_series_schema(client):
    deal = Deal(
        company_name="Analysis API Co",
        fund_number="Fund IX",
        status="Unrealized",
        investment_date=date(2020, 1, 1),
        equity_invested=100,
        realized_value=20,
        unrealized_value=130,
        entry_revenue=60,
        entry_ebitda=12,
        entry_enterprise_value=140,
        entry_net_debt=30,
        exit_revenue=90,
        exit_ebitda=20,
        exit_enterprise_value=220,
        exit_net_debt=25,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    for page in (
        "fund-liquidity",
        "underwrite-outcome",
        "valuation-quality",
        "exit-readiness",
        "stress-lab",
        "deal-trajectory",
        "nav-at-risk",
        "public-market-comparison",
        "lp-due-diligence-memo",
        "vca-ebitda",
        "vca-revenue",
        "vca-addons",
        "vca-addons-revenue",
        "benchmarking",
    ):
        response = client.get(f"/api/analysis/{page}/series")
        assert response.status_code == 200
        payload = response.get_json()
        assert payload["page"] == page
        assert "payload" in payload


def test_removed_analysis_api_series_return_404(client):
    for page in (
        "lp-liquidity-quality",
        "liquidity-forecast",
        "manager-consistency",
        "benchmark-confidence",
        "reporting-quality",
        "fee-drag",
    ):
        response = client.get(f"/api/analysis/{page}/series")
        assert response.status_code == 404


def test_analysis_vca_ebitda_api_payload_shape(client):
    deal = Deal(
        company_name="VCA API Co",
        fund_number="Fund VCA",
        status="Fully Realized",
        investment_date=date(2020, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=170,
        unrealized_value=0,
        entry_revenue=50,
        exit_revenue=80,
        entry_ebitda=10,
        exit_ebitda=16,
        entry_enterprise_value=120,
        exit_enterprise_value=210,
        entry_net_debt=35,
        exit_net_debt=20,
        irr=0.21,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/api/analysis/vca-ebitda/series")
    assert response.status_code == 200
    body = response.get_json()
    payload = body["payload"]
    assert body["page"] == "vca-ebitda"
    assert "meta" in payload
    assert "header" in payload
    assert "fund_blocks" in payload
    assert "overall_block" in payload
    assert "net_performance" in payload["fund_blocks"][0]
    assert "print_sort_metrics" in payload["fund_blocks"][0]
    assert "summary_metrics" in payload["overall_block"]


def test_analysis_vca_revenue_api_payload_shape(client):
    deal = Deal(
        company_name="VCA Revenue API Co",
        fund_number="Fund VCA Rev",
        status="Fully Realized",
        investment_date=date(2020, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=170,
        unrealized_value=0,
        entry_revenue=50,
        exit_revenue=80,
        entry_ebitda=10,
        exit_ebitda=16,
        entry_enterprise_value=120,
        exit_enterprise_value=210,
        entry_net_debt=35,
        exit_net_debt=20,
        irr=0.21,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/api/analysis/vca-revenue/series")
    assert response.status_code == 200
    body = response.get_json()
    payload = body["payload"]
    assert body["page"] == "vca-revenue"
    assert "meta" in payload
    assert "header" in payload
    assert "fund_blocks" in payload
    assert "overall_block" in payload
    assert "net_performance" in payload["fund_blocks"][0]
    assert any(col.get("key") == "revenue_cagr" for col in payload["header"]["columns"])
    assert any(col.get("key") == "entry_ev_revenue" for col in payload["header"]["columns"])
    assert not any(col.get("key") == "entry_ev_ebitda" for col in payload["header"]["columns"])


def test_analysis_vca_addons_api_payload_shape(client):
    deal = Deal(
        company_name="VCA Add-On API Co",
        fund_number="Fund VCA Add",
        status="Fully Realized",
        investment_date=date(2020, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=190,
        unrealized_value=0,
        entry_revenue=50,
        exit_revenue=95,
        entry_ebitda=10,
        exit_ebitda=20,
        entry_enterprise_value=120,
        exit_enterprise_value=240,
        entry_net_debt=35,
        exit_net_debt=20,
        acquired_revenue=15,
        acquired_ebitda=3,
        acquired_tev=45,
        irr=0.22,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/api/analysis/vca-addons/series")
    assert response.status_code == 200
    body = response.get_json()
    payload = body["payload"]
    assert body["page"] == "vca-addons"
    assert "meta" in payload
    assert "header" in payload
    assert "fund_blocks" in payload
    assert "overall_block" in payload
    row = payload["fund_blocks"][0]["deal_rows"][0]
    assert row["acquired_revenue"] == 15
    assert row["acquired_ebitda"] == 3
    assert row["acquired_tev"] == 45
    assert row["vc_add_on_ebitda_dollar"] is not None
    assert "print_sort_metrics" in payload["fund_blocks"][0]
    assert "summary_metrics" in payload["overall_block"]
    assert [group["span"] for group in payload["header"]["groups"]] == [9, 10, 2, 5, 5, 6, 2, 1, 6, 6]
    assert [col["key"] for col in payload["header"]["columns"]] == [
        "row_num",
        "platform",
        "sector",
        "geography",
        "exit_type",
        "close_date",
        "final_exit_date",
        "hold_period",
        "status",
        "fund_initial_cost",
        "fund_total_cost",
        "realized_proceeds",
        "unrealized_value",
        "total_value",
        "gross_profit",
        "gross_profit_pct_of_total",
        "gross_irr",
        "realized_moic",
        "gross_moic",
        "organic_ebitda_cagr",
        "organic_ebitda_cumulative_growth",
        "vc_organic_ebitda_growth_pct",
        "vc_add_on_ebitda_pct",
        "vc_multiple_pct",
        "vc_debt_pct",
        "vc_total_pct",
        "vc_organic_ebitda_growth_dollar",
        "vc_add_on_ebitda_dollar",
        "vc_multiple_dollar",
        "vc_debt_dollar",
        "vc_total_dollar",
        "entry_ltm_ebitda",
        "entry_ebitda_margin",
        "entry_ev_ebitda",
        "entry_net_debt",
        "entry_net_debt_ebitda",
        "entry_net_debt_ev",
        "acquired_ebitda",
        "acquired_ev_ebitda",
        "blended_ev_ebitda_with_addons",
        "exit_ltm_ebitda",
        "exit_ebitda_margin",
        "exit_ev_ebitda",
        "exit_net_debt",
        "exit_net_debt_ebitda",
        "exit_net_debt_ev",
        "diff_ebitda",
        "diff_ebitda_margin",
        "diff_ev_ebitda",
        "diff_net_debt",
        "diff_net_debt_ebitda",
        "diff_net_debt_ev",
    ]


def test_analysis_vca_addons_revenue_api_payload_shape(client):
    deal = Deal(
        company_name="VCA Add-On Revenue API Co",
        fund_number="Fund VCA Add Rev",
        status="Fully Realized",
        investment_date=date(2020, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=190,
        unrealized_value=0,
        entry_revenue=50,
        exit_revenue=95,
        entry_ebitda=10,
        exit_ebitda=20,
        entry_enterprise_value=150,
        exit_enterprise_value=260,
        entry_net_debt=35,
        exit_net_debt=20,
        acquired_revenue=15,
        acquired_ebitda=3,
        acquired_tev=45,
        irr=0.22,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/api/analysis/vca-addons-revenue/series")
    assert response.status_code == 200
    body = response.get_json()
    payload = body["payload"]
    assert body["page"] == "vca-addons-revenue"
    assert "meta" in payload
    assert "header" in payload
    assert "fund_blocks" in payload
    assert "overall_block" in payload
    row = payload["fund_blocks"][0]["deal_rows"][0]
    assert row["acquired_revenue"] == 15
    assert row["acquired_ev_revenue"] == 3
    assert row["vc_add_on_revenue_dollar"] is not None
    assert "print_sort_metrics" in payload["fund_blocks"][0]
    assert "summary_metrics" in payload["overall_block"]
    assert [group["span"] for group in payload["header"]["groups"]] == [9, 10, 2, 5, 5, 5, 2, 1, 5, 5]
    assert [col["key"] for col in payload["header"]["columns"]] == [
        "row_num",
        "platform",
        "sector",
        "geography",
        "exit_type",
        "close_date",
        "final_exit_date",
        "hold_period",
        "status",
        "fund_initial_cost",
        "fund_total_cost",
        "realized_proceeds",
        "unrealized_value",
        "total_value",
        "gross_profit",
        "gross_profit_pct_of_total",
        "gross_irr",
        "realized_moic",
        "gross_moic",
        "organic_revenue_cagr",
        "organic_revenue_cumulative_growth",
        "vc_organic_revenue_growth_pct",
        "vc_add_on_revenue_pct",
        "vc_multiple_pct",
        "vc_debt_pct",
        "vc_total_pct",
        "vc_organic_revenue_growth_dollar",
        "vc_add_on_revenue_dollar",
        "vc_multiple_dollar",
        "vc_debt_dollar",
        "vc_total_dollar",
        "entry_ltm_revenue",
        "entry_ev_revenue",
        "entry_net_debt",
        "entry_net_debt_revenue",
        "entry_net_debt_ev",
        "acquired_revenue",
        "acquired_ev_revenue",
        "blended_ev_revenue_with_addons",
        "exit_ltm_revenue",
        "exit_ev_revenue",
        "exit_net_debt",
        "exit_net_debt_revenue",
        "exit_net_debt_ev",
        "diff_revenue",
        "diff_ev_revenue",
        "diff_net_debt",
        "diff_net_debt_revenue",
        "diff_net_debt_ev",
    ]


def test_analysis_benchmarking_api_payload_shape(client):
    deal = Deal(
        company_name="Benchmark API Co",
        fund_number="Fund Bench API",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=170,
        unrealized_value=0,
        net_irr=0.22,
        net_moic=2.1,
        net_dpi=1.4,
        fund_size=450,
    )
    db.session.add(_with_active_scope(deal))
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    db.session.add_all(
        [
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="lower_quartile", value=0.12),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="median", value=0.17),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="upper_quartile", value=0.21),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="top_5", value=0.3),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="lower_quartile", value=1.5),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="median", value=1.9),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="upper_quartile", value=2.2),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="top_5", value=2.8),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="lower_quartile", value=0.6),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="median", value=1.3),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="upper_quartile", value=1.5),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="top_5", value=1.8),
        ]
    )
    db.session.commit()

    response = client.get("/api/analysis/benchmarking/series?benchmark_asset_class=Buyout")
    assert response.status_code == 200
    body = response.get_json()
    payload = body["payload"]
    assert body["page"] == "benchmarking"
    assert "meta" in payload
    assert "kpis" in payload
    assert "rank_distribution" in payload
    assert "fund_rows" in payload
    assert "threshold_rows" in payload
    assert payload["meta"]["benchmark_asset_class"] == "Buyout"
    assert payload["fund_rows"][0]["benchmark_net_irr"]["rank_code"] in {"top5", "q1", "q2", "q3", "q4", "na"}


def test_analysis_benchmarking_page_renders_ic_print_markers(client):
    deal = Deal(
        company_name="Benchmark Page Co",
        fund_number="Fund Bench Page",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=160,
        unrealized_value=0,
        net_irr=0.2,
        net_moic=1.9,
        net_dpi=1.2,
    )
    db.session.add(_with_active_scope(deal))
    membership = TeamMembership.query.order_by(TeamMembership.id.asc()).first()
    assert membership is not None
    db.session.add_all(
        [
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="lower_quartile", value=0.12),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="median", value=0.16),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_irr", quartile="upper_quartile", value=0.2),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="lower_quartile", value=1.4),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="median", value=1.7),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_moic", quartile="upper_quartile", value=2.0),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="lower_quartile", value=0.7),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="median", value=1.0),
            BenchmarkPoint(team_id=membership.team_id, asset_class="Buyout", vintage_year=2019, metric="net_dpi", quartile="upper_quartile", value=1.3),
        ]
    )
    db.session.commit()

    response = client.get("/analysis/benchmarking?benchmark_asset_class=Buyout")
    assert response.status_code == 200
    assert b"Benchmarking Analysis (IC PDF)" in response.data
    assert b"Benchmark Asset Class" in response.data
    assert b"Quartile Rank Distribution" not in response.data
    assert b"Fund Benchmarking Table" in response.data
    assert b"Benchmark Threshold Appendix" in response.data
    assert b"Vintage Year" in response.data
    assert b"Net IRR Benchmark" in response.data
    assert b"Net MOIC Benchmark" in response.data
    assert b"Net DPI Benchmark" in response.data
    assert b"vs Median" not in response.data
    assert b"<th>Composite</th>" not in response.data
    assert b'<th class="num">Score</th>' not in response.data
    assert b'id="benchPreviewToggle"' in response.data
    assert b"bench-print-layout" in response.data
    assert b"bench-print-page" in response.data
    funds_idx = response.data.find(b"bench-print-page bench-print-page-funds")
    summary_idx = response.data.find(b"bench-print-page bench-print-page-summary")
    assert funds_idx != -1
    assert summary_idx != -1
    assert funds_idx < summary_idx


def test_analysis_vca_ebitda_page_renders_group_headers_with_data(client):
    deal = Deal(
        company_name="VCA Page Co",
        fund_number="Fund VCA",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=160,
        unrealized_value=0,
        entry_revenue=55,
        exit_revenue=90,
        entry_ebitda=11,
        exit_ebitda=17,
        entry_enterprise_value=130,
        exit_enterprise_value=220,
        entry_net_debt=40,
        exit_net_debt=25,
        irr=0.2,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/analysis/vca-ebitda")
    assert response.status_code == 200
    assert b"EBITDA Growth During Hold Period" in response.data
    assert b"Value Creation (%)" in response.data
    assert b"Value Creation ($)" in response.data
    assert b"Difference Exit/Current vs Entry" in response.data
    assert b"vca-print-layout" in response.data
    assert b"vca-print-book" in response.data
    assert response.data.count(b"vca-print-fund-page") >= 1
    assert b"vca-print-fund-page" in response.data
    assert b"vca-print-overall-page" in response.data
    assert b'id="vcaPrintSort"' in response.data
    assert b'id="vcaDensityToggle"' in response.data
    assert b'id="vcaModeToggle"' in response.data
    assert b"vca-print-exec-main" in response.data
    assert b"vca-print-appendix-title" in response.data
    assert b"WIDTH_SAFETY_PX" in response.data
    assert b"HEIGHT_SAFETY_PX" in response.data
    assert b"FUND_BLOCK_BUFFER_PX" in response.data
    assert b"FOOTER_GUARD_BAND_PX" in response.data
    assert b"PAGE_MARGIN_IN = 0.22" in response.data
    assert b"vca-net-summary" not in response.data
    assert re.search(rb"Fund\s+\d+\s+of\s+\d+", response.data) is None
    assert b"Overall Portfolio (Final Block)" not in response.data
    assert b"Net IRR" not in response.data
    assert b"Net MOIC" not in response.data
    assert b"Net DPI" not in response.data
    assert b"$M" in response.data
    assert b"| $M" in response.data
    assert b"| USD $M" not in response.data


def test_analysis_vca_revenue_page_renders_group_headers_with_data(client):
    deal = Deal(
        company_name="VCA Revenue Page Co",
        fund_number="Fund VCA Rev",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=160,
        unrealized_value=0,
        entry_revenue=55,
        exit_revenue=90,
        entry_ebitda=11,
        exit_ebitda=17,
        entry_enterprise_value=130,
        exit_enterprise_value=220,
        entry_net_debt=40,
        exit_net_debt=25,
        irr=0.2,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/analysis/vca-revenue")
    assert response.status_code == 200
    assert b"Revenue Growth During Hold Period" in response.data
    assert b"Value Creation (%)" in response.data
    assert b"Value Creation ($)" in response.data
    assert b"Difference Exit/Current vs Entry" in response.data
    assert b"EBITDA Growth During Hold Period" not in response.data
    assert b"vca-print-layout" in response.data
    assert b"vca-print-book" in response.data
    assert response.data.count(b"vca-print-fund-page") >= 1
    assert b"vca-print-overall-page" in response.data
    assert b'id="vcaPrintSort"' in response.data
    assert b'id="vcaDensityToggle"' in response.data
    assert b'id="vcaModeToggle"' in response.data
    assert b"vca-print-exec-main" in response.data
    assert b"WIDTH_SAFETY_PX" in response.data
    assert b"HEIGHT_SAFETY_PX" in response.data
    assert b"FUND_BLOCK_BUFFER_PX" in response.data
    assert b"FOOTER_GUARD_BAND_PX" in response.data
    assert b"PAGE_MARGIN_IN = 0.22" in response.data
    assert b"vca-net-summary" not in response.data
    assert re.search(rb"Fund\s+\d+\s+of\s+\d+", response.data) is None
    assert b"Overall Portfolio (Final Block)" not in response.data
    assert b"Net IRR" not in response.data
    assert b"Net MOIC" not in response.data
    assert b"Net DPI" not in response.data
    assert b"| $M" in response.data
    assert b"| USD $M" not in response.data


def test_analysis_vca_addons_page_renders_group_headers_with_data(client):
    deal = Deal(
        company_name="VCA Add-On Page Co",
        fund_number="Fund VCA Add",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=175,
        unrealized_value=0,
        entry_revenue=55,
        exit_revenue=100,
        entry_ebitda=11,
        exit_ebitda=22,
        entry_enterprise_value=132,
        exit_enterprise_value=260,
        entry_net_debt=40,
        exit_net_debt=25,
        acquired_revenue=18,
        acquired_ebitda=4,
        acquired_tev=55,
        irr=0.2,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/analysis/vca-addons")
    assert response.status_code == 200
    assert b"Organic EBITDA Growth During Hold Period" in response.data
    assert b"Add-Ons" in response.data
    assert b"Acquired EBITDA" in response.data
    assert b"Acquired EV/EBITDA" in response.data
    assert b"Blended EV/EBITDA With Add-Ons" in response.data
    assert b"Add-On EBITDA" in response.data
    assert b"Value Creation (%)" in response.data
    assert b"Value Creation ($)" in response.data
    assert b"Difference Exit/Current vs Entry" in response.data
    assert b"vca-print-layout" in response.data
    assert b"vca-print-book" in response.data
    assert response.data.count(b"vca-print-fund-page") >= 1
    assert b"vca-print-overall-page" in response.data
    assert b'id="vcaPrintSort"' in response.data
    assert b'id="vcaDensityToggle"' in response.data
    assert b'id="vcaModeToggle"' in response.data
    assert b"| $M" in response.data
    assert b"| USD $M" not in response.data


def test_analysis_vca_addons_revenue_page_renders_group_headers_with_data(client):
    deal = Deal(
        company_name="VCA Add-On Revenue Page Co",
        fund_number="Fund VCA Add Rev",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=175,
        unrealized_value=0,
        entry_revenue=55,
        exit_revenue=105,
        entry_ebitda=11,
        exit_ebitda=22,
        entry_enterprise_value=165,
        exit_enterprise_value=275,
        entry_net_debt=40,
        exit_net_debt=25,
        acquired_revenue=18,
        acquired_ebitda=4,
        acquired_tev=54,
        irr=0.2,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/analysis/vca-addons-revenue")
    assert response.status_code == 200
    assert b"Organic Revenue Growth During Hold Period" in response.data
    assert b"Add-Ons" in response.data
    assert b"Acquired Revenue" in response.data
    assert b"Acquired EV/Revenue" in response.data
    assert b"Blended EV/Revenue With Add-Ons" in response.data
    assert b"Revenue Growth Through Add-Ons" in response.data
    assert b"Value Creation (%)" in response.data
    assert b"Value Creation ($)" in response.data
    assert b"Difference Exit/Current vs Entry" in response.data
    assert b"vca-print-layout" in response.data
    assert b"vca-print-book" in response.data
    assert response.data.count(b"vca-print-fund-page") >= 1
    assert b"vca-print-overall-page" in response.data
    assert b'id="vcaPrintSort"' in response.data
    assert b'id="vcaDensityToggle"' in response.data
    assert b'id="vcaModeToggle"' in response.data
    assert b"| $M" in response.data
    assert b"| USD $M" not in response.data


def test_analysis_vca_ebitda_net_conflict_omits_net_print_rows(client):
    d1 = Deal(
        company_name="Conflict One",
        fund_number="Fund Conflict",
        status="Fully Realized",
        investment_date=date(2019, 1, 1),
        exit_date=date(2024, 1, 1),
        equity_invested=100,
        realized_value=150,
        unrealized_value=0,
        entry_revenue=50,
        exit_revenue=78,
        entry_ebitda=10,
        exit_ebitda=15,
        entry_enterprise_value=120,
        exit_enterprise_value=200,
        entry_net_debt=30,
        exit_net_debt=20,
        irr=0.19,
        net_irr=0.12,
        net_moic=1.8,
        net_dpi=0.9,
    )
    d2 = Deal(
        company_name="Conflict Two",
        fund_number="Fund Conflict",
        status="Partially Realized",
        investment_date=date(2020, 1, 1),
        exit_date=date(2024, 6, 1),
        equity_invested=80,
        realized_value=90,
        unrealized_value=20,
        entry_revenue=45,
        exit_revenue=67,
        entry_ebitda=9,
        exit_ebitda=13,
        entry_enterprise_value=108,
        exit_enterprise_value=172,
        entry_net_debt=26,
        exit_net_debt=18,
        irr=0.17,
        net_irr=0.15,
        net_moic=1.8,
        net_dpi=0.9,
    )
    db.session.add_all([_with_active_scope(d1), _with_active_scope(d2)])
    db.session.commit()

    response = client.get("/analysis/vca-ebitda")
    assert response.status_code == 200
    assert b"Fund Conflict Net Performance" not in response.data
    assert b"vca-net-summary" not in response.data
    assert b"Net IRR" not in response.data
    assert b"Net MOIC" not in response.data
    assert b"Net DPI" not in response.data


def test_analysis_non_vca_page_uses_symbol_only_money(client):
    deal = Deal(
        company_name="Stress Symbol Co",
        fund_number="Fund Symbol",
        status="Unrealized",
        investment_date=date(2020, 1, 1),
        equity_invested=100,
        realized_value=0,
        unrealized_value=130,
        entry_revenue=60,
        entry_ebitda=12,
        entry_enterprise_value=140,
        entry_net_debt=30,
        exit_revenue=90,
        exit_ebitda=20,
        exit_enterprise_value=220,
        exit_net_debt=25,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get("/analysis/stress-lab")
    assert response.status_code == 200
    assert re.search(rb"\$\d+\.\dM", response.data) is not None
    assert b"Current Value</span><strong>$" in response.data


def test_stress_lab_api_supports_per_deal_overrides(client):
    deal = Deal(
        company_name="Stress Override Co",
        fund_number="Fund IX",
        status="Unrealized",
        investment_date=date(2020, 1, 1),
        equity_invested=100,
        realized_value=0,
        unrealized_value=130,
        entry_revenue=60,
        entry_ebitda=12,
        entry_enterprise_value=140,
        entry_net_debt=30,
        exit_revenue=90,
        exit_ebitda=20,
        exit_enterprise_value=220,
        exit_net_debt=25,
    )
    db.session.add(_with_active_scope(deal))
    db.session.commit()

    response = client.get(
        f"/api/analysis/stress-lab/series?ms_{deal.id}=-2.0&es_{deal.id}=-20&hp_{deal.id}=7.5"
    )
    assert response.status_code == 200
    payload = response.get_json()["payload"]
    assert payload["deal_rows"]
    row = payload["deal_rows"][0]
    assert abs(row["multiple_shock"] - (-2.0)) < 1e-9
    assert abs(row["ebitda_shock"] - (-0.20)) < 1e-9
    assert abs(row["expected_hold_period"] - 7.5) < 1e-9


def test_protected_routes_redirect_when_not_logged_in(anonymous_client):
    response = anonymous_client.get("/dashboard")
    assert response.status_code == 302
    assert "/auth/login" in response.location


def test_manage_funds_routes_are_deprecated(client):
    page = client.get("/funds", follow_redirects=False)
    assert page.status_code == 302
    assert "/firms" in page.location

    select = client.post("/funds/Fund%20One/select", follow_redirects=False)
    assert select.status_code == 302

    delete_resp = client.post("/funds/Fund%20One/delete", follow_redirects=False)
    assert delete_resp.status_code == 410
