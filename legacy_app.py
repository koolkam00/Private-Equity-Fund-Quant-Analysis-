import logging
import os
import re
import secrets
import hashlib
import json
from io import BytesIO
from pathlib import Path
from datetime import date, datetime, timedelta, timezone
from zipfile import ZIP_DEFLATED, ZipFile

import click
from flask import abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask.cli import with_appcontext
from flask_login import current_user, login_required, login_user, logout_user
from flask_migrate import stamp as migrate_stamp
from flask_migrate import upgrade as migrate_upgrade
from sqlalchemy import inspect as sa_inspect, or_, text
from sqlalchemy.exc import SQLAlchemyError
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from models import (
    BenchmarkPoint,
    ChartBuilderTemplate,
    CreditFundPerformance,
    CreditLoan,
    CreditLoanSnapshot,
    Deal,
    DealCashflowEvent,
    DealQuarterSnapshot,
    DealUnderwriteBaseline,
    Firm,
    FundCashflow,
    FundMetadata,
    FundQuarterSnapshot,
    MemoDocument,
    MemoGenerationClaim,
    MemoGenerationRun,
    MemoGenerationSection,
    MemoJob,
    MemoStyleExemplar,
    MemoStyleProfile,
    PublicMarketIndexLevel,
    Team,
    TeamFirmAccess,
    TeamInvite,
    TeamMembership,
    UploadIssue,
    User,
    db,
    ensure_schema_updates,
)
from peqa.services.memos import assemble_memo, enqueue_job, rebuild_style_profile
from peqa.services.memos.evidence_builder import build_memo_evidence_bundle
from peqa.services.memos.jobs import cancel_jobs_for_run, recover_stale_jobs, run_inline_job
from peqa.services.memos.retrieval import retrieve_section_evidence
from peqa.services.memos.storage import build_storage_key, get_document_storage
from peqa.services.memos.style_profiles import list_style_exemplars, load_style_profile
from peqa.services.memos.types import DraftSection, dataclass_to_dict
from peqa.services.memos.validation import extract_claims, validate_section
from peqa.services.memos.worker import run_memo_worker
from services.benchmark_parser import parse_benchmarks
from services.deal_parser import parse_deals
from services.fx_rates import resolve_rate_to_usd
from services.metrics import (
    build_chart_field_catalog,
    build_methodology_payload,
    compute_benchmarking_analysis,
    compute_benchmark_confidence_analysis,
    compute_bridge_aggregate,
    compute_bridge_view,
    compute_data_cuts_analytics,
    compute_data_quality,
    compute_deals_rollup_details,
    compute_deal_trajectory_analysis,
    compute_deal_metrics,
    compute_deal_track_record,
    compute_fund_performance_comparison,
    compute_executive_summary_analysis,
    compute_exit_readiness_analysis,
    compute_exit_type_performance,
    compute_fee_drag_analysis,
    compute_fund_liquidity_analysis,
    compute_ic_memo_payload,
    compute_liquidity_forecast_analysis,
    compute_lead_partner_scorecard,
    compute_loss_and_distribution,
    compute_loss_concentration_heatmap,
    compute_moic_hold_scatter,
    compute_lp_due_diligence_memo,
    compute_lp_liquidity_quality_analysis,
    compute_manager_consistency_analysis,
    compute_nav_at_risk_analysis,
    compute_organic_growth_analysis,
    compute_portfolio_analytics,
    compute_public_market_comparison_analysis,
    compute_reporting_quality_analysis,
    compute_realized_unrealized_exposure,
    compute_stress_lab_analysis,
    compute_underwrite_outcome_analysis,
    compute_value_creation_mix,
    compute_valuation_quality_analysis,
    compute_vca_addons_analysis,
    compute_vca_addons_revenue_analysis,
    compute_vca_ebitda_analysis,
    compute_vca_revenue_analysis,
    compute_vintage_series,
    rank_benchmark_metric,
    run_chart_query,
)
from services.metrics.credit import (
    CREDIT_DIMENSIONS,
    CREDIT_DIMENSION_LABELS,
    credit_data_cuts_available_dimension_keys,
    compute_credit_benchmarking_analysis,
    compute_credit_concentration,
    compute_credit_data_cuts,
    compute_credit_fundamentals,
    compute_credit_loan_metrics,
    compute_credit_loan_structure,
    compute_credit_maturity_profile,
    compute_credit_migration_matrix,
    compute_credit_portfolio_analytics,
    compute_credit_pricing_trends,
    compute_credit_risk_metrics,
    compute_credit_stress_scenarios,
    compute_credit_track_record,
    compute_credit_underwrite_outcome,
    compute_credit_vintage_comparison,
    compute_credit_watchlist,
    compute_credit_yield_attribution,
)
from services.metrics.common import resolve_analysis_as_of_date
from services.utils import (
    DEFAULT_CURRENCY_CODE,
    currency_symbol,
    currency_unit_label,
    format_currency_millions,
    normalize_currency_code,
)
from peqa.services.context import (
    benchmark_asset_classes_for_team as context_benchmark_asset_classes_for_team,
    build_analysis_context,
    load_team_benchmark_thresholds,
)
from peqa.extensions import limiter, login_manager
from peqa.route_binding import AppBinder
from peqa.services.filtering import build_deal_scope_query, build_fund_vintage_lookup, deal_vintage_year, parse_request_filters

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger(__name__)

app = AppBinder()

DEAL_TEMPLATE_FILENAME = "PE_Fund_Data_Template.xlsx"
BENCHMARK_TEMPLATE_FILENAME = "PE_Benchmark_Template.xlsx"
ALEMBIC_BASELINE_REVISION = "4a62775748c8"
SCHEMA_UPGRADE_COMMAND = "python -m flask --app app db-upgrade"
REQUIRED_SCHEMA_TABLES = (
    "alembic_version",
    "users",
    "teams",
    "team_memberships",
    "team_firm_access",
    "team_invites",
    "firms",
    "deals",
    "upload_issues",
    "benchmark_points",
    "chart_builder_templates",
    "fund_metadata",
    "fund_cashflows",
    "public_market_index_levels",
)

ANALYSIS_PAGES = {
    "executive-summary": {
        "title": "Executive Summary",
        "description": "Comprehensive one-page overview combining key metrics, performance attribution, deal rankings, concentration analysis, and portfolio health.",
    },
    "fund-liquidity": {
        "title": "Fund Liquidity & Performance Curve",
        "description": "Quarterly paid-in, distributed, NAV, unfunded, and TVPI/DPI/RVPI/PIC trends.",
    },
    "underwrite-outcome": {
        "title": "Underwrite vs Outcome",
        "description": "Planned vs realized returns, hold periods, and driver deltas by strategy segment.",
    },
    "valuation-quality": {
        "title": "Unrealized Valuation Quality",
        "description": "Mark staleness, volatility, markdown concentration, and pre-exit mark backtests.",
    },
    "exit-readiness": {
        "title": "Exit Readiness & Aging",
        "description": "Hold-age diagnostics, thesis completion, and time-above-target signals for unrealized deals.",
    },
    "stress-lab": {
        "title": "Concentration Stress Lab",
        "description": "Scenario stress on multiples, EBITDA, and timing to identify downside concentration.",
    },
    "deal-trajectory": {
        "title": "Deal Trajectory",
        "description": "Quarterly EV/net debt/equity and cash-flow path for an individual deal.",
    },
    "vca-ebitda": {
        "title": "Value Creation Analysis - by EBITDA",
        "description": "PDF-style value creation table with fund blocks, subtotal rollups, and operating deltas.",
    },
    "vca-revenue": {
        "title": "Value Creation Analysis - by Revenue",
        "description": "PDF-style value creation table with fund blocks, subtotal rollups, and operating deltas.",
    },
    "vca-addons": {
        "title": "Value Creation Analysis - with Add-Ons",
        "description": "PDF-style value creation table with fund blocks, add-on operating data, and organic versus add-on EBITDA attribution.",
    },
    "vca-addons-revenue": {
        "title": "Value Creation Analysis - with Add-Ons by Revenue",
        "description": "PDF-style value creation table with fund blocks, add-on revenue data, and organic versus add-on revenue attribution.",
    },
    "benchmarking": {
        "title": "Benchmarking Analysis (IC PDF)",
        "description": "IC-focused benchmark quartile analysis by fund with print-ready executive summaries.",
    },
    "nav-at-risk": {
        "title": "NAV at Risk",
        "description": "Tail-risk concentration across aged unrealized positions, stale marks, and below-plan holdings.",
    },
    "public-market-comparison": {
        "title": "Public Market Comparison",
        "description": "KS PME and Direct Alpha using uploaded fund cash flows and benchmark index series.",
    },
    "lp-due-diligence-memo": {
        "title": "LP Due Diligence Memo",
        "description": "Single-page LP diligence summary combining metadata, current fund liquidity, NAV concentration, and public market coverage.",
    },
    "chart-builder": {
        "title": "Chart Builder",
        "description": "Drag/drop fields, build ad-hoc visuals, and save team templates.",
    },
    "fund-comparison": {
        "title": "Fund Performance Comparison",
        "description": "Cross-firm comparison of net fund performance (IRR, TVPI, DPI) organized by vintage year.",
    },
    "deal-comparison": {
        "title": "Deal Performance Comparison",
        "description": "Cross-firm deal-level comparison of gross returns, hold periods, and sector/geography patterns.",
    },
    "organic-growth": {
        "title": "Organic vs Acquired Growth",
        "description": "Decomposition of revenue and EBITDA growth into organic vs bolt-on acquisition contributions with bridge integration.",
    },
    "data-cuts": {
        "title": "Data Cuts",
        "description": "Slice portfolio performance by any qualitative dimension with grouped metrics, charts, and drill-down.",
    },
}

WORKFLOW_SIDEBAR_ITEMS = [
    {"endpoint": "dashboard", "label": "Dashboard", "icon": "bi bi-grid-1x2"},
    {"endpoint": "analysis_page", "page_key": "benchmarking", "label": "Benchmarking Analysis", "icon": "bi bi-bar-chart-steps"},
    {"endpoint": "track_record", "label": "Track Record", "icon": "bi bi-table"},
    {"endpoint": "deals_analysis", "label": "Deal Selector", "icon": "bi bi-ui-checks-grid"},
    {"endpoint": "analysis_page", "page_key": "vca-ebitda", "label": "Value Creation (EBITDA)", "icon": "bi bi-table"},
    {"endpoint": "analysis_page", "page_key": "vca-revenue", "label": "Value Creation (Revenue)", "icon": "bi bi-table"},
    {"endpoint": "analysis_page", "page_key": "vca-addons", "label": "Value Creation (Add-Ons)", "icon": "bi bi-table"},
    {"endpoint": "analysis_page", "page_key": "vca-addons-revenue", "label": "Value Creation (Add-Ons Revenue)", "icon": "bi bi-table"},
    {"endpoint": "ic_memo", "label": "IC Memo", "icon": "bi bi-file-earmark-richtext"},
    {"endpoint": "deals", "label": "Deals", "icon": "bi bi-buildings"},
    {"endpoint": "live_ic_pdf_pack", "label": "Download 4 Analysis PDFs", "icon": "bi bi-file-earmark-zip"},
    {"endpoint": "upload", "label": "Upload Deals", "icon": "bi bi-cloud-upload"},
    {
        "endpoint": "memo_studio",
        "label": "AI Investment Memos",
        "icon": "bi bi-file-earmark-text",
        "active_endpoints": ["memo_studio", "memo_style_library", "memo_source_library", "memo_run_page"],
    },
]

ANALYSIS_SIDEBAR_ITEMS = [
    {"page_key": "executive-summary", "label": "Executive Summary", "icon": "bi bi-clipboard-data"},
    {"page_key": "nav-at-risk", "label": "NAV at Risk", "icon": "bi bi-exclamation-diamond"},
    {"page_key": "public-market-comparison", "label": "Public Market Comparison", "icon": "bi bi-graph-up"},
    {"page_key": "fund-liquidity", "label": "Fund Liquidity", "icon": "bi bi-graph-up-arrow"},
    {"page_key": "underwrite-outcome", "label": "Underwrite vs Outcome", "icon": "bi bi-bullseye"},
    {"page_key": "valuation-quality", "label": "Valuation Quality", "icon": "bi bi-shield-check"},
    {"page_key": "exit-readiness", "label": "Exit Readiness", "icon": "bi bi-hourglass-split"},
    {"page_key": "stress-lab", "label": "Stress Lab", "icon": "bi bi-activity"},
    {"page_key": "deal-trajectory", "label": "Deal Trajectory", "icon": "bi bi-bezier2"},
    {"page_key": "chart-builder", "label": "Chart Builder", "icon": "bi bi-bar-chart-line"},
    {"page_key": "fund-comparison", "label": "Fund Comparison", "icon": "bi bi-arrow-left-right"},
    {"page_key": "lp-due-diligence-memo", "label": "LP Due Diligence Memo", "icon": "bi bi-file-earmark-text"},
    {"page_key": "organic-growth", "label": "Organic vs Acquired", "icon": "bi bi-diagram-3"},
    {"page_key": "data-cuts", "label": "Data Cuts", "icon": "bi bi-sliders"},
]

CREDIT_ANALYSIS_PAGES = {
    "credit-track-record": {
        "title": "Credit Track Record",
        "description": "Deal-level track record grouped by fund with gross loan metrics and net fund performance.",
    },
    "credit-benchmarking": {
        "title": "Credit Benchmarking Analysis",
        "description": "Benchmark uploaded credit fund net performance against quartile thresholds using the credit workbook fund data.",
    },
    "credit-concentration": {
        "title": "Credit Concentration",
        "description": "Sector, geography, sponsor, and security type breakdowns sized by total unrealized value.",
    },
    "credit-fundamentals": {
        "title": "Credit Fundamentals",
        "description": "Analyze entry vs exit/current revenue, LTV, coverage ratio, and equity cushion on deal and fund level using current invested capital weightings.",
    },
    "credit-pricing-trends": {
        "title": "Credit Pricing Trends",
        "description": "Track coupon rates, floor rates, and upfront fees by entry date, by fund, and by qualitative dimensions.",
    },
    "credit-underwrite-outcome": {
        "title": "Credit Underwrite vs Outcome",
        "description": "Compare Estimated IRR at Entry against actual Gross IRR, rank the biggest underwriting misses, and inspect full loan context.",
    },
    "credit-data-cuts": {
        "title": "Credit Data Cuts",
        "description": "Slice credit portfolio performance by any dimension with cross-tab analysis.",
    },
}

CREDIT_SIDEBAR_ITEMS = [
    {"page_key": "credit-track-record", "label": "Track Record", "icon": "bi bi-table"},
    {"page_key": "credit-benchmarking", "label": "Benchmarking", "icon": "bi bi-bar-chart-steps"},
    {"page_key": "credit-concentration", "label": "Concentration", "icon": "bi bi-pie-chart"},
    {"page_key": "credit-fundamentals", "label": "Fundamentals", "icon": "bi bi-bar-chart-line"},
    {"page_key": "credit-pricing-trends", "label": "Pricing Trends", "icon": "bi bi-graph-up-arrow"},
    {"page_key": "credit-underwrite-outcome", "label": "Underwrite vs Outcome", "icon": "bi bi-bullseye"},
    {"page_key": "credit-data-cuts", "label": "Data Cuts", "icon": "bi bi-sliders"},
    {"endpoint": "live_credit_pdf_pack", "label": "Download 7 Analysis PDFs", "icon": "bi bi-file-earmark-zip"},
]

CREDIT_PDF_EXPORT_PAGES = (
    {"page": "credit-track-record", "analysis_name": "Credit Track Record"},
    {"page": "credit-benchmarking", "analysis_name": "Credit Benchmarking Analysis"},
    {"page": "credit-concentration", "analysis_name": "Credit Concentration"},
    {"page": "credit-fundamentals", "analysis_name": "Credit Fundamentals"},
    {"page": "credit-pricing-trends", "analysis_name": "Credit Pricing Trends"},
    {"page": "credit-underwrite-outcome", "analysis_name": "Credit Underwrite vs Outcome"},
    {"page": "credit-data-cuts", "analysis_name": "Credit Data Cuts Summary"},
)

TEAM_ROLE_OWNER = "owner"
TEAM_ROLE_ADMIN = "admin"
TEAM_ROLE_MEMBER = "member"
TEAM_ALLOWED_ROLES = {TEAM_ROLE_OWNER, TEAM_ROLE_ADMIN, TEAM_ROLE_MEMBER}

ROUTE_BLUEPRINTS = {
    "healthz": "dashboard",
    "readyz": "dashboard",
    "index": "dashboard",
    "dashboard": "dashboard",
    "dashboard_series_api": "dashboard",
    "login": "auth",
    "logout": "auth",
    "accept_invite": "auth",
    "team": "team",
    "create_team_invite": "team",
    "firms": "scope",
    "export_firm_excel": "scope",
    "select_firm_scope": "scope",
    "delete_firm": "scope",
    "funds": "scope",
    "select_fund_scope": "scope",
    "delete_fund": "scope",
    "analysis_page": "analysis",
    "credit_analysis_page": "credit",
    "credit_analysis_series_api": "credit",
    "upload_credit_loans": "uploads",
    "download_credit_template": "uploads",
    "analysis_series_api": "analysis",
    "ic_memo": "analysis",
    "methodology": "analysis",
    "methodology_alias": "analysis",
    "deal_bridge_api": "analysis",
    "api_fund_comparison_deals": "analysis",
    "chart_builder_catalog_api": "chart_builder_api",
    "chart_builder_query_api": "chart_builder_api",
    "chart_builder_templates_api": "chart_builder_api",
    "chart_builder_template_create_api": "chart_builder_api",
    "chart_builder_template_update_api": "chart_builder_api",
    "chart_builder_template_delete_api": "chart_builder_api",
    "upload": "uploads",
    "delete_upload_batch": "uploads",
    "download_deal_template": "uploads",
    "download_benchmark_template": "uploads",
    "upload_deals": "uploads",
    "upload_benchmarks": "uploads",
    "delete_benchmarks": "uploads",
    "deals": "reports",
    "deals_analysis": "reports",
    "track_record": "reports",
    "download_track_record_pdf": "reports",
    "live_ic_pdf_pack": "reports",
    "download_ic_pdf_pack": "reports",
    "live_credit_pdf_pack": "reports",
    "download_credit_pdf_pack": "reports",
    "download_credit_analysis_pdf": "reports",
    "memo_studio": "memos",
    "memo_style_library": "memos",
    "memo_source_library": "memos",
    "memo_run_page": "memos",
    "memo_documents_api": "memos",
    "memo_style_profiles_api": "memos",
    "memo_style_profile_rebuild_api": "memos",
    "memo_runs_api": "memos",
    "memo_run_api": "memos",
    "memo_run_sections_api": "memos",
    "memo_run_cancel_api": "memos",
    "memo_run_rerun_section_api": "memos",
    "memo_run_approve_api": "memos",
    "memo_run_export_api": "memos",
}


def _utc_now():
    return datetime.now(timezone.utc)


def _utc_now_naive():
    return _utc_now().replace(tzinfo=None)


def _schema_upgrade_message():
    return f"Database schema is not ready. Run `{SCHEMA_UPGRADE_COMMAND}` on the deployed service and retry."


def _root_db_error(exc):
    return getattr(exc, "orig", exc)


def _rollback_db_session():
    try:
        db.session.rollback()
    except Exception:
        logger.exception("Database session rollback failed")


def _handle_db_exception(exc, log_message):
    _rollback_db_session()
    logger.exception("%s [%s]: %s", log_message, type(exc).__name__, _root_db_error(exc))


def _json_schema_failure(exc, log_message, status_code=503):
    _handle_db_exception(exc, log_message)
    root = _root_db_error(exc)
    return (
        jsonify(
            {
                "error": "database_schema_not_ready",
                "message": _schema_upgrade_message(),
                "detail": f"{type(root).__name__}: {str(root)[:500]}",
            }
        ),
        status_code,
    )


def _redirect_schema_failure(exc, log_message, endpoint="dashboard"):
    _handle_db_exception(exc, log_message)
    root = _root_db_error(exc)
    detail = f" Detail: {type(root).__name__}: {str(root)[:300]}"
    flash(_schema_upgrade_message() + detail, "danger")
    return redirect(url_for(endpoint))


def _default_scope_context():
    code = DEFAULT_CURRENCY_CODE
    return {
        "app_firms": [],
        "app_active_firm": None,
        "app_active_firm_id": None,
        "app_currency_code": code,
        "app_currency_symbol": currency_symbol(code) or "",
        "app_currency_unit_label": currency_unit_label(code),
        "fmt_currency_millions": format_currency_millions,
        "app_native_currency": code,
        "app_conversion_active": False,
        "app_conversion_note": None,
        "app_conversion_warning": None,
        "app_conversion_rate": 1.0,
        "app_conversion_date": None,
        "app_conversion_source": "Identity",
        "app_fx_status": "ok",
        "app_money_scale": 1.0,
        "app_active_team": None,
        "app_active_membership": None,
        "app_team_is_admin": False,
    }


def _fund_vintage_lookup_for_scope(deals, membership=None, active_firm=None):
    return build_fund_vintage_lookup(
        deals,
        team_id=membership.team_id if membership is not None else None,
        firm_id=active_firm.id if active_firm is not None else None,
    )


def _firm_currency_code(firm):
    if firm is None:
        return DEFAULT_CURRENCY_CODE
    return normalize_currency_code(getattr(firm, "base_currency", None), default=DEFAULT_CURRENCY_CODE) or DEFAULT_CURRENCY_CODE


MONETARY_DRIVER_KEYS = ("revenue", "ebitda_growth", "margin", "multiple", "leverage", "other")
DEAL_METRIC_MONEY_KEYS = (
    "equity",
    "realized",
    "unrealized",
    "value_total",
    "value_created",
    "entry_revenue",
    "entry_ebitda",
    "entry_enterprise_value",
    "entry_net_debt",
    "exit_revenue",
    "exit_ebitda",
    "exit_enterprise_value",
    "exit_net_debt",
)


def _scale_money(value, scale):
    if scale == 1.0 or value is None or isinstance(value, bool):
        return value
    try:
        return float(value) * scale
    except (TypeError, ValueError):
        return value


def _safe_fx_rate(value):
    try:
        rate = float(value)
    except (TypeError, ValueError):
        return None
    return rate if rate > 0 else None


def _reporting_currency_context(firm):
    native_code = _firm_currency_code(firm)
    rate = _safe_fx_rate(getattr(firm, "fx_rate_to_usd", None)) if firm is not None else None
    fx_status = (getattr(firm, "fx_last_status", None) or "").strip().lower() if firm is not None else ""
    fx_date = getattr(firm, "fx_rate_date", None) if firm is not None else None
    fx_source = getattr(firm, "fx_rate_source", None) if firm is not None else None

    # Deal-level currency conversion now happens at upload time (deal_parser.py).
    # All values in the database are already in USD. The firm-level money_scale
    # must NOT re-apply the FX rate, or values get double-converted.
    # money_scale is always 1.0 — the rate is kept for display/metadata only.

    if native_code == DEFAULT_CURRENCY_CODE:
        return {
            "native_currency_code": DEFAULT_CURRENCY_CODE,
            "reporting_currency_code": DEFAULT_CURRENCY_CODE,
            "money_scale": 1.0,
            "conversion_active": False,
            "fx_status": "ok",
            "fx_rate": 1.0,
            "fx_date": fx_date,
            "fx_source": fx_source or "Identity",
            "conversion_note": None,
            "conversion_warning": None,
        }

    conversion_active = fx_status == "ok" and rate is not None
    if conversion_active:
        note = (
            f"Values converted from {native_code} to USD at upload time. "
            f"Rate: {rate:.6f} (effective {fx_date.isoformat() if fx_date else 'N/A'}, source {fx_source or 'N/A'})."
        )
        return {
            "native_currency_code": native_code,
            "reporting_currency_code": DEFAULT_CURRENCY_CODE,
            "money_scale": 1.0,  # Values already in USD from upload-time conversion
            "conversion_active": True,
            "fx_status": fx_status,
            "fx_rate": rate,
            "fx_date": fx_date,
            "fx_source": fx_source,
            "conversion_note": note,
            "conversion_warning": None,
        }

    warning = f"FX unavailable; showing native {native_code} values."
    return {
        "native_currency_code": native_code,
        "reporting_currency_code": native_code,
        "money_scale": 1.0,
        "conversion_active": False,
        "fx_status": fx_status or "lookup_failed",
        "fx_rate": rate,
        "fx_date": fx_date,
        "fx_source": fx_source,
        "conversion_note": None,
        "conversion_warning": warning,
    }


def _scale_metric_pair(metric_pair, scale):
    if not isinstance(metric_pair, dict):
        return
    metric_pair["avg"] = _scale_money(metric_pair.get("avg"), scale)
    metric_pair["wavg"] = _scale_money(metric_pair.get("wavg"), scale)


def _scale_portfolio_entry_exit(summary, scale):
    if scale == 1.0 or not isinstance(summary, dict):
        return
    for side in ("entry", "exit"):
        side_payload = summary.get(side) or {}
        for key in ("revenue", "ebitda", "tev", "net_debt"):
            _scale_metric_pair(side_payload.get(key), scale)


def _scale_bridge_view_payload(bridge, scale):
    if scale == 1.0 or not isinstance(bridge, dict):
        return
    for key in ("value_created", "fund_value_created", "company_value_created"):
        bridge[key] = _scale_money(bridge.get(key), scale)

    for map_key in ("drivers_dollar", "fund_drivers_dollar", "company_drivers_dollar"):
        drivers = bridge.get(map_key)
        if not isinstance(drivers, dict):
            continue
        for driver in MONETARY_DRIVER_KEYS:
            if driver in drivers:
                drivers[driver] = _scale_money(drivers.get(driver), scale)

    if bridge.get("unit") == "dollar" and isinstance(bridge.get("drivers"), dict):
        for driver in MONETARY_DRIVER_KEYS:
            if driver in bridge["drivers"]:
                bridge["drivers"][driver] = _scale_money(bridge["drivers"].get(driver), scale)

    for row in bridge.get("display_drivers") or []:
        if isinstance(row, dict):
            row["dollar"] = _scale_money(row.get("dollar"), scale)

    start_end = bridge.get("start_end")
    if isinstance(start_end, dict) and isinstance(start_end.get("dollar"), dict):
        start_end["dollar"]["start"] = _scale_money(start_end["dollar"].get("start"), scale)
        start_end["dollar"]["end"] = _scale_money(start_end["dollar"].get("end"), scale)


def _scale_bridge_aggregate_payload(aggregate, scale):
    if scale == 1.0 or not isinstance(aggregate, dict):
        return
    aggregate["total_value_created"] = _scale_money(aggregate.get("total_value_created"), scale)
    aggregate["total_equity"] = _scale_money(aggregate.get("total_equity"), scale)

    drivers = aggregate.get("drivers")
    if isinstance(drivers, dict) and isinstance(drivers.get("dollar"), dict):
        for driver in MONETARY_DRIVER_KEYS:
            if driver in drivers["dollar"]:
                drivers["dollar"][driver] = _scale_money(drivers["dollar"].get(driver), scale)

    for row in aggregate.get("display_drivers") or []:
        if isinstance(row, dict):
            row["dollar"] = _scale_money(row.get("dollar"), scale)

    start_end = aggregate.get("start_end")
    if isinstance(start_end, dict) and isinstance(start_end.get("dollar"), dict):
        start_end["dollar"]["start"] = _scale_money(start_end["dollar"].get("start"), scale)
        start_end["dollar"]["end"] = _scale_money(start_end["dollar"].get("end"), scale)


def _scale_dashboard_payload(payload, scale):
    if scale == 1.0:
        return
    kpis = payload.get("kpis") or {}
    for key in ("total_equity", "total_value", "total_value_created"):
        kpis[key] = _scale_money(kpis.get(key), scale)

    _scale_portfolio_entry_exit(payload.get("entry_exit_summary"), scale)
    _scale_bridge_aggregate_payload(payload.get("bridge_aggregate"), scale)

    for row in payload.get("vintage_series") or []:
        row["total_equity"] = _scale_money(row.get("total_equity"), scale)
        row["total_value_created"] = _scale_money(row.get("total_value_created"), scale)

    for point in payload.get("moic_hold_scatter") or []:
        point["equity"] = _scale_money(point.get("equity"), scale)

    for series in (payload.get("value_creation_mix") or {}).get("series", {}).values():
        totals = series.get("totals_dollar") or []
        for idx, value in enumerate(totals):
            totals[idx] = _scale_money(value, scale)

    exposure = payload.get("realized_unrealized_exposure") or {}
    for key in ("realized", "unrealized"):
        vals = exposure.get(key) or []
        for idx, value in enumerate(vals):
            vals[idx] = _scale_money(value, scale)

    heatmap = payload.get("loss_concentration_heatmap") or {}
    values = heatmap.get("values") or []
    for row in values:
        for idx, value in enumerate(row):
            row[idx] = _scale_money(value, scale)
    heatmap["max_value"] = _scale_money(heatmap.get("max_value"), scale)

    exit_perf = payload.get("exit_type_performance") or {}
    realized_vals = exit_perf.get("realized_value") or []
    for idx, value in enumerate(realized_vals):
        realized_vals[idx] = _scale_money(value, scale)

    for row in payload.get("lead_partner_scorecard") or []:
        row["capital_deployed"] = _scale_money(row.get("capital_deployed"), scale)

    for row in payload.get("fund_summary_rows") or []:
        row["fund_size"] = _scale_money(row.get("fund_size"), scale)

    _scale_deal_metrics(payload.get("deal_metrics"), scale)


def _scale_deal_metrics(metrics_by_id, scale):
    if scale == 1.0 or not isinstance(metrics_by_id, dict):
        return

    for metric in metrics_by_id.values():
        if not isinstance(metric, dict):
            continue
        for key in DEAL_METRIC_MONEY_KEYS:
            metric[key] = _scale_money(metric.get(key), scale)

        bridge = metric.get("bridge_additive_fund")
        if isinstance(bridge, dict):
            _scale_bridge_view_payload(bridge, scale)

        sens = ((metric.get("bridge_diagnostics") or {}).get("ownership_sensitivity") or {})
        for key in (
            "driver_subtotal_base",
            "driver_subtotal_up_10",
            "driver_subtotal_down_10",
            "other_up_10",
            "other_down_10",
        ):
            sens[key] = _scale_money(sens.get(key), scale)


def _scale_track_totals(totals, scale):
    if scale == 1.0 or not isinstance(totals, dict):
        return
    for key in ("invested_equity", "realized_value", "unrealized_value", "total_value"):
        totals[key] = _scale_money(totals.get(key), scale)


def _scale_track_record_payload(track_record, scale):
    if scale == 1.0 or not isinstance(track_record, dict):
        return

    for fund in track_record.get("funds", []):
        fund["fund_size"] = _scale_money(fund.get("fund_size"), scale)
        _scale_track_totals(fund.get("totals"), scale)

        for row in fund.get("rows", []):
            row["fund_size"] = _scale_money(row.get("fund_size"), scale)
            for key in ("invested_equity", "realized_value", "unrealized_value", "total_value"):
                row[key] = _scale_money(row.get(key), scale)

        for rollup in (fund.get("status_rollups") or []):
            _scale_track_totals(rollup.get("totals"), scale)
        for rollup in (fund.get("summary_rollups") or []):
            _scale_track_totals(rollup.get("totals"), scale)

    overall = track_record.get("overall") or {}
    _scale_track_totals(overall.get("totals"), scale)
    for rollup in (overall.get("status_rollups") or []):
        _scale_track_totals(rollup.get("totals"), scale)
    for rollup in (overall.get("summary_rollups") or []):
        _scale_track_totals(rollup.get("totals"), scale)


def _scale_rollup_details_payload(rollup_details, scale):
    if scale == 1.0 or not isinstance(rollup_details, dict):
        return
    for detail in rollup_details.values():
        _scale_portfolio_entry_exit(detail.get("entry_exit"), scale)
        bridge = detail.get("bridge")
        _scale_bridge_aggregate_payload(bridge, scale)


def _scale_ic_memo_payload(memo, scale):
    if scale == 1.0 or not isinstance(memo, dict):
        return

    executive = memo.get("executive") or {}
    for key in ("total_equity", "realized_value", "unrealized_value", "total_value", "total_value_created"):
        executive[key] = _scale_money(executive.get(key), scale)
    for key in ("top_5_deals", "bottom_5_deals"):
        for row in executive.get(key) or []:
            for money_key in ("invested_equity", "total_value", "value_created"):
                row[money_key] = _scale_money(row.get(money_key), scale)

    bridge = memo.get("bridge") or {}
    _scale_bridge_aggregate_payload(bridge, scale)
    for row in bridge.get("table_rows") or []:
        row["dollar"] = _scale_money(row.get("dollar"), scale)

    for dim in (memo.get("slicing") or {}).get("dimensions", {}).values():
        for key in ("groups", "top_decile", "bottom_decile"):
            for row in dim.get(key) or []:
                for money_key in ("invested_equity", "total_value", "value_created"):
                    row[money_key] = _scale_money(row.get(money_key), scale)

    for row in (memo.get("team") or {}).get("lead_partner_table", []):
        row["capital_deployed"] = _scale_money(row.get("capital_deployed"), scale)
        row["value_created"] = _scale_money(row.get("value_created"), scale)
    for row in (memo.get("team") or {}).get("entry_channel_table", []):
        row["capital_deployed"] = _scale_money(row.get("capital_deployed"), scale)
        row["value_created"] = _scale_money(row.get("value_created"), scale)


def _scale_analysis_payload(page, payload, scale):
    if scale == 1.0 or not isinstance(payload, dict):
        return

    if page == "fund-liquidity":
        for key in ("paid_in", "distributed", "nav", "unfunded"):
            vals = payload.get(key) or []
            for idx, value in enumerate(vals):
                vals[idx] = _scale_money(value, scale)
        latest = payload.get("latest") or {}
        for key in ("paid_in", "distributed", "nav", "unfunded"):
            latest[key] = _scale_money(latest.get(key), scale)
        for row in payload.get("fund_summaries") or []:
            for key in ("committed_capital", "paid_in_capital", "distributed_capital", "nav", "unfunded_commitment"):
                row[key] = _scale_money(row.get(key), scale)
        return

    if page == "underwrite-outcome":
        coverage = payload.get("coverage") or {}
        coverage["invested_equity"] = _scale_money(coverage.get("invested_equity"), scale)
        for key in ("rows", "by_partner", "by_sector", "by_entry_channel"):
            for row in payload.get(key) or []:
                row["invested_equity"] = _scale_money(row.get("invested_equity"), scale)
        return

    if page == "valuation-quality":
        for row in payload.get("unrealized_rows") or []:
            for key in ("latest_mark", "invested_equity", "unrealized_value"):
                row[key] = _scale_money(row.get(key), scale)
        for row in payload.get("mark_error_rows") or []:
            row["pre_exit_mark"] = _scale_money(row.get("pre_exit_mark"), scale)
            row["realized_value"] = _scale_money(row.get("realized_value"), scale)
        return

    if page == "exit-readiness":
        for row in payload.get("aging_buckets") or []:
            row["invested_equity"] = _scale_money(row.get("invested_equity"), scale)
            row["unrealized_value"] = _scale_money(row.get("unrealized_value"), scale)
        for key in ("aging_by_fund", "aging_by_sector"):
            for row in payload.get(key) or []:
                buckets = row.get("buckets") or {}
                for bucket_key, value in list(buckets.items()):
                    buckets[bucket_key] = _scale_money(value, scale)
        for row in payload.get("rows") or []:
            row["invested_equity"] = _scale_money(row.get("invested_equity"), scale)
            row["unrealized_value"] = _scale_money(row.get("unrealized_value"), scale)
        return

    if page == "stress-lab":
        summary = payload.get("summary") or {}
        for key in ("invested_equity", "current_value", "stressed_value", "base_total_value", "stressed_total_value", "delta_value"):
            summary[key] = _scale_money(summary.get(key), scale)

        for row in payload.get("deal_rows") or []:
            for key in ("current_ebitda", "stressed_ebitda", "invested_equity", "base_total_value", "current_total_value", "stressed_total_value", "delta_value"):
                row[key] = _scale_money(row.get(key), scale)

        for row in payload.get("fund_subtotals") or []:
            for key in ("invested_equity", "current_total_value", "stressed_total_value", "delta_value"):
                row[key] = _scale_money(row.get(key), scale)
        for row in (payload.get("fund_subtotals_map") or {}).values():
            for key in ("invested_equity", "current_total_value", "stressed_total_value", "delta_value"):
                row[key] = _scale_money(row.get(key), scale)
        return

    if page == "deal-trajectory":
        summary = payload.get("summary") or {}
        summary["current_equity_value"] = _scale_money(summary.get("current_equity_value"), scale)
        for row in payload.get("trajectory") or []:
            for key in ("revenue", "ebitda", "enterprise_value", "net_debt", "equity_value"):
                row[key] = _scale_money(row.get(key), scale)
        for row in payload.get("cashflow_curve") or []:
            for key in ("calls", "distributions", "cum_calls", "cum_distributions"):
                row[key] = _scale_money(row.get(key), scale)
        return

    if page == "nav-at-risk":
        summary = payload.get("summary") or {}
        summary["total_nav"] = _scale_money(summary.get("total_nav"), scale)
        for row in payload.get("aging_buckets") or []:
            row["nav"] = _scale_money(row.get("nav"), scale)
        for key in ("deal_rows", "fund_rows"):
            for row in payload.get(key) or []:
                if "unrealized_value" in row:
                    row["unrealized_value"] = _scale_money(row.get("unrealized_value"), scale)
                if "nav" in row:
                    row["nav"] = _scale_money(row.get("nav"), scale)
        for key in ("by_sector", "by_vintage"):
            for row in (payload.get("concentration") or {}).get(key) or []:
                row["nav"] = _scale_money(row.get("nav"), scale)
        return

    if page == "public-market-comparison":
        for row in payload.get("fund_rows") or []:
            row["nav_used"] = _scale_money(row.get("nav_used"), scale)
        for row in payload.get("series") or []:
            row["amount"] = _scale_money(row.get("amount"), scale)
            row["future_value"] = _scale_money(row.get("future_value"), scale)
        return

    if page == "lp-due-diligence-memo":
        for row in payload.get("fund_metadata") or []:
            row["fund_size"] = _scale_money(row.get("fund_size"), scale)
        _scale_analysis_payload("fund-liquidity", payload.get("fund_liquidity") or {}, scale)
        _scale_analysis_payload("nav-at-risk", payload.get("nav_at_risk") or {}, scale)
        _scale_analysis_payload("public-market-comparison", payload.get("public_market_comparison") or {}, scale)
        return

    if page == "organic-growth":
        deal_money_keys = (
            "entry_revenue", "exit_revenue", "acquired_revenue",
            "organic_revenue_growth", "acquired_revenue_contribution", "total_revenue_growth",
            "entry_ebitda", "exit_ebitda", "acquired_ebitda",
            "organic_ebitda_growth", "acquired_ebitda_contribution", "total_ebitda_growth",
            "acquired_tev", "entry_enterprise_value", "exit_enterprise_value",
            "equity", "total_value",
        )
        for row in payload.get("deal_rows") or []:
            for k in deal_money_keys:
                if k in row:
                    row[k] = _scale_money(row.get(k), scale)

        # Scale aggregate money fields in total and cohort dicts
        agg_money_keys = (
            "total_equity", "total_value",
            "total_organic_rev", "total_acquired_rev",
            "total_organic_ebitda", "total_acquired_ebitda",
        )
        for agg in [payload.get("total")] + list((payload.get("cohorts") or {}).values()):
            if agg:
                for k in agg_money_keys:
                    if k in agg:
                        agg[k] = _scale_money(agg.get(k), scale)

        # Scale fund comparison aggregates
        for fund_agg in payload.get("fund_comparison") or []:
            for k in agg_money_keys:
                if k in fund_agg:
                    fund_agg[k] = _scale_money(fund_agg.get(k), scale)

        # Scale waterfall data
        wf = payload.get("waterfall") or {}
        for wf_key in ("revenue", "ebitda"):
            wf_data = wf.get(wf_key) or {}
            for k in ("entry", "organic", "acquired", "exit"):
                if k in wf_data:
                    wf_data[k] = _scale_money(wf_data.get(k), scale)

        # Scale chart series
        charts = payload.get("charts") or {}
        for chart_key in ("revenue", "ebitda"):
            chart = charts.get(chart_key) or {}
            for series_key in ("organic", "acquired"):
                vals = chart.get(series_key) or []
                for i, v in enumerate(vals):
                    vals[i] = _scale_money(v, scale)

        # Scale scatter data equity
        for pt in payload.get("scatter_data") or []:
            pt["equity"] = _scale_money(pt.get("equity"), scale)

        # Scale bridge decomposition
        for bd in payload.get("bridge_decomposition") or []:
            for k in ("organic_revenue_contribution", "acquired_revenue_contribution",
                       "total_revenue_driver", "margin_contribution",
                       "multiple_contribution", "leverage_contribution"):
                bd[k] = _scale_money(bd.get(k), scale)
        return

    if page == "vca-ebitda":
        money_keys = (
            "fund_total_cost",
            "realized_proceeds",
            "unrealized_value",
            "total_value",
            "gross_profit",
            "vc_ebitda_growth_dollar",
            "vc_multiple_dollar",
            "vc_debt_dollar",
            "vc_total_dollar",
            "entry_ltm_ebitda",
            "exit_ltm_ebitda",
            "diff_ebitda",
        )

        for fund in payload.get("fund_blocks") or []:
            fund["fund_size"] = _scale_money(fund.get("fund_size"), scale)
            sort_metrics = fund.get("print_sort_metrics") or {}
            sort_metrics["gross_profit"] = _scale_money(sort_metrics.get("gross_profit"), scale)
            for key in ("deal_rows", "subtotal_rows", "summary_rows"):
                for row in fund.get(key) or []:
                    for money_key in money_keys:
                        row[money_key] = _scale_money(row.get(money_key), scale)

        overall = payload.get("overall_block") or {}
        for key in ("subtotal_rows", "summary_rows"):
            for row in overall.get(key) or []:
                for money_key in money_keys:
                    row[money_key] = _scale_money(row.get(money_key), scale)
        summary_metrics = overall.get("summary_metrics") or {}
        summary_metrics["gross_profit"] = _scale_money(summary_metrics.get("gross_profit"), scale)
        return

    if page == "vca-revenue":
        money_keys = (
            "fund_total_cost",
            "realized_proceeds",
            "unrealized_value",
            "total_value",
            "gross_profit",
            "vc_revenue_growth_dollar",
            "vc_multiple_dollar",
            "vc_debt_dollar",
            "vc_total_dollar",
            "entry_ltm_revenue",
            "entry_tev",
            "exit_ltm_revenue",
            "exit_tev",
            "diff_revenue",
            "diff_tev",
        )

        for fund in payload.get("fund_blocks") or []:
            fund["fund_size"] = _scale_money(fund.get("fund_size"), scale)
            sort_metrics = fund.get("print_sort_metrics") or {}
            sort_metrics["gross_profit"] = _scale_money(sort_metrics.get("gross_profit"), scale)
            for key in ("deal_rows", "subtotal_rows", "summary_rows"):
                for row in fund.get(key) or []:
                    for money_key in money_keys:
                        row[money_key] = _scale_money(row.get(money_key), scale)

        overall = payload.get("overall_block") or {}
        for key in ("subtotal_rows", "summary_rows"):
            for row in overall.get(key) or []:
                for money_key in money_keys:
                    row[money_key] = _scale_money(row.get(money_key), scale)
        summary_metrics = overall.get("summary_metrics") or {}
        summary_metrics["gross_profit"] = _scale_money(summary_metrics.get("gross_profit"), scale)
        return

    if page in {"vca-addons", "vca-addons-revenue"}:
        money_keys = (
            "fund_initial_cost",
            "fund_total_cost",
            "realized_proceeds",
            "unrealized_value",
            "total_value",
            "gross_profit",
            "vc_organic_ebitda_growth_dollar",
            "vc_add_on_ebitda_dollar",
            "vc_organic_revenue_growth_dollar",
            "vc_add_on_revenue_dollar",
            "vc_multiple_dollar",
            "vc_debt_dollar",
            "vc_total_dollar",
            "entry_ltm_revenue",
            "entry_ltm_ebitda",
            "entry_tev",
            "entry_net_debt",
            "acquired_revenue",
            "acquired_ebitda",
            "acquired_tev",
            "exit_ltm_revenue",
            "exit_ltm_ebitda",
            "exit_tev",
            "exit_net_debt",
            "diff_revenue",
            "diff_ebitda",
            "diff_tev",
            "diff_net_debt",
        )

        for fund in payload.get("fund_blocks") or []:
            fund["fund_size"] = _scale_money(fund.get("fund_size"), scale)
            sort_metrics = fund.get("print_sort_metrics") or {}
            sort_metrics["gross_profit"] = _scale_money(sort_metrics.get("gross_profit"), scale)
            for key in ("deal_rows", "subtotal_rows", "summary_rows"):
                for row in fund.get(key) or []:
                    for money_key in money_keys:
                        row[money_key] = _scale_money(row.get(money_key), scale)

        overall = payload.get("overall_block") or {}
        for key in ("subtotal_rows", "summary_rows"):
            for row in overall.get(key) or []:
                for money_key in money_keys:
                    row[money_key] = _scale_money(row.get(money_key), scale)
        summary_metrics = overall.get("summary_metrics") or {}
        summary_metrics["gross_profit"] = _scale_money(summary_metrics.get("gross_profit"), scale)
        return

    if page == "benchmarking":
        for row in payload.get("fund_rows") or []:
            row["fund_size"] = _scale_money(row.get("fund_size"), scale)
        return



def _slugify_team_name(name):
    token = re.sub(r"[^a-z0-9]+", "-", (name or "").strip().lower()).strip("-")
    return token or "team"


def _ensure_unique_team_slug(base_slug):
    candidate = base_slug
    idx = 2
    while Team.query.filter_by(slug=candidate).first() is not None:
        candidate = f"{base_slug}-{idx}"
        idx += 1
    return candidate


def _slugify_firm_name(name):
    token = re.sub(r"[^a-z0-9]+", "-", (name or "").strip().lower()).strip("-")
    return token or "firm"


def _ensure_unique_firm_slug(base_slug):
    candidate = base_slug
    idx = 2
    while Firm.query.filter_by(slug=candidate).first() is not None:
        candidate = f"{base_slug}-{idx}"
        idx += 1
    return candidate


def _hash_invite_token(raw_token):
    return hashlib.sha256(raw_token.encode("utf-8")).hexdigest()


def _build_invite_link(raw_token):
    return url_for("accept_invite", token=raw_token, _external=True)


@login_manager.user_loader
def load_user(user_id):
    try:
        return db.session.get(User, int(user_id))
    except Exception:
        return None


def _current_membership():
    if not current_user.is_authenticated:
        return None
    active_team_id = session.get("active_team_id")
    membership = None
    if active_team_id:
        membership = TeamMembership.query.filter_by(
            user_id=current_user.id,
            team_id=active_team_id,
        ).first()
    if membership is None:
        membership = (
            TeamMembership.query.filter_by(user_id=current_user.id)
            .order_by(TeamMembership.created_at.asc(), TeamMembership.id.asc())
            .first()
        )
        if membership is not None:
            session["active_team_id"] = membership.team_id
    return membership


def _current_team():
    membership = _current_membership()
    if membership is None:
        return None
    return db.session.get(Team, membership.team_id)


def _require_team_scope():
    membership = _current_membership()
    if membership is None:
        abort(403)
    if membership.role not in TEAM_ALLOWED_ROLES:
        abort(403)
    return membership


def _is_team_admin(membership):
    return membership is not None and membership.role in {TEAM_ROLE_OWNER, TEAM_ROLE_ADMIN}


def _active_firm_id_from_session():
    raw = session.get("active_firm_id")
    if raw in (None, ""):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _set_active_firm_scope(firm_id):
    if firm_id in (None, ""):
        session.pop("active_firm_id", None)
        return
    session["active_firm_id"] = int(firm_id)


def _accessible_firms_for_team(team_id):
    if team_id is None:
        return []
    return (
        db.session.query(Firm)
        .join(TeamFirmAccess, TeamFirmAccess.firm_id == Firm.id)
        .filter(TeamFirmAccess.team_id == team_id)
        .order_by(Firm.name.asc(), Firm.id.asc())
        .all()
    )


def _accessible_firms_for_current_team():
    membership = _current_membership()
    if membership is None:
        return []
    return _accessible_firms_for_team(membership.team_id)


def _resolve_active_firm_for_team():
    membership = _current_membership()
    if membership is None:
        return None

    accessible = _accessible_firms_for_team(membership.team_id)
    if not accessible:
        session.pop("active_firm_id", None)
        return None

    active_firm_id = _active_firm_id_from_session()
    if active_firm_id is not None:
        for firm in accessible:
            if firm.id == active_firm_id:
                return firm

    accessible_ids = [firm.id for firm in accessible]
    deal_firm_ids = {
        row[0]
        for row in db.session.query(Deal.firm_id)
        .filter(Deal.firm_id.in_(accessible_ids))
        .distinct()
        .all()
        if row[0] is not None
    }
    credit_firm_ids = {
        row[0]
        for row in db.session.query(CreditLoan.firm_id)
        .filter(CreditLoan.firm_id.in_(accessible_ids))
        .distinct()
        .all()
        if row[0] is not None
    }
    with_data_ids = deal_firm_ids | credit_firm_ids
    candidate = next((firm for firm in accessible if firm.id in with_data_ids), accessible[0])
    _set_active_firm_scope(candidate.id)
    return candidate


def _bootstrap_identity():
    # Seed default team and admin when no users exist and env bootstrap credentials are provided.
    admin_email = (os.environ.get("BOOTSTRAP_ADMIN_EMAIL") or "").strip().lower()
    admin_password = os.environ.get("BOOTSTRAP_ADMIN_PASSWORD") or ""
    admin_team_name = (os.environ.get("BOOTSTRAP_TEAM_NAME") or "Admin Team").strip()

    if User.query.count() == 0:
        if not admin_email or not admin_password:
            logger.warning(
                "No users found and bootstrap credentials are missing. "
                "Set BOOTSTRAP_ADMIN_EMAIL and BOOTSTRAP_ADMIN_PASSWORD."
            )
            return

        existing_team = Team.query.filter_by(name=admin_team_name).first()
        if existing_team is None:
            slug = _ensure_unique_team_slug(_slugify_team_name(admin_team_name))
            existing_team = Team(name=admin_team_name, slug=slug)
            db.session.add(existing_team)
            db.session.flush()

        user = User(
            email=admin_email,
            password_hash=generate_password_hash(admin_password),
            is_active=True,
        )
        db.session.add(user)
        db.session.flush()
        db.session.add(
            TeamMembership(
                team_id=existing_team.id,
                user_id=user.id,
                role=TEAM_ROLE_OWNER,
            )
        )
        db.session.commit()
        logger.info("Bootstrapped admin user '%s' and team '%s'.", admin_email, admin_team_name)

    # Backfill team_id for legacy rows.
    default_team = Team.query.filter_by(name=admin_team_name).first()
    if default_team is None:
        slug = _ensure_unique_team_slug(_slugify_team_name(admin_team_name))
        default_team = Team(name=admin_team_name, slug=slug)
        db.session.add(default_team)
        db.session.commit()

    # Ensure there is at least one firm.
    fallback_firm = Firm.query.filter_by(name="Admin Firm").first()
    if fallback_firm is None:
        fallback_firm_slug = _ensure_unique_firm_slug(_slugify_firm_name("Admin Firm"))
        fallback_firm = Firm(name="Admin Firm", slug=fallback_firm_slug)
        db.session.add(fallback_firm)
        db.session.commit()

    changed = False

    # Map teams to firms by name (one-to-one bootstrap mapping).
    team_to_firm = {}
    for team in Team.query.order_by(Team.id.asc()).all():
        firm = Firm.query.filter_by(name=team.name).first()
        if firm is None:
            firm_slug = _ensure_unique_firm_slug(_slugify_firm_name(team.name))
            firm = Firm(name=team.name, slug=firm_slug)
            db.session.add(firm)
            db.session.flush()
        team_to_firm[team.id] = firm.id

        if TeamFirmAccess.query.filter_by(team_id=team.id, firm_id=firm.id).first() is None:
            db.session.add(TeamFirmAccess(team_id=team.id, firm_id=firm.id))
            changed = True

    for model in (
        Deal,
        DealCashflowEvent,
        DealQuarterSnapshot,
        FundQuarterSnapshot,
        DealUnderwriteBaseline,
        UploadIssue,
    ):
        rows = model.query.filter(model.team_id.is_(None)).all()
        for row in rows:
            row.team_id = default_team.id
            changed = True

    # Backfill firm_id from team->firm mapping; fallback if team is missing.
    for model in (
        Deal,
        DealCashflowEvent,
        DealQuarterSnapshot,
        FundQuarterSnapshot,
        DealUnderwriteBaseline,
        UploadIssue,
    ):
        rows = model.query.filter(model.firm_id.is_(None)).all()
        for row in rows:
            if row.team_id is not None and row.team_id in team_to_firm:
                row.firm_id = team_to_firm[row.team_id]
            else:
                row.firm_id = fallback_firm.id
            changed = True

    # Backfill team-firm access from historical deals with explicit team and firm.
    deal_pairs = (
        db.session.query(Deal.team_id, Deal.firm_id)
        .filter(Deal.team_id.isnot(None), Deal.firm_id.isnot(None))
        .distinct()
        .all()
    )
    for team_id, firm_id in deal_pairs:
        if TeamFirmAccess.query.filter_by(team_id=team_id, firm_id=firm_id).first() is None:
            db.session.add(TeamFirmAccess(team_id=team_id, firm_id=firm_id))
            changed = True
    if changed:
        db.session.commit()
        logger.info(
            "Backfilled legacy rows to default team '%s' and mapped firm scopes.",
            default_team.name,
        )


def _is_missing_table_error(exc):
    message = str(_root_db_error(exc)).lower()
    missing_object_markers = (
        "no such table:",
        "no such column:",
        "undefinedtable",
        "undefinedcolumn",
        "relation ",
        "column ",
    )
    return any(marker in message for marker in missing_object_markers) and any(
        marker in message
        for marker in (
            "deals",
            "upload_issues",
            "firms",
            "users",
            "teams",
            "team_memberships",
            "team_firm_access",
            "team_invites",
            "benchmark_points",
            "chart_builder_templates",
            "fund_metadata",
            "fund_cashflows",
            "public_market_index_levels",
            "credit_loans",
            "credit_loan_snapshots",
            "credit_fund_performance",
        )
    )


def _recover_missing_tables(exc):
    _rollback_db_session()
    if _is_missing_table_error(exc):
        logger.exception(
            "Detected missing schema objects at runtime. Run '%s'. Root cause [%s]: %s",
            SCHEMA_UPGRADE_COMMAND,
            type(exc).__name__,
            _root_db_error(exc),
        )
    else:
        logger.exception(
            "Database request failed at runtime [%s]: %s",
            type(exc).__name__,
            _root_db_error(exc),
        )
    return False


def _run_db_migrations():
    inspector = sa_inspect(db.engine)
    table_names = set(inspector.get_table_names())
    user_tables = {name for name in table_names if name != "alembic_version"}

    if user_tables and "alembic_version" not in table_names:
        logger.info("Detected legacy non-versioned schema; normalizing and stamping current head revision.")
        db.create_all()
        ensure_schema_updates()
        migrate_stamp(revision="head")
    migrate_upgrade()
    # Always run idempotent schema updates (adds new columns, indexes, etc.)
    # even on Alembic-managed databases. This covers columns added to models
    # that don't yet have a dedicated Alembic migration file.
    ensure_schema_updates()


@app.cli.command("db-upgrade")
@with_appcontext
def db_upgrade_command():
    """Create or upgrade the schema for the current environment."""
    _run_db_migrations()
    click.echo("Database schema is up to date.")


@app.cli.command("bootstrap-admin")
@with_appcontext
def bootstrap_admin_command():
    """Bootstrap the default admin user and team if missing."""
    _run_db_migrations()
    _bootstrap_identity()
    click.echo("Bootstrap completed.")


@app.cli.command("fx-refresh")
@click.option("--firm-id", type=int, default=None, help="Refresh a single firm by id.")
@click.option("--failed-only/--all", "failed_only", default=True, show_default=True, help="Refresh failed firms only or all firms.")
@click.option("--as-of", type=click.DateTime(formats=["%Y-%m-%d"]), default=None, help="Override FX lookup date (YYYY-MM-DD).")
@with_appcontext
def fx_refresh_command(firm_id, failed_only, as_of):
    """Refresh firm FX metadata for USD reporting conversion."""
    import app as app_module

    query = Firm.query.order_by(Firm.id.asc())
    if firm_id is not None:
        query = query.filter(Firm.id == firm_id)
    if failed_only:
        query = query.filter(or_(Firm.fx_last_status.is_(None), Firm.fx_last_status != "ok"))

    firms = query.all()
    if not firms:
        click.echo("No firms matched the requested scope.")
        return

    scanned = 0
    updated_ok = 0
    still_failed = 0
    skipped_usd = 0
    as_of_override = as_of.date() if as_of is not None else None

    for firm in firms:
        scanned += 1
        code = normalize_currency_code(getattr(firm, "base_currency", None), default=DEFAULT_CURRENCY_CODE) or DEFAULT_CURRENCY_CODE
        firm.base_currency = code

        if code == DEFAULT_CURRENCY_CODE:
            skipped_usd += 1
            continue

        if as_of_override is not None:
            lookup_date = as_of_override
        else:
            latest_created = db.session.query(db.func.max(Deal.created_at)).filter(Deal.firm_id == firm.id).scalar()
            if latest_created is not None and hasattr(latest_created, "date"):
                lookup_date = latest_created.date()
            else:
                lookup_date = date.today()

        fx = app_module.resolve_rate_to_usd(code, lookup_date)
        if fx.get("ok"):
            firm.fx_rate_to_usd = fx.get("rate")
            firm.fx_rate_date = fx.get("effective_date") or lookup_date
            firm.fx_rate_source = fx.get("source")
            firm.fx_last_status = "ok"
            updated_ok += 1
            click.echo(
                f"[ok] firm_id={firm.id} name={firm.name} {code}->USD "
                f"rate={float(firm.fx_rate_to_usd):.6f} date={firm.fx_rate_date.isoformat() if firm.fx_rate_date else 'N/A'}"
            )
            continue

        firm.fx_rate_to_usd = None
        firm.fx_rate_date = None
        firm.fx_rate_source = fx.get("source")
        firm.fx_last_status = "lookup_failed"
        still_failed += 1
        warning = str(fx.get("warning") or "FX lookup failed").splitlines()[0].strip()
        click.echo(f"[fail] firm_id={firm.id} name={firm.name} {warning}")

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        raise click.ClickException(f"FX refresh failed: {exc}") from exc

    click.echo(
        f"Summary: scanned={scanned}, updated_ok={updated_ok}, still_failed={still_failed}, skipped_usd={skipped_usd}"
    )


def _allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in app.config["ALLOWED_EXTENSIONS"]


@app.context_processor
def inject_global_scope_context():
    if not current_user.is_authenticated:
        return _default_scope_context()

    try:
        membership = _current_membership()
        team = db.session.get(Team, membership.team_id) if membership is not None else None
        active_firm = _resolve_active_firm_for_team()
        active_firm_id = active_firm.id if active_firm is not None else None
        reporting = _reporting_currency_context(active_firm)
        currency_code = reporting["reporting_currency_code"]
        firms = _accessible_firms_for_current_team()
        workflow_sidebar_items = [
            item for item in WORKFLOW_SIDEBAR_ITEMS
            if item.get("endpoint") != "analysis_page" or item.get("page_key") in ANALYSIS_PAGES
        ]
        analysis_sidebar_items = [
            item for item in ANALYSIS_SIDEBAR_ITEMS
            if item.get("page_key") in ANALYSIS_PAGES
        ]
        credit_sidebar_items = [
            item for item in CREDIT_SIDEBAR_ITEMS
            if (
                item.get("page_key") in CREDIT_ANALYSIS_PAGES
                or item.get("endpoint")
            )
        ]
        return {
            "app_firms": firms,
            "app_active_firm": active_firm,
            "app_active_firm_id": active_firm_id,
            "app_currency_code": currency_code,
            "app_currency_symbol": currency_symbol(currency_code) or "",
            "app_currency_unit_label": currency_unit_label(currency_code),
            "fmt_currency_millions": format_currency_millions,
            "app_native_currency": reporting["native_currency_code"],
            "app_conversion_active": reporting["conversion_active"],
            "app_conversion_note": reporting["conversion_note"],
            "app_conversion_warning": reporting["conversion_warning"],
            "app_conversion_rate": reporting["fx_rate"],
            "app_conversion_date": reporting["fx_date"],
            "app_conversion_source": reporting["fx_source"],
            "app_fx_status": reporting["fx_status"],
            "app_money_scale": reporting["money_scale"],
            "app_active_team": team,
            "app_active_membership": membership,
            "app_team_is_admin": _is_team_admin(membership) if membership is not None else False,
            "app_workflow_sidebar_items": workflow_sidebar_items,
            "app_analysis_sidebar_items": analysis_sidebar_items,
            "app_credit_sidebar_items": credit_sidebar_items,
        }
    except SQLAlchemyError as exc:
        _handle_db_exception(exc, "Global scope context load failed")
        return _default_scope_context()


def _deal_vintage_year(deal):
    return deal_vintage_year(deal)


def _benchmark_asset_classes_for_team(team_id):
    return context_benchmark_asset_classes_for_team(team_id)


def _load_team_benchmark_thresholds(team_id, asset_class):
    return load_team_benchmark_thresholds(team_id, asset_class)


def _rank_benchmark_metric(metric_value, vintage_year, metric_name, thresholds, asset_class_selected):
    return rank_benchmark_metric(metric_value, vintage_year, metric_name, thresholds, asset_class_selected)


def _fmt_track_date(value):
    if value is None:
        return "—"
    return value.strftime("%b-%y")


def _fmt_track_years(value):
    if value is None:
        return "—"
    return f"{value:.1f}"


def _fmt_track_pct(value):
    if value is None:
        return "—"
    return f"{value * 100:.1f}%"


def _fmt_track_multiple(value):
    if value is None:
        return "—"
    return f"{value:.2f}x"


def _fmt_track_currency(value, currency_code=DEFAULT_CURRENCY_CODE):
    return format_currency_millions(value, currency_code=currency_code, show_code=True)


def _track_totals_to_pdf_row(label, totals, include_fund_size=True, currency_code=DEFAULT_CURRENCY_CODE):
    return [
        "",
        label,
        "",
        "—",
        "—",
        _fmt_track_years(totals.get("hold_period")),
        _fmt_track_pct(totals.get("ownership_pct")),
        _fmt_track_pct(totals.get("pct_total_invested")),
        _fmt_track_pct(totals.get("pct_fund_size")) if include_fund_size else "—",
        _fmt_track_currency(totals.get("invested_equity"), currency_code=currency_code),
        _fmt_track_currency(totals.get("realized_value"), currency_code=currency_code),
        _fmt_track_currency(totals.get("unrealized_value"), currency_code=currency_code),
        _fmt_track_currency(totals.get("total_value"), currency_code=currency_code),
        _fmt_track_pct(totals.get("gross_irr")),
        _fmt_track_multiple(totals.get("gross_moic")),
        _fmt_track_multiple(totals.get("realized_gross_moic")),
        _fmt_track_multiple(totals.get("unrealized_gross_moic")),
    ]


VCA_PRINT_LABEL_MAPS = {
    "ebitda": {
        "row_num": "#",
        "platform": "Platform",
        "close_date": "Close",
        "final_exit_date": "Exit",
        "hold_period": "Hold*",
        "status": "Status",
        "fund_total_cost": "Cost",
        "realized_proceeds": "Realized",
        "unrealized_value": "Unrealized",
        "total_value": "Total Val",
        "gross_profit": "Gross Pft",
        "gross_profit_pct_of_total": "GP % Tot",
        "gross_irr": "Gross IRR",
        "realized_moic": "Real MOIC",
        "gross_moic": "Gross MOIC",
        "ebitda_cagr": "EBITDA CAGR*",
        "ebitda_cumulative_growth": "EBITDA Cum*",
        "vc_ebitda_growth_pct": "EBITDA %",
        "vc_multiple_pct": "Multiple %",
        "vc_debt_pct": "Debt %",
        "vc_total_pct": "Total %",
        "vc_ebitda_growth_dollar": "EBITDA $",
        "vc_multiple_dollar": "Multiple $",
        "vc_debt_dollar": "Debt $",
        "vc_total_dollar": "Total $",
        "entry_ltm_ebitda": "Ent EBITDA",
        "entry_ebitda_margin": "Ent Margin",
        "entry_ev_ebitda": "Ent EV/EBITDA*",
        "entry_net_debt_ebitda": "Ent ND/EBITDA*",
        "entry_net_debt_ev": "Ent ND/EV",
        "exit_ltm_ebitda": "Exit EBITDA",
        "exit_ebitda_margin": "Exit Margin",
        "exit_ev_ebitda": "Exit EV/EBITDA*",
        "exit_net_debt_ebitda": "Exit ND/EBITDA*",
        "exit_net_debt_ev": "Exit ND/EV",
        "diff_ebitda": "Diff EBITDA",
        "diff_ebitda_margin": "Diff Margin",
        "diff_ev_ebitda": "Diff EV/EBITDA*",
        "diff_net_debt_ebitda": "Diff ND/EBITDA*",
        "diff_net_debt_ev": "Diff ND/EV",
    },
    "revenue": {
        "row_num": "#",
        "platform": "Platform",
        "close_date": "Close",
        "final_exit_date": "Exit",
        "hold_period": "Hold*",
        "status": "Status",
        "fund_total_cost": "Cost",
        "realized_proceeds": "Realized",
        "unrealized_value": "Unrealized",
        "total_value": "Total Val",
        "gross_profit": "Gross Pft",
        "gross_profit_pct_of_total": "GP % Tot",
        "gross_irr": "Gross IRR",
        "realized_moic": "Real MOIC",
        "gross_moic": "Gross MOIC",
        "revenue_cagr": "Revenue CAGR*",
        "revenue_cumulative_growth": "Revenue Cum*",
        "vc_revenue_growth_pct": "Revenue %",
        "vc_multiple_pct": "Multiple %",
        "vc_debt_pct": "Debt %",
        "vc_total_pct": "Total %",
        "vc_revenue_growth_dollar": "Revenue $",
        "vc_multiple_dollar": "Multiple $",
        "vc_debt_dollar": "Debt $",
        "vc_total_dollar": "Total $",
        "entry_ltm_revenue": "Ent Revenue",
        "entry_tev": "Ent TEV",
        "entry_ev_revenue": "Ent EV/Revenue*",
        "entry_net_debt_revenue": "Ent ND/Revenue*",
        "entry_net_debt_ev": "Ent ND/EV",
        "exit_ltm_revenue": "Exit Revenue",
        "exit_tev": "Exit TEV",
        "exit_ev_revenue": "Exit EV/Revenue*",
        "exit_net_debt_revenue": "Exit ND/Revenue*",
        "exit_net_debt_ev": "Exit ND/EV",
        "diff_revenue": "Diff Revenue",
        "diff_tev": "Diff TEV",
        "diff_ev_revenue": "Diff EV/Revenue*",
        "diff_net_debt_revenue": "Diff ND/Revenue*",
        "diff_net_debt_ev": "Diff ND/EV",
    },
}

VCA_EBITDA_MONEY_KEYS = {
    "fund_total_cost",
    "realized_proceeds",
    "unrealized_value",
    "total_value",
    "gross_profit",
    "vc_ebitda_growth_dollar",
    "vc_multiple_dollar",
    "vc_debt_dollar",
    "vc_total_dollar",
    "entry_ltm_ebitda",
    "exit_ltm_ebitda",
    "diff_ebitda",
}
VCA_EBITDA_PCT_KEYS = {
    "gross_profit_pct_of_total",
    "gross_irr",
    "vc_ebitda_growth_pct",
    "vc_multiple_pct",
    "vc_debt_pct",
    "vc_total_pct",
    "entry_net_debt_ev",
    "exit_net_debt_ev",
    "diff_net_debt_ev",
}
VCA_EBITDA_PP_KEYS = {
    "ebitda_cagr",
    "ebitda_cumulative_growth",
    "entry_ebitda_margin",
    "exit_ebitda_margin",
    "diff_ebitda_margin",
}
VCA_EBITDA_MULTIPLE_KEYS = {
    "realized_moic",
    "gross_moic",
    "entry_ev_ebitda",
    "entry_net_debt_ebitda",
    "exit_ev_ebitda",
    "exit_net_debt_ebitda",
    "diff_ev_ebitda",
    "diff_net_debt_ebitda",
}
VCA_REVENUE_MONEY_KEYS = {
    "fund_total_cost",
    "realized_proceeds",
    "unrealized_value",
    "total_value",
    "gross_profit",
    "vc_revenue_growth_dollar",
    "vc_multiple_dollar",
    "vc_debt_dollar",
    "vc_total_dollar",
    "entry_ltm_revenue",
    "entry_tev",
    "exit_ltm_revenue",
    "exit_tev",
    "diff_revenue",
    "diff_tev",
}
VCA_REVENUE_PCT_KEYS = {
    "gross_profit_pct_of_total",
    "gross_irr",
    "vc_revenue_growth_pct",
    "vc_multiple_pct",
    "vc_debt_pct",
    "vc_total_pct",
    "entry_net_debt_ev",
    "exit_net_debt_ev",
    "diff_net_debt_ev",
}
VCA_REVENUE_PP_KEYS = {
    "revenue_cagr",
    "revenue_cumulative_growth",
}
VCA_REVENUE_MULTIPLE_KEYS = {
    "realized_moic",
    "gross_moic",
    "entry_ev_revenue",
    "entry_net_debt_revenue",
    "exit_ev_revenue",
    "exit_net_debt_revenue",
    "diff_ev_revenue",
    "diff_net_debt_revenue",
}


def _sanitize_filename_component(text):
    cleaned = re.sub(r'[\\/:*?"<>|]+', " ", str(text or "")).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned or "Unknown"


def _as_of_ymd(value):
    resolved = None
    if isinstance(value, datetime):
        resolved = value.date()
    elif isinstance(value, date):
        resolved = value
    else:
        resolved = date.today()
    return f"{resolved.month}.{resolved.day}.{resolved.strftime('%y')}"


def _report_title(firm_name, analysis_name, as_of_date):
    return f"{firm_name} {analysis_name} {_as_of_ymd(as_of_date)}"


def _safe_pdf_download_name(value, fallback):
    candidate = (value or "").strip() or fallback
    stem = candidate[:-4] if candidate.lower().endswith(".pdf") else candidate
    return f"{_sanitize_filename_component(stem)}.pdf"


def _fmt_symbol_currency(value, currency_code=DEFAULT_CURRENCY_CODE):
    return format_currency_millions(value, currency_code=currency_code, show_code=False)


def _fmt_pp(value):
    if value is None:
        return "—"
    return f"{value * 100:.1f}pp"


def _format_vca_value(key, value, analysis_kind, currency_code):
    if key == "row_num":
        return str(value) if value is not None else ""
    if key == "platform":
        return value or "—"
    if key in {"close_date", "final_exit_date"}:
        return _fmt_track_date(value)
    if key == "hold_period":
        return _fmt_track_years(value)
    if key == "status":
        return value or "—"

    money_keys = VCA_EBITDA_MONEY_KEYS if analysis_kind == "ebitda" else VCA_REVENUE_MONEY_KEYS
    pct_keys = VCA_EBITDA_PCT_KEYS if analysis_kind == "ebitda" else VCA_REVENUE_PCT_KEYS
    pp_keys = VCA_EBITDA_PP_KEYS if analysis_kind == "ebitda" else VCA_REVENUE_PP_KEYS
    multiple_keys = VCA_EBITDA_MULTIPLE_KEYS if analysis_kind == "ebitda" else VCA_REVENUE_MULTIPLE_KEYS

    if key in money_keys:
        return _fmt_symbol_currency(value, currency_code=currency_code)
    if key in pct_keys:
        return _fmt_track_pct(value)
    if key in pp_keys:
        return _fmt_pp(value)
    if key in multiple_keys:
        return _fmt_track_multiple(value)
    return "—" if value is None else str(value)


def _build_vca_pdf(analysis_payload, report_title, currency_code=DEFAULT_CURRENCY_CODE, analysis_kind="ebitda"):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    columns = analysis_payload.get("header", {}).get("columns") or []
    groups = analysis_payload.get("header", {}).get("groups") or []
    label_map = VCA_PRINT_LABEL_MAPS.get(analysis_kind, {})

    if not columns:
        styles = getSampleStyleSheet()
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(legal), leftMargin=18, rightMargin=18, topMargin=16, bottomMargin=16)
        doc.build([Paragraph(report_title, styles["Heading4"]), Spacer(1, 6), Paragraph("No rows available for export.", styles["Normal"])])
        return buffer.getvalue()

    group_row = []
    for group in groups:
        span = max(int(group.get("span") or 1), 1)
        group_row.append(group.get("label") or "")
        for _ in range(span - 1):
            group_row.append("")
    if len(group_row) < len(columns):
        group_row.extend([""] * (len(columns) - len(group_row)))
    group_row = group_row[: len(columns)]
    col_row = [label_map.get(col.get("key"), col.get("label") or "") for col in columns]

    rows = [group_row, col_row]
    row_tags = ["group_header", "column_header"]

    for fund in analysis_payload.get("fund_blocks") or []:
        fund_label = fund.get("fund_name") or "Unknown Fund"
        if fund.get("fund_size") is not None:
            fund_label = f"{fund_label} ({_fmt_symbol_currency(fund.get('fund_size'), currency_code=currency_code)})"
        if fund.get("fund_size_conflict"):
            fund_label = f"{fund_label} [fund size conflict]"
        rows.append([fund_label] + [""] * (len(columns) - 1))
        row_tags.append("fund_header")

        for bucket in ("deal_rows", "subtotal_rows", "summary_rows"):
            for row in fund.get(bucket) or []:
                rows.append(
                    [
                        _format_vca_value(col.get("key"), row.get(col.get("key")), analysis_kind, currency_code)
                        for col in columns
                    ]
                )
                row_tags.append(row.get("row_kind") or "deal")

    overall = analysis_payload.get("overall_block") or {}
    rows.append(["Overall Portfolio"] + [""] * (len(columns) - 1))
    row_tags.append("overall_header")
    for bucket in ("subtotal_rows", "summary_rows"):
        for row in overall.get(bucket) or []:
            rows.append(
                [
                    _format_vca_value(col.get("key"), row.get(col.get("key")), analysis_kind, currency_code)
                    for col in columns
                ]
            )
            row_tags.append(row.get("row_kind") or "summary")

    page_width, _ = landscape(legal)
    left_margin = right_margin = 12
    available_width = page_width - left_margin - right_margin
    weights = []
    for col in columns:
        key = col.get("key")
        if key == "row_num":
            weights.append(0.7)
        elif key == "platform":
            weights.append(3.0)
        elif key in {"close_date", "final_exit_date", "status"}:
            weights.append(1.6)
        else:
            weights.append(1.05)
    weight_total = sum(weights) or 1.0
    col_widths = [available_width * (w / weight_total) for w in weights]

    table = Table(rows, colWidths=col_widths, repeatRows=2)
    style_cmds = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", 4.8),
        ("FONT", (0, 0), (-1, 1), "Helvetica-Bold", 4.9),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4d78")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#2f648f")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.whitesmoke),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.28, colors.HexColor("#9eb3c5")),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.4),
        ("TOPPADDING", (0, 0), (-1, -1), 0.8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0.8),
    ]

    for idx, col in enumerate(columns):
        if col.get("numeric"):
            style_cmds.append(("ALIGN", (idx, 2), (idx, -1), "RIGHT"))

    separators = [5, 14, 16, 20, 24, 29, 34]
    for sep in separators:
        if sep < len(columns):
            style_cmds.append(("LINEBEFORE", (sep, 0), (sep, -1), 0.45, colors.black))

    detail_alt = False
    for idx, tag in enumerate(row_tags):
        if idx < 2:
            continue
        if tag == "fund_header":
            style_cmds.extend(
                [
                    ("SPAN", (0, idx), (-1, idx)),
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#9bb9d5")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 5.0),
                    ("LINEABOVE", (0, idx), (-1, idx), 0.45, colors.HexColor("#4f6e89")),
                    ("LINEBELOW", (0, idx), (-1, idx), 0.45, colors.HexColor("#4f6e89")),
                ]
            )
            detail_alt = False
        elif tag == "overall_header":
            style_cmds.extend(
                [
                    ("SPAN", (0, idx), (-1, idx)),
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#729bc0")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 5.0),
                ]
            )
        elif tag == "subtotal":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#d8e5f0")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 4.9),
                ]
            )
        elif tag == "summary":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#c7d9ea")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 4.9),
                ]
            )
        else:
            style_cmds.append(
                ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#f8fbff") if detail_alt else colors.HexColor("#f1f6fb"))
            )
            detail_alt = not detail_alt

    table.setStyle(TableStyle(style_cmds))

    styles = getSampleStyleSheet()
    as_of = analysis_payload.get("meta", {}).get("as_of_date")
    as_of_label = _as_of_ymd(as_of)
    meta = Paragraph(
        f"As of {as_of_label} | Unit {_fmt_symbol_currency(1, currency_code=currency_code).replace('1.0', '').strip()}",
        styles["Normal"],
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(legal),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=16,
        bottomMargin=12,
        title=report_title,
    )
    doc.build([Paragraph(report_title, styles["Heading4"]), meta, Spacer(1, 6), table])
    return buffer.getvalue()


def _build_data_cuts_summary_pdf(all_cuts, report_title, currency_code=DEFAULT_CURRENCY_CODE):
    """Build a multi-page PDF showing all 9 data-cut dimensions in one document."""
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from services.metrics.data_cuts import DIMENSION_LABELS

    COLUMNS = [
        ("", "label", 3.0),
        ("Deals", "deal_count", 0.8),
        ("Invested", "invested_equity", 1.3),
        ("Realized", "realized_value", 1.3),
        ("Unrealized", "unrealized_value", 1.3),
        ("Total Value", "total_value", 1.3),
        ("MOIC", "weighted_moic", 0.9),
        ("IRR", "weighted_irr", 0.9),
        ("Entry TEV/EBITDA", "weighted_entry_tev_ebitda", 1.1),
        ("Exit TEV/EBITDA", "weighted_exit_tev_ebitda", 1.1),
        ("Entry Margin", "weighted_entry_ebitda_margin", 1.0),
        ("Exit Margin", "weighted_exit_ebitda_margin", 1.0),
        ("Loss (Capital)", "loss_ratio_capital", 1.0),
    ]

    DIM_ORDER = [
        "sector", "geography", "vintage_year", "fund",
        "status", "deal_type", "exit_type", "entry_channel", "lead_partner",
    ]

    def _fmt_cell(key, value):
        if key == "label":
            return str(value or "")
        if key == "deal_count":
            return str(value) if value is not None else "\u2014"
        if key in ("invested_equity", "realized_value", "unrealized_value", "total_value"):
            return _fmt_symbol_currency(value, currency_code=currency_code)
        if key in ("weighted_moic", "weighted_entry_tev_ebitda", "weighted_exit_tev_ebitda"):
            return _fmt_track_multiple(value)
        if key in ("weighted_irr", "loss_ratio_capital"):
            return _fmt_track_pct(value)
        if key in ("weighted_entry_ebitda_margin", "weighted_exit_ebitda_margin"):
            if value is None:
                return "\u2014"
            return f"{value:.1f}%"
        return "\u2014" if value is None else str(value)

    def _table_style(num_rows):
        cmds = [
            ("FONT", (0, 0), (-1, -1), "Helvetica", 5.0),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 5.2),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4d78")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.28, colors.HexColor("#9eb3c5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.4),
            ("TOPPADDING", (0, 0), (-1, -1), 0.8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0.8),
        ]
        for idx in range(1, num_rows - 1):
            bg = colors.HexColor("#f8fbff") if (idx % 2 == 1) else colors.HexColor("#f1f6fb")
            cmds.append(("BACKGROUND", (0, idx), (-1, idx), bg))
        last = num_rows - 1
        if last > 0:
            cmds.extend([
                ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#d8e5f0")),
                ("FONT", (0, last), (-1, last), "Helvetica-Bold", 5.0),
                ("LINEABOVE", (0, last), (-1, last), 0.45, colors.HexColor("#4f6e89")),
            ])
        return cmds

    page_width, page_height = landscape(legal)
    left_margin = right_margin = 12
    available_width = page_width - left_margin - right_margin
    weights = [col[2] for col in COLUMNS]
    weight_total = sum(weights)
    col_widths = [available_width * (w / weight_total) for w in weights]

    styles = getSampleStyleSheet()
    story = [
        Paragraph(report_title, styles["Heading4"]),
        Spacer(1, 8),
    ]

    for i, dim_key in enumerate(DIM_ORDER):
        payload = all_cuts.get(dim_key)
        if payload is None:
            continue
        dim_label = DIMENSION_LABELS.get(dim_key, dim_key)
        groups = payload.get("groups") or []
        totals = payload.get("totals") or {}

        story.append(Paragraph(f"Performance by {dim_label}", styles["Heading5"]))
        story.append(Spacer(1, 4))

        header_row = [col[0] for col in COLUMNS]
        header_row[0] = dim_label

        table_data = [header_row]
        for group in groups:
            table_data.append([_fmt_cell(col[1], group.get(col[1])) for col in COLUMNS])
        totals_row = [_fmt_cell(col[1], totals.get(col[1])) for col in COLUMNS]
        totals_row[0] = "Total"
        table_data.append(totals_row)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(_table_style(len(table_data))))
        story.append(table)

        if i < len(DIM_ORDER) - 1:
            story.append(PageBreak())

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(legal),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=16,
        bottomMargin=12,
        title=report_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _build_benchmarking_pdf(
    analysis_payload,
    report_title,
    currency_code=DEFAULT_CURRENCY_CODE,
    benchmark_asset_class="",
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    as_of_label = _as_of_ymd((analysis_payload.get("meta") or {}).get("as_of_date"))
    bench_label = benchmark_asset_class or "Not Selected"
    metadata = Paragraph(f"As of {as_of_label} | Benchmark Asset Class: {bench_label}", styles["Normal"])

    fund_rows = analysis_payload.get("fund_rows") or []
    fund_table_data = [[
        "Fund",
        "Vintage Year",
        "Fund Size",
        "Net IRR",
        "Net MOIC",
        "Net DPI",
        "Net IRR Benchmark",
        "Net MOIC Benchmark",
        "Net DPI Benchmark",
    ]]
    for row in fund_rows:
        fund_table_data.append(
            [
                row.get("fund_name") or "Unknown Fund",
                str(row.get("vintage_year")) if row.get("vintage_year") is not None else "—",
                "N/A" if row.get("fund_size_conflict") else _fmt_symbol_currency(row.get("fund_size"), currency_code=currency_code),
                "N/A" if row.get("net_irr_conflict") or row.get("net_irr") is None else _fmt_track_pct(row.get("net_irr")),
                "N/A" if row.get("net_moic_conflict") or row.get("net_moic") is None else _fmt_track_multiple(row.get("net_moic")),
                "N/A" if row.get("net_dpi_conflict") or row.get("net_dpi") is None else _fmt_track_multiple(row.get("net_dpi")),
                ((row.get("benchmark_net_irr") or {}).get("label") or "N/A"),
                ((row.get("benchmark_net_moic") or {}).get("label") or "N/A"),
                ((row.get("benchmark_net_dpi") or {}).get("label") or "N/A"),
            ]
        )

    fund_table = Table(
        fund_table_data,
        colWidths=[165, 62, 78, 58, 58, 58, 88, 88, 88],
        repeatRows=1,
    )
    fund_style = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", 7.0),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 7.2),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4d78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#b5c6d6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for idx in (1, 2, 3, 4, 5):
        fund_style.append(("ALIGN", (idx, 1), (idx, -1), "RIGHT"))
    for row_idx in range(1, len(fund_table_data)):
        fund_style.append(
            ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#f5f9ff") if row_idx % 2 else colors.HexColor("#ebf2fa"))
        )
    fund_table.setStyle(TableStyle(fund_style))

    kpis = analysis_payload.get("kpis") or {}
    summary_table = Table(
        [
            ["Funds in Scope", kpis.get("fund_count"), "Any Benchmark Coverage", _fmt_track_pct(kpis.get("any_coverage_pct"))],
            ["Full Benchmark Coverage", _fmt_track_pct(kpis.get("full_coverage_pct")), "Average Composite Score", f"{(kpis.get('avg_composite_score') or 0):.2f} / 5.00" if kpis.get("avg_composite_score") is not None else "—"],
        ],
        colWidths=[180, 120, 180, 120],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), "Helvetica", 8.5),
                ("FONT", (0, 0), (-1, -1), "Helvetica-Bold", 8.5),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b5c6d6")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f4f7fb")),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ]
        )
    )

    threshold_rows = analysis_payload.get("threshold_rows") or []
    threshold_data = [[
        "Vintage",
        "IRR LQ",
        "IRR Median",
        "IRR UQ",
        "IRR Top 5%",
        "MOIC LQ",
        "MOIC Median",
        "MOIC UQ",
        "MOIC Top 5%",
        "DPI LQ",
        "DPI Median",
        "DPI UQ",
        "DPI Top 5%",
    ]]
    for row in threshold_rows:
        threshold_data.append(
            [
                row.get("vintage_year"),
                _fmt_track_pct(row.get("net_irr_lower_quartile")),
                _fmt_track_pct(row.get("net_irr_median")),
                _fmt_track_pct(row.get("net_irr_upper_quartile")),
                _fmt_track_pct(row.get("net_irr_top_5")),
                _fmt_track_multiple(row.get("net_moic_lower_quartile")),
                _fmt_track_multiple(row.get("net_moic_median")),
                _fmt_track_multiple(row.get("net_moic_upper_quartile")),
                _fmt_track_multiple(row.get("net_moic_top_5")),
                _fmt_track_multiple(row.get("net_dpi_lower_quartile")),
                _fmt_track_multiple(row.get("net_dpi_median")),
                _fmt_track_multiple(row.get("net_dpi_upper_quartile")),
                _fmt_track_multiple(row.get("net_dpi_top_5")),
            ]
        )

    threshold_table = Table(
        threshold_data,
        colWidths=[64, 60, 60, 60, 64, 60, 60, 60, 64, 60, 60, 60, 64],
        repeatRows=1,
    )
    threshold_style = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", 7.0),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 7.2),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4d78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#b5c6d6")),
        ("ALIGN", (0, 1), (-1, -1), "RIGHT"),
    ]
    for row_idx in range(1, len(threshold_data)):
        threshold_style.append(
            ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#f9fbfe") if row_idx % 2 else colors.HexColor("#edf4fc"))
        )
    threshold_table.setStyle(TableStyle(threshold_style))

    story = [
        Paragraph(report_title, styles["Heading4"]),
        metadata,
        Spacer(1, 8),
        Paragraph("Fund Benchmarking Table", styles["Heading5"]),
        fund_table,
        PageBreak(),
        Paragraph("Executive Summary", styles["Heading5"]),
        summary_table,
        Spacer(1, 6),
        Paragraph((analysis_payload.get("meta") or {}).get("coverage_note") or "", styles["Normal"]),
        PageBreak(),
        Paragraph("Appendix: Benchmark Threshold Matrix", styles["Heading5"]),
        threshold_table,
    ]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=18,
        rightMargin=18,
        topMargin=16,
        bottomMargin=16,
        title=report_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _build_track_record_pdf(track_record, currency_code=DEFAULT_CURRENCY_CODE, report_title=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    headers = [
        "#",
        "Company",
        "Status",
        "Investment Date",
        "Exit Date",
        "Hold Period",
        "Ownership",
        "% of Total Invested",
        "% of Fund Size",
        "Total Invested",
        "Realized Value",
        "Unrealized Value",
        "Total Value",
        "Gross IRR",
        "Gross MOIC",
        "Realized Gross MOIC",
        "Unrealized Gross MOIC",
    ]
    rows = [headers]
    row_tags = ["header"]

    for fund in track_record.get("funds", []):
        fund_title = fund.get("fund_name") or "Unknown Fund"
        if fund.get("fund_size") is not None:
            fund_title = f"{fund_title} ({format_currency_millions(fund['fund_size'], currency_code=currency_code, show_code=True)})"
        if fund.get("fund_size_conflict"):
            fund_title = f"{fund_title} [fund size conflict]"
        rows.append([fund_title] + [""] * 16)
        row_tags.append("fund_header")

        for row in fund.get("rows", []):
            rows.append(
                [
                    str(row.get("row_num") or ""),
                    row.get("company_name") or "Unknown Company",
                    row.get("status") or "Other",
                    _fmt_track_date(row.get("investment_date")),
                    _fmt_track_date(row.get("exit_date")),
                    _fmt_track_years(row.get("hold_period")),
                    _fmt_track_pct(row.get("ownership_pct")),
                    _fmt_track_pct(row.get("pct_total_invested")),
                    _fmt_track_pct(row.get("pct_fund_size")),
                    _fmt_track_currency(row.get("invested_equity"), currency_code=currency_code),
                    _fmt_track_currency(row.get("realized_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("unrealized_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("total_value"), currency_code=currency_code),
                    _fmt_track_pct(row.get("gross_irr")),
                    _fmt_track_multiple(row.get("gross_moic")),
                    _fmt_track_multiple(row.get("realized_gross_moic")),
                    _fmt_track_multiple(row.get("unrealized_gross_moic")),
                ]
            )
            row_tags.append("detail")

        for rollup in fund.get("status_rollups", []):
            rows.append(
                _track_totals_to_pdf_row(
                    rollup.get("label", "Status Rollup"),
                    rollup.get("totals", {}),
                    include_fund_size=True,
                    currency_code=currency_code,
                )
            )
            row_tags.append("rollup_status")

        for rollup in fund.get("summary_rollups", []):
            rows.append(
                _track_totals_to_pdf_row(
                    rollup.get("label", "Fund Rollup"),
                    rollup.get("totals", {}),
                    include_fund_size=True,
                    currency_code=currency_code,
                )
            )
            row_tags.append("rollup_summary")

        net = fund.get("net_performance", {})
        irr_conflict = (net.get("conflicts") or {}).get("net_irr")
        moic_conflict = (net.get("conflicts") or {}).get("net_moic")
        dpi_conflict = (net.get("conflicts") or {}).get("net_dpi")
        net_irr_val = "N/A" if irr_conflict or net.get("net_irr") is None else _fmt_track_pct(net.get("net_irr"))
        net_moic_val = "N/A" if moic_conflict or net.get("net_moic") is None else _fmt_track_multiple(net.get("net_moic"))
        net_dpi_val = "N/A" if dpi_conflict or net.get("net_dpi") is None else _fmt_track_multiple(net.get("net_dpi"))

        rows.append(
            [
                "",
                f"{fund.get('fund_name', 'Fund')} Net Performance",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "IRR:",
                net_irr_val,
                "Net MOIC:",
                net_moic_val,
            ]
        )
        row_tags.append("net")
        rows.append([""] * 13 + ["DPI:", net_dpi_val, "", ""])
        row_tags.append("net")
        rows.append([""] * 17)
        row_tags.append("gap")

    rows.append(["All Funds Summary"] + [""] * 16)
    row_tags.append("overall_header")
    for rollup in track_record.get("overall", {}).get("status_rollups", []):
        rows.append(
            _track_totals_to_pdf_row(
                rollup.get("label", "Status Rollup"),
                rollup.get("totals", {}),
                include_fund_size=False,
                currency_code=currency_code,
            )
        )
        row_tags.append("rollup_status")
    for rollup in track_record.get("overall", {}).get("summary_rollups", []):
        rows.append(
            _track_totals_to_pdf_row(
                rollup.get("label", "Overall Rollup"),
                rollup.get("totals", {}),
                include_fund_size=False,
                currency_code=currency_code,
            )
        )
        row_tags.append("rollup_overall")

    page_width, _ = landscape(A3)
    left_margin = right_margin = 18
    available_width = page_width - left_margin - right_margin
    base_widths = [18, 110, 72, 52, 52, 42, 52, 64, 60, 65, 65, 65, 65, 52, 52, 60, 64]
    scale = available_width / sum(base_widths)
    col_widths = [w * scale for w in base_widths]

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", 6.4),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 6.6),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4d78")),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (5, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9eb3c5")),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]

    detail_alt = False
    for idx, tag in enumerate(row_tags):
        if idx == 0:
            continue
        if tag in {"fund_header", "overall_header"}:
            style_cmds.extend(
                [
                    ("SPAN", (0, idx), (-1, idx)),
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#8eb0cf") if tag == "fund_header" else colors.HexColor("#6f97bd")),
                    ("TEXTCOLOR", (0, idx), (-1, idx), colors.HexColor("#0d2740")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 7.0),
                    ("ALIGN", (0, idx), (-1, idx), "LEFT"),
                    ("LINEABOVE", (0, idx), (-1, idx), 0.6, colors.HexColor("#4f6e89")),
                    ("LINEBELOW", (0, idx), (-1, idx), 0.6, colors.HexColor("#4f6e89")),
                ]
            )
            detail_alt = False
        elif tag == "detail":
            style_cmds.append(
                ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#e5e5e5") if detail_alt else colors.HexColor("#f1f1f1"))
            )
            detail_alt = not detail_alt
        elif tag == "rollup_status":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#9dbbda")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.4),
                    ("LINEABOVE", (0, idx), (-1, idx), 0.5, colors.HexColor("#557894")),
                ]
            )
        elif tag == "rollup_summary":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#8db0d0")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.4),
                    ("LINEABOVE", (0, idx), (-1, idx), 0.5, colors.HexColor("#557894")),
                ]
            )
        elif tag == "rollup_overall":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#7ca4cb")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.4),
                    ("LINEABOVE", (0, idx), (-1, idx), 0.5, colors.HexColor("#557894")),
                ]
            )
        elif tag == "net":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#84a9cc")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.4),
                ]
            )
        elif tag == "gap":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#c8d9e8")),
                    ("LINEABOVE", (0, idx), (-1, idx), 0, colors.white),
                    ("LINEBELOW", (0, idx), (-1, idx), 0, colors.white),
                ]
            )

    table.setStyle(TableStyle(style_cmds))

    styles = getSampleStyleSheet()
    final_title = report_title or "Deal Level Track Record (Print-Ready PDF)"
    title = Paragraph(final_title, styles["Heading4"])
    generated = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=16,
        bottomMargin=16,
        title=final_title,
    )
    doc.build([title, generated, Spacer(1, 8), table])
    return buffer.getvalue()


def _resolve_credit_analysis_as_of_date(loans, fund_performance=None):
    report_dates = [
        getattr(row, "report_date", None)
        for row in (fund_performance or {}).values()
        if getattr(row, "report_date", None) is not None
    ]
    if report_dates:
        return max(report_dates)

    explicit_as_of_dates = [
        getattr(loan, "as_of_date", None)
        for loan in loans
        if getattr(loan, "as_of_date", None) is not None
    ]
    if explicit_as_of_dates:
        return max(explicit_as_of_dates)

    exit_dates = [
        getattr(loan, "exit_date", None)
        for loan in loans
        if getattr(loan, "exit_date", None) is not None
    ]
    if exit_dates:
        return max(exit_dates)

    close_dates = [
        getattr(loan, "close_date", None)
        for loan in loans
        if getattr(loan, "close_date", None) is not None
    ]
    if close_dates:
        return max(close_dates)

    return date.today()


def _credit_pdf_meta_line(as_of_date, currency_code=DEFAULT_CURRENCY_CODE, extras=None):
    parts = [
        f"As of {_as_of_ymd(as_of_date)}",
        currency_unit_label(currency_code),
    ]
    parts.extend([extra for extra in (extras or []) if extra])
    return " | ".join(parts)


def _build_empty_report_pdf(
    report_title,
    *,
    as_of_date=None,
    currency_code=DEFAULT_CURRENCY_CODE,
    note="No rows available for export.",
    pagesize=None,
):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = BytesIO()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize or landscape(letter),
        leftMargin=18,
        rightMargin=18,
        topMargin=16,
        bottomMargin=16,
        title=report_title,
    )
    story = [
        Paragraph(report_title, styles["Heading4"]),
        Paragraph(_credit_pdf_meta_line(as_of_date, currency_code), styles["Normal"]),
        Spacer(1, 8),
        Paragraph(note, styles["Normal"]),
    ]
    doc.build(story)
    return buffer.getvalue()


def _build_pdf_table(
    table_data,
    *,
    col_widths,
    numeric_cols=None,
    repeat_rows=1,
    header_rows=1,
    font_size=7.0,
    emphasized_rows=None,
    span_rows=None,
):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    numeric_cols = set(numeric_cols or [])
    emphasized_rows = emphasized_rows or {}
    span_rows = span_rows or {}

    table = Table(table_data, colWidths=col_widths, repeatRows=repeat_rows)
    style_cmds = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", font_size),
        ("FONT", (0, 0), (-1, header_rows - 1), "Helvetica-Bold", font_size + 0.2),
        ("BACKGROUND", (0, 0), (-1, header_rows - 1), colors.HexColor("#1f4d78")),
        ("TEXTCOLOR", (0, 0), (-1, header_rows - 1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#b5c6d6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]

    for idx in numeric_cols:
        style_cmds.append(("ALIGN", (idx, header_rows), (idx, -1), "RIGHT"))

    for row_idx in range(header_rows, len(table_data)):
        if row_idx in span_rows:
            bg = span_rows[row_idx]
            style_cmds.extend(
                [
                    ("SPAN", (0, row_idx), (-1, row_idx)),
                    ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor(bg)),
                    ("TEXTCOLOR", (0, row_idx), (-1, row_idx), colors.HexColor("#0d2740")),
                    ("FONT", (0, row_idx), (-1, row_idx), "Helvetica-Bold", font_size + 0.2),
                    ("ALIGN", (0, row_idx), (-1, row_idx), "LEFT"),
                ]
            )
            continue

        if row_idx in emphasized_rows:
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor(emphasized_rows[row_idx])),
                    ("FONT", (0, row_idx), (-1, row_idx), "Helvetica-Bold", font_size),
                ]
            )
            continue

        style_cmds.append(
            (
                "BACKGROUND",
                (0, row_idx),
                (-1, row_idx),
                colors.HexColor("#f8fbff") if (row_idx - header_rows) % 2 == 0 else colors.HexColor("#edf4fc"),
            )
        )

    table.setStyle(TableStyle(style_cmds))
    return table


def _credit_track_totals_to_pdf_row(label, totals, include_fund_size=True, currency_code=DEFAULT_CURRENCY_CODE):
    return [
        "",
        label,
        "",
        "",
        "",
        _fmt_track_years(totals.get("hold_period")),
        _fmt_track_pct(totals.get("pct_total_invested")),
        _fmt_track_pct(totals.get("pct_fund_size")) if include_fund_size else "—",
        _fmt_track_currency(totals.get("invested_equity"), currency_code=currency_code),
        _fmt_track_currency(totals.get("realized_value"), currency_code=currency_code),
        _fmt_track_currency(totals.get("unrealized_value"), currency_code=currency_code),
        _fmt_track_currency(
            totals.get("unrealized_warrant_equity_value"), currency_code=currency_code
        ),
        _fmt_track_currency(totals.get("total_value"), currency_code=currency_code),
        _fmt_track_pct(totals.get("gross_irr")),
        _fmt_track_multiple(totals.get("gross_moic")),
        _fmt_track_multiple(totals.get("realized_gross_moic")),
        _fmt_track_multiple(totals.get("unrealized_gross_moic")),
    ]


def _build_credit_track_record_pdf(
    track_record,
    *,
    currency_code=DEFAULT_CURRENCY_CODE,
    report_title=None,
    as_of_date=None,
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    headers = [
        "#",
        "Company",
        "Status",
        "Close Date",
        "Exit Date",
        "Hold Period",
        "% of Fund Invested",
        "% of Fund Size",
        "Current Invested Capital",
        "Realized Value",
        "Unrealized Value",
        "Unrealized Equity Value",
        "Total Value",
        "Gross IRR",
        "Gross MOIC",
        "Realized MOIC",
        "Unrealized MOIC",
    ]
    rows = [headers]
    row_tags = ["header"]

    for fund in track_record.get("funds", []):
        fund_title = fund.get("fund_name") or "Unknown Fund"
        if fund.get("vintage_year") is not None:
            fund_title = f"{fund_title} · Vintage {fund['vintage_year']}"
        if fund.get("fund_size") is not None:
            fund_title = f"{fund_title} ({format_currency_millions(fund['fund_size'], currency_code=currency_code, show_code=True)})"
        rows.append([fund_title] + [""] * 16)
        row_tags.append("fund_header")

        for row in fund.get("rows", []):
            company_label = row.get("company_name") or "Unknown Company"
            if row.get("sector"):
                company_label = f"{company_label} · {row['sector']}"
            rows.append(
                [
                    str(row.get("row_num") or ""),
                    company_label,
                    row.get("status") or "—",
                    _fmt_track_date(row.get("investment_date")),
                    _fmt_track_date(row.get("exit_date")),
                    _fmt_track_years(row.get("hold_period")),
                    _fmt_track_pct(row.get("pct_total_invested")),
                    _fmt_track_pct(row.get("pct_fund_size")),
                    _fmt_track_currency(row.get("current_invested_capital"), currency_code=currency_code),
                    _fmt_track_currency(row.get("realized_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("unrealized_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("unrealized_warrant_equity_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("total_value"), currency_code=currency_code),
                    _fmt_track_pct(row.get("gross_irr")),
                    _fmt_track_multiple(row.get("gross_moic")),
                    _fmt_track_multiple(row.get("realized_gross_moic")),
                    _fmt_track_multiple(row.get("unrealized_gross_moic")),
                ]
            )
            row_tags.append("detail")

        for rollup in fund.get("status_rollups", []):
            rows.append(
                _credit_track_totals_to_pdf_row(
                    rollup.get("label", "Status Rollup"),
                    rollup.get("totals", {}),
                    include_fund_size=True,
                    currency_code=currency_code,
                )
            )
            row_tags.append("rollup_status")

        for rollup in fund.get("summary_rollups", []):
            rows.append(
                _credit_track_totals_to_pdf_row(
                    rollup.get("label", "Fund Rollup"),
                    rollup.get("totals", {}),
                    include_fund_size=True,
                    currency_code=currency_code,
                )
            )
            row_tags.append("rollup_summary")

        net = fund.get("net_performance", {})
        net_irr_val = "N/A" if net.get("net_irr") is None else _fmt_track_pct(net.get("net_irr"))
        net_tvpi_val = "N/A" if net.get("net_tvpi") is None else _fmt_track_multiple(net.get("net_tvpi"))
        net_dpi_val = "N/A" if net.get("net_dpi") is None else _fmt_track_multiple(net.get("net_dpi"))

        rows.append(
            ["", f"{fund.get('fund_name', 'Fund')} Net Performance"] + [""] * 11 + ["Net IRR:", net_irr_val, "Net TVPI:", net_tvpi_val]
        )
        row_tags.append("net")
        rows.append([""] * 13 + ["Net DPI:", net_dpi_val, "", ""])
        row_tags.append("net")
        rows.append([""] * 17)
        row_tags.append("gap")

    rows.append(["All Funds Summary"] + [""] * 16)
    row_tags.append("overall_header")
    for rollup in track_record.get("overall", {}).get("status_rollups", []):
        rows.append(
            _credit_track_totals_to_pdf_row(
                rollup.get("label", "Status Rollup"),
                rollup.get("totals", {}),
                include_fund_size=False,
                currency_code=currency_code,
            )
        )
        row_tags.append("rollup_status")
    for rollup in track_record.get("overall", {}).get("summary_rollups", []):
        rows.append(
            _credit_track_totals_to_pdf_row(
                rollup.get("label", "Overall Rollup"),
                rollup.get("totals", {}),
                include_fund_size=False,
                currency_code=currency_code,
            )
        )
        row_tags.append("rollup_overall")

    page_width, _ = landscape(A3)
    left_margin = right_margin = 18
    available_width = page_width - left_margin - right_margin
    base_widths = [18, 122, 72, 52, 52, 42, 58, 56, 68, 62, 62, 68, 66, 52, 52, 56, 60]
    scale = available_width / sum(base_widths)
    col_widths = [w * scale for w in base_widths]

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", 6.2),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 6.4),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4d78")),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9eb3c5")),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]

    detail_alt = False
    for idx, tag in enumerate(row_tags):
        if idx == 0:
            continue
        if tag in {"fund_header", "overall_header"}:
            style_cmds.extend(
                [
                    ("SPAN", (0, idx), (-1, idx)),
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#8eb0cf") if tag == "fund_header" else colors.HexColor("#6f97bd")),
                    ("TEXTCOLOR", (0, idx), (-1, idx), colors.HexColor("#0d2740")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.8),
                    ("ALIGN", (0, idx), (-1, idx), "LEFT"),
                ]
            )
            detail_alt = False
        elif tag == "detail":
            style_cmds.append(
                ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#f8fbff") if detail_alt else colors.HexColor("#edf4fc"))
            )
            detail_alt = not detail_alt
        elif tag == "rollup_status":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#d8e5f0")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.2),
                ]
            )
        elif tag == "rollup_summary":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#c7d9ea")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.2),
                ]
            )
        elif tag == "rollup_overall":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#b5cee5")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.2),
                ]
            )
        elif tag == "net":
            style_cmds.extend(
                [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#dce8f3")),
                    ("FONT", (0, idx), (-1, idx), "Helvetica-Bold", 6.2),
                ]
            )
        elif tag == "gap":
            style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#ffffff")))

    table.setStyle(TableStyle(style_cmds))

    styles = getSampleStyleSheet()
    final_title = report_title or "Credit Track Record"
    story = [
        Paragraph(final_title, styles["Heading4"]),
        Paragraph(_credit_pdf_meta_line(as_of_date, currency_code), styles["Normal"]),
        Spacer(1, 8),
        table,
    ]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=16,
        bottomMargin=16,
        title=final_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _build_credit_concentration_pdf(
    analysis_payload,
    *,
    currency_code=DEFAULT_CURRENCY_CODE,
    report_title,
    as_of_date=None,
):
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

    if not analysis_payload.get("loan_count"):
        return _build_empty_report_pdf(
            report_title,
            as_of_date=as_of_date,
            currency_code=currency_code,
            pagesize=landscape(legal),
        )

    styles = getSampleStyleSheet()
    story = [
        Paragraph(report_title, styles["Heading4"]),
        Paragraph(
            _credit_pdf_meta_line(
                as_of_date,
                currency_code,
                extras=["Concentration basis: Total Unrealized Value"],
            ),
            styles["Normal"],
        ),
        Spacer(1, 8),
    ]

    summary_table = _build_pdf_table(
        [
            ["Metric", "Value", "Metric", "Value"],
            ["Loans", str(analysis_payload.get("loan_count") or 0), "Funds", str(analysis_payload.get("fund_count") or 0)],
            ["Invested", _fmt_track_currency(analysis_payload.get("total_invested"), currency_code=currency_code), "Realized Value", _fmt_track_currency(analysis_payload.get("total_realized_value"), currency_code=currency_code)],
            ["Unrealized Loan Value", _fmt_track_currency(analysis_payload.get("total_unrealized_loan_value"), currency_code=currency_code), "Unrealized Equity Value", _fmt_track_currency(analysis_payload.get("total_unrealized_equity_value"), currency_code=currency_code)],
            ["Total Unrealized Value", _fmt_track_currency(analysis_payload.get("total_unrealized_value"), currency_code=currency_code), "Total Value", _fmt_track_currency(analysis_payload.get("total_value"), currency_code=currency_code)],
        ],
        col_widths=[145, 125, 145, 125],
        numeric_cols=[1, 3],
        font_size=7.5,
    )
    story.extend([Paragraph("Portfolio Summary", styles["Heading5"]), summary_table, Spacer(1, 10)])

    top_10 = analysis_payload.get("top_10") or []
    top_table = _build_pdf_table(
        [["Company", "% Portfolio Value", "Invested", "Realized", "Unrealized", "Unrealized Equity", "Total Unrealized", "Total Value"]]
        + [
            [
                row.get("company") or "Unknown Company",
                _fmt_track_pct(row.get("pct")),
                _fmt_track_currency(row.get("invested"), currency_code=currency_code),
                _fmt_track_currency(row.get("realized_value"), currency_code=currency_code),
                _fmt_track_currency(row.get("unrealized_value"), currency_code=currency_code),
                _fmt_track_currency(row.get("unrealized_equity_value"), currency_code=currency_code),
                _fmt_track_currency(row.get("total_unrealized_value"), currency_code=currency_code),
                _fmt_track_currency(row.get("total_value"), currency_code=currency_code),
            ]
            for row in top_10
        ],
        col_widths=[150, 72, 72, 72, 72, 78, 84, 72],
        numeric_cols=[1, 2, 3, 4, 5, 6, 7],
        font_size=6.9,
    )
    story.extend([Paragraph("Top Single-Name Exposures", styles["Heading5"]), top_table])

    breakdown_specs = [
        ("By Sector", analysis_payload.get("by_sector") or []),
        ("By Geography", analysis_payload.get("by_geography") or []),
        ("By Sponsor", analysis_payload.get("by_sponsor") or []),
        ("By Security Type", analysis_payload.get("by_security") or []),
        ("By Rating", analysis_payload.get("by_rating") or []),
    ]
    if analysis_payload.get("has_sourcing_data"):
        breakdown_specs.append(("By Sourcing Channel", analysis_payload.get("by_sourcing") or []))
    if analysis_payload.get("has_public_data"):
        breakdown_specs.append(("By Public / Private", analysis_payload.get("by_public") or []))

    for idx, (label, rows) in enumerate(breakdown_specs, start=1):
        story.extend([PageBreak(), Paragraph(label, styles["Heading5"]), Spacer(1, 4)])
        table_rows = [["Group", "Loans", "% Portfolio Value", "Invested", "Realized", "Unrealized", "Unrealized Equity", "Total Unrealized", "Total Value"]]
        for row in rows[:15]:
            table_rows.append(
                [
                    row.get("name") or "Unknown",
                    str(row.get("count") or 0),
                    _fmt_track_pct(row.get("pct")),
                    _fmt_track_currency(row.get("invested"), currency_code=currency_code),
                    _fmt_track_currency(row.get("realized_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("unrealized_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("unrealized_equity_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("total_unrealized_value"), currency_code=currency_code),
                    _fmt_track_currency(row.get("total_value"), currency_code=currency_code),
                ]
            )
        story.append(
            _build_pdf_table(
                table_rows,
                col_widths=[145, 44, 72, 68, 68, 68, 74, 80, 70],
                numeric_cols=[1, 2, 3, 4, 5, 6, 7, 8],
                font_size=6.8,
            )
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(legal),
        leftMargin=18,
        rightMargin=18,
        topMargin=16,
        bottomMargin=16,
        title=report_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _build_credit_fundamentals_pdf(
    analysis_payload,
    *,
    currency_code=DEFAULT_CURRENCY_CODE,
    report_title,
    as_of_date=None,
):
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

    if not analysis_payload.get("loan_count"):
        return _build_empty_report_pdf(
            report_title,
            as_of_date=as_of_date,
            currency_code=currency_code,
            pagesize=landscape(legal),
        )

    styles = getSampleStyleSheet()
    exit_current_label = analysis_payload.get("exit_current_label") or "Exit / Current"
    story = [
        Paragraph(report_title, styles["Heading4"]),
        Paragraph(
            _credit_pdf_meta_line(
                as_of_date,
                currency_code,
                extras=["Weighted averages use Current Invested Capital"],
            ),
            styles["Normal"],
        ),
        Spacer(1, 8),
        Paragraph("Portfolio Summary", styles["Heading5"]),
    ]

    summary_rows = [["Metric", "Paired Deals", "Wtd Avg Entry", f"Wtd Avg {exit_current_label}", "Wtd Avg Δ", "Avg Entry", f"Avg {exit_current_label}", "Avg Δ"]]
    for metric in analysis_payload.get("summary_metrics") or []:
        kind = metric.get("kind")
        summary_rows.append(
            [
                metric.get("label") or "Metric",
                str(metric.get("paired_count") or 0),
                _render_pdf_metric(metric.get("weighted_average_entry"), kind, currency_code=currency_code),
                _render_pdf_metric(metric.get("weighted_average_exit_current"), kind, currency_code=currency_code),
                _render_pdf_metric(metric.get("weighted_average_delta"), kind, currency_code=currency_code),
                _render_pdf_metric(metric.get("average_entry"), kind, currency_code=currency_code),
                _render_pdf_metric(metric.get("average_exit_current"), kind, currency_code=currency_code),
                _render_pdf_metric(metric.get("average_delta"), kind, currency_code=currency_code),
            ]
        )
    story.append(
        _build_pdf_table(
            summary_rows,
            col_widths=[120, 54, 78, 86, 78, 74, 82, 74],
            numeric_cols=[1, 2, 3, 4, 5, 6, 7],
            font_size=6.9,
        )
    )

    story.extend([Spacer(1, 10), Paragraph("Loan Term by Fund", styles["Heading5"])])
    term_rows = [["Fund", "Loans", "Loans with Term", "Current Invested Capital", "Wtd Avg Term", "Avg Term"]]
    for row in analysis_payload.get("term_by_fund_rows") or []:
        term_rows.append(
            [
                row.get("fund_name") or "Unknown Fund",
                str(row.get("loan_count") or 0),
                str(row.get("term_count") or 0),
                _fmt_track_currency(row.get("total_current_invested_capital"), currency_code=currency_code),
                _fmt_track_years(row.get("weighted_average_term_years")),
                _fmt_track_years(row.get("average_term_years")),
            ]
        )
    story.append(
        _build_pdf_table(
            term_rows,
            col_widths=[170, 46, 70, 108, 70, 70],
            numeric_cols=[1, 2, 3, 4, 5],
            font_size=7.0,
        )
    )

    for idx, table in enumerate(analysis_payload.get("fund_metric_tables") or []):
        story.extend([PageBreak(), Paragraph(f"{table.get('label', 'Metric')} by Fund", styles["Heading5"])])
        table_rows = [["Fund", "Loans", "Current Invested Capital", "Paired Deals", "Wtd Avg Entry", f"Wtd Avg {exit_current_label}", "Wtd Avg Δ", "Avg Entry", f"Avg {exit_current_label}", "Avg Δ"]]
        for row in table.get("rows") or []:
            table_rows.append(
                [
                    row.get("fund_name") or "Unknown Fund",
                    str(row.get("loan_count") or 0),
                    _fmt_track_currency(row.get("total_current_invested_capital"), currency_code=currency_code),
                    str(row.get("paired_count") or 0),
                    _render_pdf_metric(row.get("weighted_average_entry"), table.get("kind"), currency_code=currency_code),
                    _render_pdf_metric(row.get("weighted_average_exit_current"), table.get("kind"), currency_code=currency_code),
                    _render_pdf_metric(row.get("weighted_average_delta"), table.get("kind"), currency_code=currency_code),
                    _render_pdf_metric(row.get("average_entry"), table.get("kind"), currency_code=currency_code),
                    _render_pdf_metric(row.get("average_exit_current"), table.get("kind"), currency_code=currency_code),
                    _render_pdf_metric(row.get("average_delta"), table.get("kind"), currency_code=currency_code),
                ]
            )
        story.append(
            _build_pdf_table(
                table_rows,
                col_widths=[130, 40, 100, 50, 74, 78, 74, 70, 74, 70],
                numeric_cols=[1, 2, 3, 4, 5, 6, 7, 8, 9],
                font_size=6.6,
            )
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(legal),
        leftMargin=18,
        rightMargin=18,
        topMargin=16,
        bottomMargin=16,
        title=report_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _render_pdf_metric(value, kind, *, currency_code=DEFAULT_CURRENCY_CODE):
    if kind == "currency":
        return _fmt_track_currency(value, currency_code=currency_code)
    if kind == "percent":
        return _fmt_track_pct(value)
    if kind == "multiple":
        return _fmt_track_multiple(value)
    if value is None:
        return "N/A"
    return f"{value:.1f}"


def _build_credit_pricing_trends_pdf(
    analysis_payload,
    *,
    currency_code=DEFAULT_CURRENCY_CODE,
    report_title,
    as_of_date=None,
):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

    if not analysis_payload.get("loan_count"):
        return _build_empty_report_pdf(
            report_title,
            as_of_date=as_of_date,
            currency_code=currency_code,
            pagesize=landscape(letter),
        )

    styles = getSampleStyleSheet()
    story = [
        Paragraph(report_title, styles["Heading4"]),
        Paragraph(
            _credit_pdf_meta_line(
                as_of_date,
                currency_code,
                extras=[
                    f"Time Group: {analysis_payload.get('time_group_label')}",
                    f"Dimension: {analysis_payload.get('primary_dim_label')}",
                ],
            ),
            styles["Normal"],
        ),
        Spacer(1, 8),
        Paragraph("Summary", styles["Heading5"]),
    ]

    summary = analysis_payload.get("summary") or {}
    summary_rows = [
        ["Metric", "Value", "Metric", "Value"],
        ["Loans", str(analysis_payload.get("loan_count") or 0), "Funds", str(analysis_payload.get("fund_count") or 0)],
        ["Weighted Loans", str(analysis_payload.get("weighted_loan_count") or 0), "Current Invested Capital", _fmt_track_currency(analysis_payload.get("total_current_invested_capital"), currency_code=currency_code)],
        ["Wtd Avg Coupon", _fmt_track_pct(summary.get("weighted_average_coupon_rate")), "Avg Coupon", _fmt_track_pct(summary.get("average_coupon_rate"))],
        ["Wtd Avg Floor", _fmt_track_pct(summary.get("weighted_average_floor_rate")), "Avg Floor", _fmt_track_pct(summary.get("average_floor_rate"))],
        ["Wtd Avg Upfront Fee", _fmt_track_pct(summary.get("weighted_average_upfront_fee")), "Avg Upfront Fee", _fmt_track_pct(summary.get("average_upfront_fee"))],
    ]
    story.append(_build_pdf_table(summary_rows, col_widths=[150, 120, 150, 120], numeric_cols=[1, 3], font_size=7.4))

    def _pricing_table(title, rows, label_key):
        story.extend([Spacer(1, 10), Paragraph(title, styles["Heading5"])])
        table_rows = [[label_key, "Loans", "Current Invested", "Wtd Avg Coupon", "Avg Coupon", "Wtd Avg Floor", "Avg Floor", "Wtd Avg Upfront", "Avg Upfront Fee"]]
        for row in rows:
            table_rows.append(
                [
                    row.get("label") or row.get("fund_name") or row.get("dimension_value") or "Unknown",
                    str(row.get("loan_count") or 0),
                    _fmt_track_currency(row.get("total_current_invested_capital"), currency_code=currency_code),
                    _fmt_track_pct(row.get("weighted_average_coupon_rate")),
                    _fmt_track_pct(row.get("average_coupon_rate")),
                    _fmt_track_pct(row.get("weighted_average_floor_rate")),
                    _fmt_track_pct(row.get("average_floor_rate")),
                    _fmt_track_pct(row.get("weighted_average_upfront_fee")),
                    _fmt_track_pct(row.get("average_upfront_fee")),
                ]
            )
        story.append(
            _build_pdf_table(
                table_rows,
                col_widths=[120, 40, 92, 66, 62, 66, 62, 68, 68],
                numeric_cols=[1, 2, 3, 4, 5, 6, 7, 8],
                font_size=6.8,
            )
        )

    _pricing_table(f"Pricing Trend by {analysis_payload.get('time_group_label')}", analysis_payload.get("time_rows") or [], analysis_payload.get("time_group_label") or "Period")
    story.append(PageBreak())
    _pricing_table("By Fund", analysis_payload.get("fund_rows") or [], "Fund")
    story.append(PageBreak())
    _pricing_table(f"By {analysis_payload.get('primary_dim_label')}", analysis_payload.get("dimension_rows") or [], analysis_payload.get("primary_dim_label") or "Dimension")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=18,
        rightMargin=18,
        topMargin=16,
        bottomMargin=16,
        title=report_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _build_credit_underwrite_outcome_pdf(
    analysis_payload,
    *,
    currency_code=DEFAULT_CURRENCY_CODE,
    report_title,
    as_of_date=None,
):
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

    styles = getSampleStyleSheet()
    story = [
        Paragraph(report_title, styles["Heading4"]),
        Paragraph(_credit_pdf_meta_line(as_of_date, currency_code), styles["Normal"]),
        Spacer(1, 8),
        Paragraph(analysis_payload.get("coverage_note") or "", styles["Normal"]),
        Spacer(1, 8),
    ]

    summary = analysis_payload.get("summary") or {}
    summary_rows = [
        ["Metric", "Value", "Metric", "Value"],
        ["Compared Loans", str(summary.get("loan_count") or 0), "Funds", str(summary.get("fund_count") or 0)],
        ["Weighted Loans", str(summary.get("weighted_loan_count") or 0), "Current Invested Capital", _fmt_track_currency(summary.get("total_current_invested_capital"), currency_code=currency_code)],
        ["Wtd Est. IRR", _fmt_track_pct(summary.get("weighted_estimated_irr")), "Wtd Actual Gross IRR", _fmt_track_pct(summary.get("weighted_actual_gross_irr"))],
        ["Wtd Delta IRR", _fmt_track_pct(summary.get("weighted_delta_irr")), "Avg Delta IRR", _fmt_track_pct(summary.get("average_delta_irr"))],
        ["Hit Rate", _fmt_track_pct(summary.get("hit_rate")), "Miss Count", str(summary.get("miss_count") or 0)],
        ["Missing Estimate", str(summary.get("missing_estimate_count") or 0), "Missing Actual", str(summary.get("missing_actual_count") or 0)],
    ]
    story.extend([Paragraph("Summary", styles["Heading5"]), _build_pdf_table(summary_rows, col_widths=[150, 118, 150, 118], numeric_cols=[1, 3], font_size=7.4)])

    fund_rows = analysis_payload.get("fund_rows") or []
    if fund_rows:
        story.extend([Spacer(1, 10), Paragraph("By Fund", styles["Heading5"])])
        fund_table_rows = [["Fund", "Loans", "Current Invested", "Wtd Est. IRR", "Wtd Actual IRR", "Wtd Delta IRR", "Hit Rate", "Miss Count", "Worst Company"]]
        for row in fund_rows:
            fund_table_rows.append(
                [
                    row.get("fund_name") or "Unknown Fund",
                    str(row.get("loan_count") or 0),
                    _fmt_track_currency(row.get("total_current_invested_capital"), currency_code=currency_code),
                    _fmt_track_pct(row.get("weighted_estimated_irr")),
                    _fmt_track_pct(row.get("weighted_actual_gross_irr")),
                    _fmt_track_pct(row.get("weighted_delta_irr")),
                    _fmt_track_pct(row.get("hit_rate")),
                    str(row.get("miss_count") or 0),
                    row.get("worst_company_name") or "—",
                ]
            )
        story.append(
            _build_pdf_table(
                fund_table_rows,
                col_widths=[110, 40, 94, 64, 70, 70, 54, 48, 122],
                numeric_cols=[1, 2, 3, 4, 5, 6, 7],
                font_size=6.7,
            )
        )

    worst_rows = analysis_payload.get("worst_rows") or []
    if worst_rows:
        story.extend([PageBreak(), Paragraph("Worst IRR Misses", styles["Heading5"])])
        worst_table_rows = [["Company", "Fund", "Sector", "Status", "Current Invested", "Est. IRR", "Actual IRR", "Delta IRR", "Gross MOIC", "Total Value"]]
        for row in worst_rows:
            worst_table_rows.append(
                [
                    row.get("company_name") or "Unknown Company",
                    row.get("fund_name") or "Unknown Fund",
                    row.get("sector") or "Unknown",
                    row.get("status") or "—",
                    _fmt_track_currency(row.get("current_invested_capital"), currency_code=currency_code),
                    _fmt_track_pct(row.get("estimated_irr_at_entry")),
                    _fmt_track_pct(row.get("actual_gross_irr")),
                    _fmt_track_pct(row.get("delta_irr")),
                    _fmt_track_multiple(row.get("gross_moic")),
                    _fmt_track_currency(row.get("total_value"), currency_code=currency_code),
                ]
            )
        story.append(
            _build_pdf_table(
                worst_table_rows,
                col_widths=[120, 92, 70, 72, 86, 58, 58, 58, 56, 78],
                numeric_cols=[4, 5, 6, 7, 8, 9],
                font_size=6.7,
            )
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(legal),
        leftMargin=18,
        rightMargin=18,
        topMargin=16,
        bottomMargin=16,
        title=report_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _build_credit_data_cuts_summary_pdf(
    all_cuts,
    *,
    currency_code=DEFAULT_CURRENCY_CODE,
    report_title,
    as_of_date=None,
):
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

    styles = getSampleStyleSheet()
    story = [
        Paragraph(report_title, styles["Heading4"]),
        Paragraph(_credit_pdf_meta_line(as_of_date, currency_code), styles["Normal"]),
        Spacer(1, 8),
    ]

    dim_keys = list(all_cuts.keys())
    if not dim_keys:
        return _build_empty_report_pdf(
            report_title,
            as_of_date=as_of_date,
            currency_code=currency_code,
            note="No credit data-cut dimensions have populated qualitative data for export.",
            pagesize=landscape(legal),
        )

    for idx, dim_key in enumerate(dim_keys):
        payload = all_cuts.get(dim_key) or {}
        groups = payload.get("groups") or []
        totals = payload.get("totals") or {}
        dim_label = CREDIT_DIMENSION_LABELS.get(dim_key, dim_key)

        if idx > 0:
            story.append(PageBreak())
        story.extend([Paragraph(f"Performance by {dim_label}", styles["Heading5"]), Spacer(1, 4)])

        table_rows = [[
            dim_label,
            "Loans",
            "Invested",
            "Realized",
            "Unrealized",
            "Total Value",
            "Wtd MOIC",
            "Wtd IRR",
            "Wtd Coupon",
            "Entry LTV",
            "Entry Coverage",
            "Entry Cushion",
            "Hold Period",
        ]]
        for group in groups[:15]:
            table_rows.append(
                [
                    group.get("label") or "Unknown",
                    str(group.get("loan_count") or 0),
                    _fmt_track_currency(group.get("invested"), currency_code=currency_code),
                    _fmt_track_currency(group.get("realized_value"), currency_code=currency_code),
                    _fmt_track_currency(group.get("unrealized_value"), currency_code=currency_code),
                    _fmt_track_currency(group.get("total_value"), currency_code=currency_code),
                    _fmt_track_multiple(group.get("weighted_moic")),
                    _fmt_track_pct(group.get("weighted_irr")),
                    _fmt_track_pct(group.get("weighted_yield")),
                    _fmt_track_pct(group.get("weighted_ltv")),
                    _fmt_track_multiple(group.get("weighted_entry_coverage_ratio")),
                    _fmt_track_pct(group.get("weighted_entry_equity_cushion")),
                    _fmt_track_years(group.get("weighted_hold_years")),
                ]
            )
        table_rows.append(
            [
                "Total",
                str(totals.get("loan_count") or 0),
                _fmt_track_currency(totals.get("invested"), currency_code=currency_code),
                _fmt_track_currency(totals.get("realized_value"), currency_code=currency_code),
                _fmt_track_currency(totals.get("unrealized_value"), currency_code=currency_code),
                _fmt_track_currency(totals.get("total_value"), currency_code=currency_code),
                _fmt_track_multiple(totals.get("weighted_moic")),
                _fmt_track_pct(totals.get("weighted_irr")),
                _fmt_track_pct(totals.get("weighted_yield")),
                _fmt_track_pct(totals.get("weighted_ltv")),
                _fmt_track_multiple(totals.get("weighted_entry_coverage_ratio")),
                _fmt_track_pct(totals.get("weighted_entry_equity_cushion")),
                _fmt_track_years(totals.get("weighted_hold_years")),
            ]
        )
        story.append(
            _build_pdf_table(
                table_rows,
                col_widths=[124, 40, 66, 66, 66, 68, 54, 54, 54, 54, 58, 54, 50],
                numeric_cols=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
                font_size=6.2,
                emphasized_rows={len(table_rows) - 1: "#d8e5f0"},
            )
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(legal),
        leftMargin=18,
        rightMargin=18,
        topMargin=16,
        bottomMargin=16,
        title=report_title,
    )
    doc.build(story)
    return buffer.getvalue()


def _credit_default_benchmark_asset_class(team_id):
    benchmark_asset_classes = _benchmark_asset_classes_for_team(team_id)
    benchmark_session_key = "selected_benchmark_asset_class"
    current_benchmark_asset_class = (session.get(benchmark_session_key, "") or "").strip()
    if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
        current_benchmark_asset_class = ""
        session[benchmark_session_key] = ""
    return current_benchmark_asset_class


def _build_credit_pdf_export_context():
    from peqa.services.credit_filtering import build_credit_analysis_context

    membership = _require_team_scope()
    active_firm = _resolve_active_firm_for_team()
    if active_firm is None:
        return None

    ctx = build_credit_analysis_context(
        team_id=membership.team_id,
        firm_id=active_firm.id,
        filters={},
    )
    return {
        "membership": membership,
        "active_firm": active_firm,
        "ctx": ctx,
        "benchmark_asset_class": _credit_default_benchmark_asset_class(membership.team_id),
    }


def _credit_pdf_payload_for_page(page, export_ctx):
    ctx = export_ctx["ctx"]
    loans = ctx["loans"]
    metrics_by_id = ctx["metrics_by_id"]
    snapshots_by_loan = ctx["snapshots_by_loan"]
    fund_performance = ctx.get("fund_performance", {})
    team_id = export_ctx["membership"].team_id
    benchmark_asset_class = export_ctx["benchmark_asset_class"]

    if page == "credit-track-record":
        return compute_credit_track_record(loans, metrics_by_id, fund_performance=fund_performance)
    if page == "credit-benchmarking":
        return compute_credit_benchmarking_analysis(
            loans,
            fund_performance=fund_performance,
            benchmark_thresholds=_load_team_benchmark_thresholds(team_id, benchmark_asset_class),
            benchmark_asset_class=benchmark_asset_class,
        )
    if page == "credit-concentration":
        return compute_credit_concentration(loans, metrics_by_id)
    if page == "credit-fundamentals":
        return compute_credit_fundamentals(loans, metrics_by_id, snapshots_by_loan=snapshots_by_loan)
    if page == "credit-pricing-trends":
        return compute_credit_pricing_trends(
            loans,
            metrics_by_id,
            primary_dim="sector",
            time_group="quarter",
        )
    if page == "credit-underwrite-outcome":
        return compute_credit_underwrite_outcome(loans, metrics_by_id, snapshots_by_loan=snapshots_by_loan)
    if page == "credit-data-cuts":
        available_dim_keys = credit_data_cuts_available_dimension_keys(loans)
        return {
            "all_cuts": {
                dim_key: compute_credit_data_cuts(loans, metrics_by_id, primary_dim=dim_key)
                for dim_key in available_dim_keys
            }
        }
    raise KeyError(page)


def _build_credit_analysis_pdf(page, payload, *, report_title, currency_code, as_of_date, benchmark_asset_class=""):
    if page == "credit-track-record":
        return _build_credit_track_record_pdf(
            payload,
            currency_code=currency_code,
            report_title=report_title,
            as_of_date=as_of_date,
        )
    if page == "credit-benchmarking":
        return _build_benchmarking_pdf(
            payload,
            report_title=report_title,
            currency_code=currency_code,
            benchmark_asset_class=benchmark_asset_class,
        )
    if page == "credit-concentration":
        return _build_credit_concentration_pdf(
            payload,
            currency_code=currency_code,
            report_title=report_title,
            as_of_date=as_of_date,
        )
    if page == "credit-fundamentals":
        return _build_credit_fundamentals_pdf(
            payload,
            currency_code=currency_code,
            report_title=report_title,
            as_of_date=as_of_date,
        )
    if page == "credit-pricing-trends":
        return _build_credit_pricing_trends_pdf(
            payload,
            currency_code=currency_code,
            report_title=report_title,
            as_of_date=as_of_date,
        )
    if page == "credit-underwrite-outcome":
        return _build_credit_underwrite_outcome_pdf(
            payload,
            currency_code=currency_code,
            report_title=report_title,
            as_of_date=as_of_date,
        )
    if page == "credit-data-cuts":
        return _build_credit_data_cuts_summary_pdf(
            payload.get("all_cuts") or {},
            currency_code=currency_code,
            report_title=report_title,
            as_of_date=as_of_date,
        )
    raise KeyError(page)


def _handle_upload(parse_func, redirect_route):
    membership = _require_team_scope()

    if "file" not in request.files:
        flash("No file selected.", "danger")
        return redirect(url_for("upload"))

    file = request.files["file"]
    if file.filename == "":
        flash("No file selected.", "danger")
        return redirect(url_for("upload"))

    if not _allowed_file(file.filename):
        flash("Invalid file type. Please upload an .xlsx or .xls file.", "danger")
        return redirect(url_for("upload"))

    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(file_path)

    try:
        try:
            result = parse_func(
                file_path,
                team_id=membership.team_id,
                uploader_user_id=current_user.id,
                replace_mode="replace_fund",
            )
        except SQLAlchemyError as exc:
            if not _recover_missing_tables(exc):
                raise
            result = parse_func(
                file_path,
                team_id=membership.team_id,
                uploader_user_id=current_user.id,
                replace_mode="replace_fund",
            )
        if result["success"] > 0:
            firms_processed = result.get("firms_processed") or []
            firm_count = int(result.get("firm_count") or (len(firms_processed) or 1))
            if firm_count > 1 and firms_processed:
                flash(
                    f"Successfully imported {result['success']} deal records across {firm_count} firms "
                    f"(batch {result['batch_id']}).",
                    "success",
                )
                for firm_summary in firms_processed[:12]:
                    firm_name = firm_summary.get("firm_name") or "Unknown Firm"
                    firm_currency = (
                        normalize_currency_code(firm_summary.get("currency"), default=DEFAULT_CURRENCY_CODE)
                        or DEFAULT_CURRENCY_CODE
                    )
                    upload_as_of_date = firm_summary.get("as_of_date")
                    as_of_text = (
                        upload_as_of_date.isoformat()
                        if hasattr(upload_as_of_date, "isoformat")
                        else str(upload_as_of_date or "N/A")
                    )
                    # Show currency conversion info
                    perf_ccy = firm_summary.get("perf_currency", firm_currency)
                    fin_ccys = firm_summary.get("fin_currencies", [perf_ccy])
                    perf_rate = firm_summary.get("perf_fx_rate")
                    ccy_parts = [f"Perf: {perf_ccy}"]
                    if perf_rate and perf_ccy != "USD":
                        ccy_parts[0] += f" @{perf_rate:.4f}"
                    if fin_ccys and fin_ccys != [perf_ccy]:
                        ccy_parts.append(f"Fin: {', '.join(fin_ccys)}")
                    flash(f"{firm_name} | {' | '.join(ccy_parts)} | As Of: {as_of_text}", "info")

                    for cw in firm_summary.get("conversion_warnings", []):
                        flash(str(cw).splitlines()[0].strip(), "warning")
                    fx_warning = firm_summary.get("fx_warning")
                    if fx_warning:
                        flash(
                            f"FX warning ({firm_name}): {str(fx_warning).splitlines()[0].strip()}",
                            "warning",
                        )
                if len(firms_processed) > 12:
                    flash(f"...and {len(firms_processed) - 12} additional firms in this upload.", "info")
            else:
                flash(f"Successfully imported {result['success']} deal records (batch {result['batch_id']}).", "success")
                firm_name = result.get("firm_name")
                if firm_name:
                    flash(f"Upload firm scope: {firm_name}.", "info")
                firm_currency = normalize_currency_code(result.get("firm_currency"), default=DEFAULT_CURRENCY_CODE) or DEFAULT_CURRENCY_CODE
                flash(f"Firm currency: {firm_currency}.", "info")
                upload_as_of_date = result.get("as_of_date")
                if upload_as_of_date is not None:
                    as_of_text = upload_as_of_date.isoformat() if hasattr(upload_as_of_date, "isoformat") else str(upload_as_of_date)
                    flash(f"Upload As Of Date: {as_of_text}.", "info")
                fx_status = (result.get("fx_status") or "").lower()
                fx_rate = result.get("fx_rate_to_usd")
                fx_date = result.get("fx_rate_date")
                if firm_currency != DEFAULT_CURRENCY_CODE and fx_status == "ok" and fx_rate:
                    fx_date_text = fx_date.isoformat() if hasattr(fx_date, "isoformat") else str(fx_date or "N/A")
                    flash(
                        f"Reporting conversion active: {firm_currency}->USD at {float(fx_rate):.6f} (effective {fx_date_text}).",
                        "info",
                    )
                fx_warning = result.get("fx_warning")
                if fx_warning:
                    flash(f"FX warning: {str(fx_warning).splitlines()[0].strip()}", "warning")
        replaced_funds = result.get("replaced_funds") or {}
        if replaced_funds:
            replaced_summaries = ", ".join(f"{name} ({count} old deals replaced)" for name, count in replaced_funds.items())
            if result.get("firm_name"):
                flash(f"Replaced existing fund data in {result['firm_name']}: {replaced_summaries}.", "info")
            else:
                flash(f"Replaced existing fund data: {replaced_summaries}.", "info")
        elif result.get("firm_count", 1) > 1:
            for firm_summary in (result.get("firms_processed") or []):
                firm_replaced = firm_summary.get("replaced_funds") or {}
                if not firm_replaced:
                    continue
                replaced_summaries = ", ".join(
                    f"{name} ({count} old deals replaced)" for name, count in firm_replaced.items()
                )
                firm_name = firm_summary.get("firm_name") or "Unknown Firm"
                flash(f"Replaced existing fund data in {firm_name}: {replaced_summaries}.", "info")
        if result.get("duplicates_skipped", 0) > 0:
            flash(f"Skipped {result['duplicates_skipped']} duplicate deal records.", "warning")
        if result.get("quarantined_count", 0) > 0:
            flash(
                f"Quarantined {result['quarantined_count']} invalid row(s). "
                f"Issue Report ID: {result.get('issue_report_id')}",
                "warning",
            )
        if result.get("bridge_complete") is not None and result["success"] > 0:
            pct = result["bridge_complete"] / result["success"] * 100
            flash(f"Bridge-ready deals: {result['bridge_complete']}/{result['success']} ({pct:.0f}%).", "info")
        supplemental = result.get("supplemental_counts") or {}
        if supplemental:
            extras = []
            if supplemental.get("cashflows", 0) > 0:
                extras.append(f"{supplemental['cashflows']} cashflow events")
            if supplemental.get("deal_quarterly", 0) > 0:
                extras.append(f"{supplemental['deal_quarterly']} deal quarterly marks")
            if supplemental.get("fund_quarterly", 0) > 0:
                extras.append(f"{supplemental['fund_quarterly']} fund quarterly snapshots")
            if supplemental.get("underwrite", 0) > 0:
                extras.append(f"{supplemental['underwrite']} underwrite baseline rows")
            if supplemental.get("fund_metadata", 0) > 0:
                extras.append(f"{supplemental['fund_metadata']} fund metadata rows")
            if supplemental.get("fund_cashflows", 0) > 0:
                extras.append(f"{supplemental['fund_cashflows']} fund cashflow rows")
            if supplemental.get("public_market_benchmarks", 0) > 0:
                extras.append(f"{supplemental['public_market_benchmarks']} public market benchmark rows")
            if extras:
                flash("Imported supplemental analysis rows: " + ", ".join(extras) + ".", "info")
        for err in result.get("errors", [])[:8]:
            flash(err, "warning")
        if len(result.get("errors", [])) > 8:
            flash(f"...and {len(result['errors']) - 8} additional warnings.", "warning")
        if result["success"] == 0 and not result.get("errors"):
            flash("No records found in the file.", "warning")
    except Exception as exc:
        logger.exception("Error processing upload")
        flash(f"Error processing file: {str(exc)}", "danger")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

    return redirect(url_for(redirect_route))


def _empty_dashboard_context():
    metric_pair = {"avg": None, "wavg": None}
    empty_entry_exit = {
        "entry": {
            "tev_ebitda": metric_pair,
            "tev_revenue": metric_pair,
            "net_debt_ebitda": metric_pair,
            "net_debt_tev": metric_pair,
            "ebitda_margin": metric_pair,
        },
        "exit": {
            "tev_ebitda": metric_pair,
            "tev_revenue": metric_pair,
            "net_debt_ebitda": metric_pair,
            "net_debt_tev": metric_pair,
            "ebitda_margin": metric_pair,
        },
        "growth": {
            "revenue_growth": metric_pair,
            "ebitda_growth": metric_pair,
            "revenue_cagr": metric_pair,
            "ebitda_cagr": metric_pair,
        },
        "returns": {
            "gross_moic": metric_pair,
            "gross_irr": metric_pair,
            "implied_irr": metric_pair,
            "hold_period": metric_pair,
        },
    }

    return {
        "kpis": {
            "total_deals": 0,
            "total_equity": 0,
            "total_value": 0,
            "total_value_created": 0,
            "gross_moic": None,
            "gross_irr": None,
        },
        "loss": {"count_pct": None, "capital_pct": None, "loss_count": 0, "total_count": 0},
        "moic_distribution": [],
        "entry_exit_summary": empty_entry_exit,
        "bridge_aggregate": {
            "model": "additive",
            "basis": "fund",
            "ready_count": 0,
            "fallback_ready_count": 0,
            "drivers": {"dollar": {}, "moic": {}, "pct": {}},
            "total_value_created": 0,
            "total_equity": 0,
            "start_end": {
                "dollar": {"start": 0, "end": 0},
                "moic": {"start": None, "end": None},
                "pct": {"start": 0, "end": 1},
            },
        },
        "vintage_series": [],
        "moic_hold_scatter": [],
        "value_creation_mix": {
            "current": "fund",
            "series": {
                "fund": {"labels": [], "drivers": {}, "totals_dollar": []},
                "sector": {"labels": [], "drivers": {}, "totals_dollar": []},
                "exit_type": {"labels": [], "drivers": {}, "totals_dollar": []},
            },
        },
        "realized_unrealized_exposure": {"labels": [], "realized": [], "unrealized": []},
        "loss_concentration_heatmap": {"sectors": [], "geographies": [], "values": [], "max_value": 0},
        "exit_type_performance": {
            "labels": [],
            "calculated_moic": [],
            "deal_count": [],
            "realized_value": [],
        },
        "lead_partner_scorecard": [],
        "deal_metrics": {},
        "deals": [],
        "data_quality": {"total_deals": 0, "complete_deals": 0, "bridge_ready": 0, "warnings": []},
        "takeaway": {
            "text": "Load deal, quarter, and benchmark data to turn the dashboard into an LP-ready briefing surface.",
            "tone": "neutral",
            "why_it_matters": "The dashboard is strongest when current scope, benchmark coverage, and reporting quality are all available.",
        },
        "coverage": {"items": []},
        "confidence": {
            "level": "low",
            "label": "Low Confidence",
            "tone": "risk",
            "note": "No portfolio data is currently loaded.",
        },
        "risk_flags": [],
        "fund_summary_rows": [],
        "funds": [],
        "statuses": [],
        "sectors": [],
        "geographies": [],
        "vintages": [],
        "exit_types": [],
        "lead_partners": [],
        "security_types": [],
        "deal_types": [],
        "entry_channels": [],
        "current_fund": "",
        "current_status": "",
        "current_sector": "",
        "current_geography": "",
        "current_vintage": "",
        "current_exit_type": "",
        "current_lead_partner": "",
        "current_security_type": "",
        "current_deal_type": "",
        "current_entry_channel": "",
        "display_as_of_date": date.today(),
        "benchmark_asset_classes": [],
        "current_benchmark_asset_class": "",
    }


def _build_filtered_deals_context(fund_override=None):
    membership = _current_membership()
    active_firm = _resolve_active_firm_for_team()
    active_team = db.session.get(Team, membership.team_id) if membership is not None else None
    reporting = _reporting_currency_context(active_firm)
    context = build_analysis_context(
        membership=membership,
        active_team=active_team,
        active_firm=active_firm,
        request_values=request.args,
        session_store=session,
        reporting=reporting,
        fund_override=fund_override,
    )
    return context.as_legacy_dict()


def _extract_chart_builder_global_filters_from_request():
    filters = parse_request_filters(request.values)
    filters["benchmark_asset_class"] = (request.values.get("benchmark_asset_class", "") or "").strip()
    return filters


def _serialize_chart_builder_template(row):
    config = {}
    try:
        config = json.loads(row.config_json or "{}")
    except (TypeError, ValueError):
        config = {}
    return {
        "id": row.id,
        "team_id": row.team_id,
        "name": row.name,
        "source": row.source,
        "config": config,
        "created_by_user_id": row.created_by_user_id,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


def _build_dashboard_payload(filtered_deals, team_id=None, benchmark_asset_class="", metrics_by_id=None):
    metrics_by_id = metrics_by_id or {d.id: compute_deal_metrics(d) for d in filtered_deals}
    firm_id = filtered_deals[0].firm_id if filtered_deals else None
    fund_vintage_years = build_fund_vintage_lookup(filtered_deals, team_id=team_id, firm_id=firm_id)

    portfolio = compute_portfolio_analytics(filtered_deals, metrics_by_id=metrics_by_id)
    risk = compute_loss_and_distribution(filtered_deals, metrics_by_id=metrics_by_id)
    additive = compute_bridge_aggregate(filtered_deals, basis="fund")
    vintage = compute_vintage_series(filtered_deals, metrics_by_id=metrics_by_id)
    moic_hold_scatter = compute_moic_hold_scatter(filtered_deals, metrics_by_id=metrics_by_id)
    value_creation_mix = {
        "current": "fund",
        "series": {
            "fund": compute_value_creation_mix(filtered_deals, metrics_by_id=metrics_by_id, group_by="fund"),
            "sector": compute_value_creation_mix(filtered_deals, metrics_by_id=metrics_by_id, group_by="sector"),
            "exit_type": compute_value_creation_mix(filtered_deals, metrics_by_id=metrics_by_id, group_by="exit_type"),
        },
    }
    realized_unrealized_exposure = compute_realized_unrealized_exposure(filtered_deals)
    loss_concentration_heatmap = compute_loss_concentration_heatmap(filtered_deals, metrics_by_id=metrics_by_id)
    exit_type_performance = compute_exit_type_performance(filtered_deals, metrics_by_id=metrics_by_id)
    lead_partner_scorecard = compute_lead_partner_scorecard(filtered_deals, metrics_by_id=metrics_by_id)
    quality = compute_data_quality(filtered_deals, metrics_by_id)
    track_record = compute_deal_track_record(filtered_deals, metrics_by_id=metrics_by_id, fund_vintage_lookup=fund_vintage_years)
    benchmark_thresholds = _load_team_benchmark_thresholds(team_id, benchmark_asset_class)

    fund_summary_rows = []
    for fund in track_record.get("funds", []):
        net = fund.get("net_performance") or {}
        conflicts = net.get("conflicts") or {}
        vintage_year = fund_vintage_years.get(fund.get("fund_name"))
        net_irr_value = None if conflicts.get("net_irr") else net.get("net_irr")
        net_moic_value = None if conflicts.get("net_moic") else net.get("net_moic")
        net_dpi_value = None if conflicts.get("net_dpi") else net.get("net_dpi")
        fund_summary_rows.append(
            {
                "fund_name": fund.get("fund_name"),
                "vintage_year": vintage_year,
                "fund_size": fund.get("fund_size"),
                "net_irr": net.get("net_irr"),
                "net_moic": net.get("net_moic"),
                "net_dpi": net.get("net_dpi"),
                "conflicts": {
                    "fund_size": bool(fund.get("fund_size_conflict")),
                    "net_irr": bool(conflicts.get("net_irr")),
                    "net_moic": bool(conflicts.get("net_moic")),
                    "net_dpi": bool(conflicts.get("net_dpi")),
                },
                "benchmark_net_irr": _rank_benchmark_metric(
                    net_irr_value,
                    vintage_year,
                    "net_irr",
                    benchmark_thresholds,
                    benchmark_asset_class,
                ),
                "benchmark_net_moic": _rank_benchmark_metric(
                    net_moic_value,
                    vintage_year,
                    "net_moic",
                    benchmark_thresholds,
                    benchmark_asset_class,
                ),
                "benchmark_net_dpi": _rank_benchmark_metric(
                    net_dpi_value,
                    vintage_year,
                    "net_dpi",
                    benchmark_thresholds,
                    benchmark_asset_class,
                ),
            }
        )

    fund_summary_rows.sort(
        key=lambda row: (
            row.get("vintage_year") is None,
            row.get("vintage_year") if row.get("vintage_year") is not None else 9999,
            (row.get("fund_name") or "").lower(),
        )
    )

    kpis = {
        "total_deals": len(filtered_deals),
        "total_equity": portfolio["total_equity"],
        "total_value": portfolio["total_value"],
        "total_value_created": portfolio["total_value_created"],
        "gross_moic": portfolio["returns"]["gross_moic"]["avg"],
        "gross_irr": portfolio["returns"]["gross_irr"]["wavg"],
    }
    risk_flags = []
    if filtered_deals and not benchmark_asset_class:
        risk_flags.append("No benchmark asset class is selected, so peer comparisons are not visible on the dashboard.")
    if quality["warnings"]:
        risk_flags.append(f"{len(quality['warnings'])} data-quality warning(s) are present in the current filter set.")
    capital_loss_pct = risk["loss_ratios"].get("capital_pct")
    if capital_loss_pct is not None and capital_loss_pct > 30:
        risk_flags.append("Capital loss concentration remains elevated in the current scope.")

    bridge_ready_pct = (quality["bridge_ready"] / quality["total_deals"]) if quality["total_deals"] else None
    complete_deal_pct = (quality["complete_deals"] / quality["total_deals"]) if quality["total_deals"] else None

    return {
        "kpis": kpis,
        "loss": risk["loss_ratios"],
        "moic_distribution": risk["moic_distribution"],
        "entry_exit_summary": {
            "entry": portfolio["entry"],
            "exit": portfolio["exit"],
            "growth": portfolio["growth"],
            "returns": portfolio["returns"],
        },
        "bridge_aggregate": additive,
        "vintage_series": vintage,
        "moic_hold_scatter": moic_hold_scatter,
        "value_creation_mix": value_creation_mix,
        "realized_unrealized_exposure": realized_unrealized_exposure,
        "loss_concentration_heatmap": loss_concentration_heatmap,
        "exit_type_performance": exit_type_performance,
        "lead_partner_scorecard": lead_partner_scorecard,
        "deal_metrics": metrics_by_id,
        "data_quality": quality,
        "takeaway": {
            "text": (
                "Use the dashboard as the opening briefing page: it should tell you whether the current portfolio is compounding cleanly, "
                "where value creation is coming from, and whether the underlying data is strong enough to trust the signal."
            ),
            "tone": "warning" if risk_flags else "positive",
            "why_it_matters": "This is the fastest way to assess portfolio health before drilling into benchmarking, liquidity, or individual deal pages.",
        },
        "coverage": {
            "items": [
                {"label": "Funds", "value": len(fund_summary_rows), "display": len(fund_summary_rows), "tone": "neutral"},
                {
                    "label": "Bridge Ready",
                    "value": bridge_ready_pct,
                    "display": f"{bridge_ready_pct * 100:.1f}%" if bridge_ready_pct is not None else "—",
                    "tone": "positive" if bridge_ready_pct is not None and bridge_ready_pct >= 0.6 else "warning",
                },
                {
                    "label": "Complete Deals",
                    "value": complete_deal_pct,
                    "display": f"{complete_deal_pct * 100:.1f}%" if complete_deal_pct is not None else "—",
                    "tone": "positive" if complete_deal_pct is not None and complete_deal_pct >= 0.75 else "warning",
                },
                {
                    "label": "Benchmark",
                    "value": benchmark_asset_class or "No Benchmark",
                    "display": benchmark_asset_class or "No Benchmark",
                    "tone": "positive" if benchmark_asset_class else "warning",
                },
            ]
        },
        "confidence": {
            "level": "high" if not risk_flags and filtered_deals else ("medium" if filtered_deals else "low"),
            "label": "High Confidence" if not risk_flags and filtered_deals else ("Moderate Confidence" if filtered_deals else "Low Confidence"),
            "tone": "positive" if not risk_flags and filtered_deals else ("warning" if filtered_deals else "risk"),
            "note": "Confidence blends reporting completeness, bridge coverage, and current benchmark context.",
        },
        "risk_flags": risk_flags,
        "fund_summary_rows": fund_summary_rows,
    }


def _analysis_route_payload(page, filtered_deals, firm_id=None, team_id=None, benchmark_asset_class="", metrics_by_id=None):
    metrics_by_id = metrics_by_id or {d.id: compute_deal_metrics(d) for d in filtered_deals}

    if page == "executive-summary":
        rank_by = (request.args.get("rank_by", "") or "moic").strip()
        if rank_by not in ("moic", "irr"):
            rank_by = "moic"
        # Load benchmark thresholds for Performance and Peer Context signals
        bench_thresholds = _load_team_benchmark_thresholds(team_id, benchmark_asset_class)
        return compute_executive_summary_analysis(
            filtered_deals,
            metrics_by_id=metrics_by_id,
            firm_id=firm_id,
            team_id=team_id,
            rank_by=rank_by,
            benchmark_thresholds=bench_thresholds,
            benchmark_asset_class=benchmark_asset_class,
        )
    if page == "chart-builder":
        return {}
    if page == "fund-liquidity":
        return compute_fund_liquidity_analysis(filtered_deals, firm_id=firm_id)
    if page == "underwrite-outcome":
        return compute_underwrite_outcome_analysis(filtered_deals, metrics_by_id=metrics_by_id)
    if page == "valuation-quality":
        return compute_valuation_quality_analysis(filtered_deals)
    if page == "exit-readiness":
        return compute_exit_readiness_analysis(filtered_deals, metrics_by_id=metrics_by_id)
    if page == "stress-lab":
        deal_overrides = {}
        for deal in filtered_deals:
            multiple_raw = request.args.get(f"ms_{deal.id}")
            ebitda_raw = request.args.get(f"es_{deal.id}")
            hold_raw = request.args.get(f"hp_{deal.id}")
            if multiple_raw in (None, "") and ebitda_raw in (None, "") and hold_raw in (None, ""):
                continue

            override = {}
            if multiple_raw not in (None, ""):
                try:
                    override["multiple_shock"] = float(multiple_raw)
                except ValueError:
                    pass
            if ebitda_raw not in (None, ""):
                try:
                    override["ebitda_shock"] = float(ebitda_raw) / 100.0
                except ValueError:
                    pass
            if hold_raw not in (None, ""):
                try:
                    override["expected_hold_years"] = float(hold_raw)
                except ValueError:
                    pass
            if override:
                deal_overrides[deal.id] = override

        scenario = {
            "default_multiple_shock": 0.0,
            "default_ebitda_shock": 0.0,
        }
        return compute_stress_lab_analysis(
            filtered_deals,
            scenario=scenario,
            deal_overrides=deal_overrides,
            metrics_by_id=metrics_by_id,
        )
    if page == "deal-trajectory":
        return compute_deal_trajectory_analysis(
            filtered_deals,
            deal_id=request.args.get("deal_id"),
            metrics_by_id=metrics_by_id,
        )
    if page == "vca-ebitda":
        return compute_vca_ebitda_analysis(filtered_deals, metrics_by_id=metrics_by_id)
    if page == "vca-revenue":
        return compute_vca_revenue_analysis(filtered_deals, metrics_by_id=metrics_by_id)
    if page == "vca-addons":
        return compute_vca_addons_analysis(filtered_deals, metrics_by_id=metrics_by_id)
    if page == "vca-addons-revenue":
        return compute_vca_addons_revenue_analysis(filtered_deals, metrics_by_id=metrics_by_id)
    if page == "organic-growth":
        return compute_organic_growth_analysis(filtered_deals, metrics_by_id=metrics_by_id)
    if page == "data-cuts":
        primary_dim = request.args.get("dim", "sector")
        secondary_dim = request.args.get("dim2", "")
        return compute_data_cuts_analytics(
            filtered_deals, metrics_by_id,
            primary_dim=primary_dim,
            secondary_dim=secondary_dim or None,
        )
    if page == "benchmarking":
        thresholds = _load_team_benchmark_thresholds(team_id, benchmark_asset_class)
        return compute_benchmarking_analysis(
            filtered_deals,
            benchmark_thresholds=thresholds,
            benchmark_asset_class=benchmark_asset_class,
            metrics_by_id=metrics_by_id,
            fund_vintage_lookup=build_fund_vintage_lookup(filtered_deals, team_id=team_id, firm_id=firm_id),
        )
    if page == "nav-at-risk":
        return compute_nav_at_risk_analysis(
            filtered_deals,
            firm_id=firm_id,
            team_id=team_id,
            metrics_by_id=metrics_by_id,
            as_of_date=resolve_analysis_as_of_date(filtered_deals),
        )
    if page == "public-market-comparison":
        return compute_public_market_comparison_analysis(
            filtered_deals,
            team_id=team_id,
            firm_id=firm_id,
            benchmark_asset_class=benchmark_asset_class,
            as_of_date=resolve_analysis_as_of_date(filtered_deals),
        )
    if page == "lp-due-diligence-memo":
        return compute_lp_due_diligence_memo(
            filtered_deals,
            team_id=team_id,
            firm_id=firm_id,
            benchmark_asset_class=benchmark_asset_class,
            metrics_by_id=metrics_by_id,
            as_of_date=resolve_analysis_as_of_date(filtered_deals),
        )
    abort(404)


def _safe_next_url(candidate):
    if not candidate:
        return None
    if candidate.startswith("/") and not candidate.startswith("//"):
        return candidate
    return None


def _purge_fund_for_firm(firm_id, fund_name):
    deal_ids = [
        row[0]
        for row in db.session.query(Deal.id)
        .filter(Deal.firm_id == firm_id, Deal.fund_number == fund_name)
        .all()
    ]
    if deal_ids:
        DealCashflowEvent.query.filter(
            DealCashflowEvent.firm_id == firm_id,
            DealCashflowEvent.deal_id.in_(deal_ids),
        ).delete(synchronize_session=False)
        DealQuarterSnapshot.query.filter(
            DealQuarterSnapshot.firm_id == firm_id,
            DealQuarterSnapshot.deal_id.in_(deal_ids),
        ).delete(synchronize_session=False)
        DealUnderwriteBaseline.query.filter(
            DealUnderwriteBaseline.firm_id == firm_id,
            DealUnderwriteBaseline.deal_id.in_(deal_ids),
        ).delete(synchronize_session=False)

    FundQuarterSnapshot.query.filter_by(firm_id=firm_id, fund_number=fund_name).delete(synchronize_session=False)
    FundMetadata.query.filter_by(firm_id=firm_id, fund_number=fund_name).delete(synchronize_session=False)
    FundCashflow.query.filter_by(firm_id=firm_id, fund_number=fund_name).delete(synchronize_session=False)
    deleted_deals = Deal.query.filter_by(firm_id=firm_id, fund_number=fund_name).delete(synchronize_session=False)
    db.session.commit()
    return deleted_deals


def _upload_batches_for_firm(firm_id, limit=30):
    if firm_id is None:
        return []

    rows = (
        db.session.query(
            Deal.upload_batch.label("batch_id"),
            db.func.count(Deal.id).label("deal_count"),
            db.func.max(Deal.created_at).label("uploaded_at"),
        )
        .filter(
            Deal.firm_id == firm_id,
            Deal.upload_batch.isnot(None),
            Deal.upload_batch != "",
        )
        .group_by(Deal.upload_batch)
        .order_by(db.func.max(Deal.created_at).desc(), Deal.upload_batch.desc())
        .limit(limit)
        .all()
    )
    if not rows:
        return []

    batch_ids = [row.batch_id for row in rows if row.batch_id]
    funds_by_batch = {}
    if batch_ids:
        fund_rows = (
            db.session.query(Deal.upload_batch, Deal.fund_number)
            .filter(
                Deal.firm_id == firm_id,
                Deal.upload_batch.in_(batch_ids),
            )
            .all()
        )
        for batch_id, fund_number in fund_rows:
            if not batch_id or not fund_number:
                continue
            bucket = funds_by_batch.setdefault(batch_id, set())
            bucket.add(fund_number)

        issue_rows = (
            db.session.query(UploadIssue.upload_batch, db.func.count(UploadIssue.id))
            .filter(
                UploadIssue.firm_id == firm_id,
                UploadIssue.upload_batch.in_(batch_ids),
            )
            .group_by(UploadIssue.upload_batch)
            .all()
        )
        issue_counts = {batch_id: int(count or 0) for batch_id, count in issue_rows}
    else:
        issue_counts = {}

    history_rows = []
    for row in rows:
        batch_id = row.batch_id
        fund_names = sorted(funds_by_batch.get(batch_id, set()))
        history_rows.append(
            {
                "batch_id": batch_id,
                "deal_count": int(row.deal_count or 0),
                "uploaded_at": row.uploaded_at,
                "fund_count": len(fund_names),
                "fund_names": fund_names,
                "issue_count": int(issue_counts.get(batch_id, 0)),
            }
        )
    return history_rows


def _benchmark_dataset_for_team(team_id):
    if team_id is None:
        return None

    try:
        rows = (
            BenchmarkPoint.query.filter(BenchmarkPoint.team_id == team_id)
            .order_by(BenchmarkPoint.created_at.desc(), BenchmarkPoint.id.desc())
            .all()
        )
    except SQLAlchemyError as exc:
        _handle_db_exception(exc, "Benchmark dataset lookup failed")
        return None
    if not rows:
        return None

    asset_classes = sorted({row.asset_class for row in rows if row.asset_class})
    vintages = [int(row.vintage_year) for row in rows if row.vintage_year is not None]
    latest = rows[0]
    return {
        "rows_loaded": len(rows),
        "upload_batch": latest.upload_batch,
        "asset_classes": asset_classes,
        "vintage_min": min(vintages) if vintages else None,
        "vintage_max": max(vintages) if vintages else None,
        "updated_at": latest.created_at,
    }


def _delete_upload_batch_for_firm(firm_id, batch_id):
    batch_id = (batch_id or "").strip()
    if firm_id is None or not batch_id:
        return {
            "deals": 0,
            "cashflows": 0,
            "deal_quarterly": 0,
            "fund_quarterly": 0,
            "underwrite": 0,
            "fund_metadata": 0,
            "fund_cashflows": 0,
            "public_market_benchmarks": 0,
            "upload_issues": 0,
        }

    deal_ids = [
        row[0]
        for row in db.session.query(Deal.id)
        .filter(
            Deal.firm_id == firm_id,
            Deal.upload_batch == batch_id,
        )
        .all()
    ]

    if deal_ids:
        cashflow_scope = or_(DealCashflowEvent.upload_batch == batch_id, DealCashflowEvent.deal_id.in_(deal_ids))
        deal_quarter_scope = or_(
            DealQuarterSnapshot.upload_batch == batch_id,
            DealQuarterSnapshot.deal_id.in_(deal_ids),
        )
        underwrite_scope = or_(
            DealUnderwriteBaseline.upload_batch == batch_id,
            DealUnderwriteBaseline.deal_id.in_(deal_ids),
        )
    else:
        cashflow_scope = DealCashflowEvent.upload_batch == batch_id
        deal_quarter_scope = DealQuarterSnapshot.upload_batch == batch_id
        underwrite_scope = DealUnderwriteBaseline.upload_batch == batch_id

    deleted_counts = {
        "cashflows": DealCashflowEvent.query.filter(
            DealCashflowEvent.firm_id == firm_id,
            cashflow_scope,
        ).delete(synchronize_session=False),
        "deal_quarterly": DealQuarterSnapshot.query.filter(
            DealQuarterSnapshot.firm_id == firm_id,
            deal_quarter_scope,
        ).delete(synchronize_session=False),
        "underwrite": DealUnderwriteBaseline.query.filter(
            DealUnderwriteBaseline.firm_id == firm_id,
            underwrite_scope,
        ).delete(synchronize_session=False),
        "fund_quarterly": FundQuarterSnapshot.query.filter(
            FundQuarterSnapshot.firm_id == firm_id,
            FundQuarterSnapshot.upload_batch == batch_id,
        ).delete(synchronize_session=False),
        "fund_metadata": FundMetadata.query.filter(
            FundMetadata.firm_id == firm_id,
            FundMetadata.upload_batch == batch_id,
        ).delete(synchronize_session=False),
        "fund_cashflows": FundCashflow.query.filter(
            FundCashflow.firm_id == firm_id,
            FundCashflow.upload_batch == batch_id,
        ).delete(synchronize_session=False),
        "public_market_benchmarks": PublicMarketIndexLevel.query.filter(
            PublicMarketIndexLevel.upload_batch == batch_id,
        ).delete(synchronize_session=False),
        "upload_issues": UploadIssue.query.filter(
            UploadIssue.firm_id == firm_id,
            UploadIssue.upload_batch == batch_id,
        ).delete(synchronize_session=False),
        "deals": Deal.query.filter(
            Deal.firm_id == firm_id,
            Deal.upload_batch == batch_id,
        ).delete(synchronize_session=False),
    }
    db.session.commit()
    return deleted_counts


def _schema_readiness_status():
    errors = []
    revision = None

    db.session.execute(text("SELECT 1"))
    inspector = sa_inspect(db.engine)
    table_names = set(inspector.get_table_names())
    missing_tables = [name for name in REQUIRED_SCHEMA_TABLES if name not in table_names]
    if missing_tables:
        errors.append(f"missing tables: {', '.join(missing_tables)}")

    if "alembic_version" in table_names:
        revision = db.session.execute(text("SELECT version_num FROM alembic_version LIMIT 1")).scalar()
        if not revision:
            errors.append("alembic_version is present but has no version row")

    status = "ok" if not errors else "error"
    payload = {"status": status}
    if not app.config.get("IS_PRODUCTION"):
        payload["revision"] = revision
        payload["missing_tables"] = missing_tables
        payload["errors"] = errors
    return payload, (200 if status == "ok" else 500)


@app.route("/healthz")
def healthz():
    try:
        db.session.execute(text("SELECT 1"))
    except Exception as exc:
        if app.config.get("IS_PRODUCTION"):
            logger.exception("Health check failed")
            return jsonify({"status": "error"}), 500
        return jsonify({"status": "error", "detail": str(exc)}), 500
    return jsonify({"status": "ok"}), 200


@app.route("/readyz")
def readyz():
    try:
        payload, status_code = _schema_readiness_status()
    except SQLAlchemyError as exc:
        _handle_db_exception(exc, "Schema readiness check failed")
        payload = {"status": "error"}
        if not app.config.get("IS_PRODUCTION"):
            payload["detail"] = str(_root_db_error(exc))
        return jsonify(payload), 500
    return jsonify(payload), status_code


@app.route("/auth/login", methods=["GET", "POST"])
@limiter.limit("5/minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        next_url = _safe_next_url(request.form.get("next")) or url_for("dashboard")

        user = User.query.filter_by(email=email).first()
        if user is None or not check_password_hash(user.password_hash, password):
            flash("Invalid email or password.", "danger")
            return render_template("login.html", next_url=next_url)
        if not user.is_active:
            flash("Your account is inactive. Contact your team admin.", "danger")
            return render_template("login.html", next_url=next_url)

        membership = (
            TeamMembership.query.filter_by(user_id=user.id)
            .order_by(TeamMembership.created_at.asc(), TeamMembership.id.asc())
            .first()
        )
        if membership is None:
            flash("No team membership found for this account.", "danger")
            return render_template("login.html", next_url=next_url)

        login_user(user)
        user.last_login_at = _utc_now_naive()
        session["active_team_id"] = membership.team_id
        _resolve_active_firm_for_team()
        db.session.commit()

        return redirect(next_url)

    next_url = _safe_next_url(request.args.get("next")) or url_for("dashboard")
    return render_template("login.html", next_url=next_url)


@app.route("/auth/logout", methods=["POST"])
@login_required
def logout():
    logout_user()
    session.pop("active_team_id", None)
    session.pop("active_firm_id", None)
    flash("Signed out.", "info")
    return redirect(url_for("login"))


@app.route("/team")
@login_required
def team():
    membership = _require_team_scope()
    team_obj = db.session.get(Team, membership.team_id)
    member_rows = (
        db.session.query(TeamMembership, User)
        .join(User, TeamMembership.user_id == User.id)
        .filter(TeamMembership.team_id == membership.team_id)
        .order_by(TeamMembership.created_at.asc(), User.email.asc())
        .all()
    )
    invites = (
        TeamInvite.query.filter(
            TeamInvite.team_id == membership.team_id,
            TeamInvite.accepted_at.is_(None),
        )
        .order_by(TeamInvite.created_at.desc())
        .all()
    )
    return render_template(
        "team.html",
        team=team_obj,
        membership=membership,
        member_rows=member_rows,
        invites=invites,
        now_utc=_utc_now_naive(),
    )


@app.route("/team/invites", methods=["POST"])
@login_required
@limiter.limit("10/hour")
def create_team_invite():
    membership = _require_team_scope()
    if not _is_team_admin(membership):
        abort(403)

    email = (request.form.get("email") or "").strip().lower()
    if not email:
        flash("Invite email is required.", "danger")
        return redirect(url_for("team"))

    if "@" not in email:
        flash("Enter a valid invite email.", "danger")
        return redirect(url_for("team"))

    existing_user = User.query.filter_by(email=email).first()
    if existing_user is not None:
        existing_membership = TeamMembership.query.filter_by(
            team_id=membership.team_id,
            user_id=existing_user.id,
        ).first()
        if existing_membership is not None:
            flash("That user is already a member of this team.", "warning")
            return redirect(url_for("team"))

    raw_token = secrets.token_urlsafe(32)
    invite = TeamInvite(
        team_id=membership.team_id,
        email=email,
        token_hash=_hash_invite_token(raw_token),
        expires_at=_utc_now_naive() + timedelta(days=7),
        created_by_user_id=current_user.id,
    )
    db.session.add(invite)
    db.session.commit()

    flash("Invite created. Share this link securely: " + _build_invite_link(raw_token), "info")
    return redirect(url_for("team"))


@app.route("/auth/accept-invite/<token>", methods=["GET", "POST"])
def accept_invite(token):
    token_hash = _hash_invite_token(token)
    invite = TeamInvite.query.filter_by(token_hash=token_hash).first()
    if invite is None or invite.accepted_at is not None or invite.expires_at < _utc_now_naive():
        flash("Invite is invalid or expired.", "danger")
        return redirect(url_for("login"))

    team_obj = db.session.get(Team, invite.team_id)
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""

        if email != invite.email.lower():
            flash("Email must match the invited address.", "danger")
            return render_template("accept_invite.html", invite=invite, team=team_obj)
        if len(password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return render_template("accept_invite.html", invite=invite, team=team_obj)
        if password != password_confirm:
            flash("Password confirmation does not match.", "danger")
            return render_template("accept_invite.html", invite=invite, team=team_obj)

        user = User.query.filter_by(email=email).first()
        if user is None:
            user = User(
                email=email,
                password_hash=generate_password_hash(password),
                is_active=True,
                last_login_at=_utc_now_naive(),
            )
            db.session.add(user)
            db.session.flush()
        else:
            user.password_hash = generate_password_hash(password)
            user.last_login_at = _utc_now_naive()

        existing_membership = TeamMembership.query.filter_by(team_id=invite.team_id, user_id=user.id).first()
        if existing_membership is None:
            db.session.add(
                TeamMembership(
                    team_id=invite.team_id,
                    user_id=user.id,
                    role=TEAM_ROLE_MEMBER,
                )
            )

        invite.accepted_at = _utc_now_naive()
        db.session.commit()

        login_user(user)
        session["active_team_id"] = invite.team_id
        _resolve_active_firm_for_team()
        flash("Welcome. Your invite has been accepted.", "success")
        return redirect(url_for("dashboard"))

    return render_template("accept_invite.html", invite=invite, team=team_obj)


@app.route("/team/members/<int:user_id>/toggle-active", methods=["POST"])
@login_required
def toggle_member_active(user_id):
    """Owner/admin can activate or deactivate any team member (except themselves)."""
    membership = _require_team_scope()
    if not _is_team_admin(membership):
        abort(403)

    # Can't deactivate yourself
    if user_id == current_user.id:
        flash("You cannot deactivate your own account.", "warning")
        return redirect(url_for("team"))

    target_membership = TeamMembership.query.filter_by(
        team_id=membership.team_id, user_id=user_id
    ).first_or_404()

    # Only owners can touch other owners/admins
    if target_membership.role in {TEAM_ROLE_OWNER, TEAM_ROLE_ADMIN} and membership.role != TEAM_ROLE_OWNER:
        flash("Only owners can activate or deactivate admins.", "warning")
        return redirect(url_for("team"))

    target_user = db.session.get(User, user_id)
    if target_user is None:
        abort(404)

    target_user.is_active = not target_user.is_active
    db.session.commit()
    state = "activated" if target_user.is_active else "deactivated"
    flash(f"{target_user.email} has been {state}.", "success")
    return redirect(url_for("team"))


@app.route("/team/members/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_team_member(user_id):
    """Owner can remove a user from the team (and deactivate them if they have no other team)."""
    membership = _require_team_scope()
    if membership.role != TEAM_ROLE_OWNER:
        abort(403)

    if user_id == current_user.id:
        flash("You cannot remove yourself from the team.", "warning")
        return redirect(url_for("team"))

    target_membership = TeamMembership.query.filter_by(
        team_id=membership.team_id, user_id=user_id
    ).first_or_404()

    target_user = db.session.get(User, user_id)
    email = target_user.email if target_user else f"User #{user_id}"

    db.session.delete(target_membership)

    # If the user has no remaining team memberships, deactivate their account
    remaining = TeamMembership.query.filter_by(user_id=user_id).count()
    if remaining == 0 and target_user is not None:
        target_user.is_active = False

    db.session.commit()
    flash(f"{email} has been removed from the team.", "success")
    return redirect(url_for("team"))


@app.route("/team/members/add", methods=["POST"])
@login_required
def add_team_member_direct():
    """Owner can add a user directly (create account + membership in one step, no invite needed)."""
    membership = _require_team_scope()
    if membership.role != TEAM_ROLE_OWNER:
        abort(403)

    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    role = request.form.get("role") or TEAM_ROLE_MEMBER

    if not email or "@" not in email:
        flash("A valid email is required.", "danger")
        return redirect(url_for("team"))
    if len(password) < 8:
        flash("Password must be at least 8 characters.", "danger")
        return redirect(url_for("team"))
    if role not in TEAM_ALLOWED_ROLES:
        role = TEAM_ROLE_MEMBER

    existing_user = User.query.filter_by(email=email).first()
    if existing_user is not None:
        # User exists — reactivate and add to team if not already a member
        existing_user.is_active = True
        existing_user.password_hash = generate_password_hash(password)
        user = existing_user
    else:
        user = User(
            email=email,
            password_hash=generate_password_hash(password),
            is_active=True,
        )
        db.session.add(user)
        db.session.flush()

    existing_membership = TeamMembership.query.filter_by(
        team_id=membership.team_id, user_id=user.id
    ).first()
    if existing_membership is not None:
        existing_membership.role = role
        flash(f"{email} already on team — role updated to {role} and account reactivated.", "success")
    else:
        db.session.add(
            TeamMembership(team_id=membership.team_id, user_id=user.id, role=role)
        )
        flash(f"{email} added to team as {role}.", "success")

    db.session.commit()
    return redirect(url_for("team"))


@app.route("/firms")
@login_required
def firms():
    membership = _require_team_scope()
    firm_rows = _accessible_firms_for_team(membership.team_id)
    stats = {}
    for firm in firm_rows:
        deal_rows = (
            db.session.query(Deal)
            .filter(Deal.firm_id == firm.id)
            .order_by(Deal.created_at.desc())
            .all()
        )
        fund_count = len({d.fund_number for d in deal_rows if d.fund_number})
        credit_count = db.session.query(CreditLoan).filter_by(firm_id=firm.id).count()
        stats[firm.id] = {
            "deal_count": len(deal_rows),
            "fund_count": fund_count,
            "credit_count": credit_count,
            "last_updated": deal_rows[0].created_at if deal_rows else None,
        }

    return render_template(
        "firms.html",
        firm_rows=firm_rows,
        firm_stats=stats,
        active_firm_id=_active_firm_id_from_session(),
    )


@app.route("/firms/<int:firm_id>/export-excel")
@login_required
def export_firm_excel(firm_id):
    membership = _require_team_scope()
    accessible_ids = {f.id for f in _accessible_firms_for_team(membership.team_id)}
    if firm_id not in accessible_ids:
        flash("Firm not accessible.", "warning")
        return redirect(url_for("firms"))

    firm = db.session.get(Firm, firm_id)
    if firm is None:
        abort(404)

    from services.excel_exporter import export_firm_to_excel

    buffer = export_firm_to_excel(firm_id, membership.team_id)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{firm.name} Data Export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/firms/<int:firm_id>/select", methods=["POST"])
@login_required
def select_firm_scope(firm_id):
    membership = _require_team_scope()
    accessible_ids = {firm.id for firm in _accessible_firms_for_team(membership.team_id)}
    if firm_id not in accessible_ids:
        flash("Selected firm is not available for your team.", "warning")
        return redirect(request.referrer or url_for("firms"))

    firm = db.session.get(Firm, firm_id)
    if firm is None:
        flash("Selected firm was not found.", "warning")
        return redirect(request.referrer or url_for("dashboard"))

    _set_active_firm_scope(firm.id)
    flash(f"Switched active firm to {firm.name}.", "success")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/firms/<int:firm_id>/delete", methods=["POST"])
@login_required
def delete_firm(firm_id):
    """Delete a firm and ALL associated data across every table."""
    try:
        membership = _require_team_scope()
        team_id = membership.team_id
        accessible_ids = {f.id for f in _accessible_firms_for_team(team_id)}
        if firm_id not in accessible_ids:
            flash("You do not have access to this firm.", "danger")
            return redirect(url_for("firms"))

        firm = db.session.get(Firm, firm_id)
        if firm is None:
            flash("Firm not found.", "danger")
            return redirect(url_for("firms"))

        firm_name = firm.name

        # Use raw SQL for the cascade delete. This avoids ORM issues with
        # tables that might not exist on older Postgres schemas, and handles
        # FK ordering in a single explicit sequence.
        from sqlalchemy import text as sa_text
        conn = db.session.connection()

        # Credit cascade
        conn.execute(sa_text(
            "DELETE FROM credit_loan_snapshots WHERE credit_loan_id IN "
            "(SELECT id FROM credit_loans WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM credit_loans WHERE firm_id = :fid"), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM credit_fund_performance WHERE firm_id = :fid"), {"fid": firm_id})

        # PE cascade: children first
        conn.execute(sa_text(
            "DELETE FROM deal_cashflow_events WHERE deal_id IN "
            "(SELECT id FROM deals WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text(
            "DELETE FROM deal_quarter_snapshots WHERE deal_id IN "
            "(SELECT id FROM deals WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text(
            "DELETE FROM deal_underwrite_baselines WHERE deal_id IN "
            "(SELECT id FROM deals WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM deals WHERE firm_id = :fid"), {"fid": firm_id})

        # Fund data
        conn.execute(sa_text("DELETE FROM fund_quarter_snapshots WHERE firm_id = :fid"), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM fund_metadata WHERE firm_id = :fid"), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM fund_cashflows WHERE firm_id = :fid"), {"fid": firm_id})

        # Memo cascade: deepest children first
        conn.execute(sa_text(
            "DELETE FROM memo_generation_claims WHERE run_id IN "
            "(SELECT id FROM memo_generation_runs WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text(
            "DELETE FROM memo_generation_sections WHERE run_id IN "
            "(SELECT id FROM memo_generation_runs WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text(
            "DELETE FROM memo_jobs WHERE run_id IN "
            "(SELECT id FROM memo_generation_runs WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM memo_generation_runs WHERE firm_id = :fid"), {"fid": firm_id})
        conn.execute(sa_text(
            "DELETE FROM memo_style_exemplars WHERE document_id IN "
            "(SELECT id FROM memo_documents WHERE firm_id = :fid)"
        ), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM memo_document_chunks WHERE firm_id = :fid"), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM memo_documents WHERE firm_id = :fid"), {"fid": firm_id})

        # Upload issues
        conn.execute(sa_text("DELETE FROM upload_issues WHERE firm_id = :fid"), {"fid": firm_id})

        # Access record and the firm itself
        conn.execute(sa_text("DELETE FROM team_firm_access WHERE firm_id = :fid"), {"fid": firm_id})
        conn.execute(sa_text("DELETE FROM firms WHERE id = :fid"), {"fid": firm_id})

        db.session.commit()

        # If the deleted firm was the active scope, clear it
        if session.get("active_firm_id") == firm_id:
            session.pop("active_firm_id", None)

        flash(f"Deleted firm '{firm_name}' and all associated data.", "success")
    except Exception as exc:
        db.session.rollback()
        logger.exception("Delete firm %s failed", firm_id)
        flash(f"Delete failed: {type(exc).__name__}: {str(exc)[:300]}", "danger")

    return redirect(url_for("firms"))


@app.route("/funds")
@login_required
def funds():
    flash("Manage Funds has been retired. Use Manage Firms to switch analytics scope.", "info")
    return redirect(url_for("firms"))


@app.route("/funds/<path:fund_name>/select", methods=["POST"])
@login_required
def select_fund_scope(fund_name):
    flash("Global fund scope switching has been retired. Use page filters for fund cuts.", "info")
    return redirect(request.referrer or url_for("firms"))


@app.route("/funds/<path:fund_name>/delete", methods=["POST"])
@login_required
def delete_fund(fund_name):
    return (
        "Fund deletion endpoint has been retired. Replace data through uploads at firm scope.",
        410,
    )


@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    try:
        filter_ctx = _build_filtered_deals_context()
        membership = filter_ctx.get("active_membership")
        payload = _build_dashboard_payload(
            filter_ctx["deals"],
            team_id=membership.team_id if membership is not None else None,
            benchmark_asset_class=filter_ctx.get("current_benchmark_asset_class", ""),
            metrics_by_id=filter_ctx.get("metrics_by_id"),
        )
        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        _scale_dashboard_payload(payload, reporting["money_scale"])
        return render_template(
            "dashboard.html",
            deals=filter_ctx["deals"],
            kpis=payload["kpis"],
            loss=payload["loss"],
            moic_distribution=payload["moic_distribution"],
            entry_exit_summary=payload["entry_exit_summary"],
            bridge_aggregate=payload["bridge_aggregate"],
            vintage_series=payload["vintage_series"],
            deal_metrics=payload["deal_metrics"],
            data_quality=payload["data_quality"],
            fund_summary_rows=payload["fund_summary_rows"],
            funds=filter_ctx["funds"],
            statuses=filter_ctx["statuses"],
            sectors=filter_ctx["sectors"],
            geographies=filter_ctx["geographies"],
            vintages=filter_ctx["vintages"],
            exit_types=filter_ctx["exit_types"],
            lead_partners=filter_ctx["lead_partners"],
            security_types=filter_ctx["security_types"],
            deal_types=filter_ctx["deal_types"],
            entry_channels=filter_ctx["entry_channels"],
            takeaway=payload["takeaway"],
            coverage=payload["coverage"],
            confidence=payload["confidence"],
            risk_flags=payload["risk_flags"],
            current_fund=filter_ctx["current_fund"],
            current_status=filter_ctx["current_status"],
            current_sector=filter_ctx["current_sector"],
            current_geography=filter_ctx["current_geography"],
            current_vintage=filter_ctx["current_vintage"],
            current_exit_type=filter_ctx["current_exit_type"],
            current_lead_partner=filter_ctx["current_lead_partner"],
            current_security_type=filter_ctx["current_security_type"],
            current_deal_type=filter_ctx["current_deal_type"],
            current_entry_channel=filter_ctx["current_entry_channel"],
            display_as_of_date=filter_ctx["display_as_of_date"],
            benchmark_asset_classes=filter_ctx["benchmark_asset_classes"],
            current_benchmark_asset_class=filter_ctx["current_benchmark_asset_class"],
        )
    except SQLAlchemyError as exc:
        if not _recover_missing_tables(exc):
            flash(_schema_upgrade_message(), "danger")
            return render_template("dashboard.html", **_empty_dashboard_context()), 503

        filter_ctx = _build_filtered_deals_context()
        membership = filter_ctx.get("active_membership")
        payload = _build_dashboard_payload(
            filter_ctx["deals"],
            team_id=membership.team_id if membership is not None else None,
            benchmark_asset_class=filter_ctx.get("current_benchmark_asset_class", ""),
            metrics_by_id=filter_ctx.get("metrics_by_id"),
        )
        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        _scale_dashboard_payload(payload, reporting["money_scale"])
        return render_template(
            "dashboard.html",
            deals=filter_ctx["deals"],
            kpis=payload["kpis"],
            loss=payload["loss"],
            moic_distribution=payload["moic_distribution"],
            entry_exit_summary=payload["entry_exit_summary"],
            bridge_aggregate=payload["bridge_aggregate"],
            vintage_series=payload["vintage_series"],
            deal_metrics=payload["deal_metrics"],
            data_quality=payload["data_quality"],
            fund_summary_rows=payload["fund_summary_rows"],
            funds=filter_ctx["funds"],
            statuses=filter_ctx["statuses"],
            sectors=filter_ctx["sectors"],
            geographies=filter_ctx["geographies"],
            vintages=filter_ctx["vintages"],
            exit_types=filter_ctx["exit_types"],
            lead_partners=filter_ctx["lead_partners"],
            security_types=filter_ctx["security_types"],
            deal_types=filter_ctx["deal_types"],
            entry_channels=filter_ctx["entry_channels"],
            takeaway=payload["takeaway"],
            coverage=payload["coverage"],
            confidence=payload["confidence"],
            risk_flags=payload["risk_flags"],
            current_fund=filter_ctx["current_fund"],
            current_status=filter_ctx["current_status"],
            current_sector=filter_ctx["current_sector"],
            current_geography=filter_ctx["current_geography"],
            current_vintage=filter_ctx["current_vintage"],
            current_exit_type=filter_ctx["current_exit_type"],
            current_lead_partner=filter_ctx["current_lead_partner"],
            current_security_type=filter_ctx["current_security_type"],
            current_deal_type=filter_ctx["current_deal_type"],
            current_entry_channel=filter_ctx["current_entry_channel"],
            display_as_of_date=filter_ctx["display_as_of_date"],
            benchmark_asset_classes=filter_ctx["benchmark_asset_classes"],
            current_benchmark_asset_class=filter_ctx["current_benchmark_asset_class"],
        )
    except Exception as exc:
        logger.exception("Dashboard computation failed")
        flash(f"Error computing dashboard metrics: {str(exc)}", "danger")
        return render_template("dashboard.html", **_empty_dashboard_context())


@app.route("/api/dashboard/series")
@login_required
def dashboard_series_api():
    try:
        filter_ctx = _build_filtered_deals_context()
        membership = filter_ctx.get("active_membership")
        payload = _build_dashboard_payload(
            filter_ctx["deals"],
            team_id=membership.team_id if membership is not None else None,
            benchmark_asset_class=filter_ctx.get("current_benchmark_asset_class", ""),
            metrics_by_id=filter_ctx.get("metrics_by_id"),
        )
        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        _scale_dashboard_payload(payload, reporting["money_scale"])
    except SQLAlchemyError as exc:
        return _json_schema_failure(exc, "Dashboard series computation failed")
    return jsonify(
        {
            "kpis": payload["kpis"],
            "loss_ratios": payload["loss"],
            "moic_distribution": payload["moic_distribution"],
            "entry_exit_summary": payload["entry_exit_summary"],
            "bridge_aggregate": payload["bridge_aggregate"],
            "vintage_series": payload["vintage_series"],
            "moic_hold_scatter": payload["moic_hold_scatter"],
            "value_creation_mix": payload["value_creation_mix"],
            "realized_unrealized_exposure": payload["realized_unrealized_exposure"],
            "loss_concentration_heatmap": payload["loss_concentration_heatmap"],
            "exit_type_performance": payload["exit_type_performance"],
            "lead_partner_scorecard": payload["lead_partner_scorecard"],
            "takeaway": payload["takeaway"],
            "coverage": payload["coverage"],
            "confidence": payload["confidence"],
            "risk_flags": payload["risk_flags"],
        }
    )


@app.route("/analysis/<page>")
@login_required
def analysis_page(page):
    if page not in ANALYSIS_PAGES:
        abort(404)

    if page == "chart-builder":
        try:
            filter_ctx = _build_filtered_deals_context()
            membership = filter_ctx.get("active_membership")
            team_id = membership.team_id if membership is not None else None
            catalog = build_chart_field_catalog(
                team_id=team_id,
                firm_id=filter_ctx.get("firm_id"),
                global_filters=_extract_chart_builder_global_filters_from_request(),
            )
            return render_template(
                "chart_builder.html",
                page_key=page,
                page_meta=ANALYSIS_PAGES[page],
                chart_builder_catalog=catalog,
                **filter_ctx,
            )
        except SQLAlchemyError as exc:
            return _redirect_schema_failure(exc, "Chart Builder page failed")

    if page == "fund-comparison":
        try:
            membership = _current_membership()
            if membership is None:
                abort(403)
            team_id = membership.team_id

            accessible_firms = _accessible_firms_for_team(team_id)
            accessible_firm_ids = {f.id for f in accessible_firms}
            firm_name_map = {f.id: f.name for f in accessible_firms}

            raw_firm_ids = request.args.getlist("firm_ids", type=int)
            selected_firm_ids = [fid for fid in raw_firm_ids if fid in accessible_firm_ids]

            vintage_filter_str = (request.args.get("vintage", "") or "").strip()
            vintage_filter = None
            if vintage_filter_str:
                try:
                    vintage_filter = int(vintage_filter_str)
                except (TypeError, ValueError):
                    pass
            metric_filter = (request.args.get("metric", "") or "").strip()

            firms_data = []
            all_vintages = set()
            for fid in selected_firm_ids:
                deals = build_deal_scope_query(team_id=team_id, firm_id=fid).all()
                fvl = build_fund_vintage_lookup(deals, team_id=team_id, firm_id=fid)
                all_vintages.update(v for v in fvl.values() if v is not None)
                firms_data.append({
                    "firm_id": fid,
                    "firm_name": firm_name_map[fid],
                    "deals": deals,
                    "fund_vintage_lookup": fvl,
                })

            payload = compute_fund_performance_comparison(
                firms_data,
                vintage_filter=vintage_filter,
                metric_filter=metric_filter,
            )

            return render_template(
                "analysis_fund_comparison.html",
                page_key=page,
                page_meta=ANALYSIS_PAGES[page],
                analysis=payload,
                available_firms=accessible_firms,
                selected_firm_ids=selected_firm_ids,
                selected_vintage=vintage_filter_str,
                selected_metric=metric_filter,
                available_vintages=sorted(all_vintages),
            )
        except SQLAlchemyError as exc:
            return _redirect_schema_failure(exc, "Fund comparison page failed")

    if page == "deal-comparison":
        try:
            from services.metrics.deal_comparison import compute_deal_level_comparison

            membership = _current_membership()
            if membership is None:
                abort(403)
            team_id = membership.team_id

            accessible_firms = _accessible_firms_for_team(team_id)
            accessible_firm_ids = {f.id for f in accessible_firms}
            firm_name_map = {f.id: f.name for f in accessible_firms}

            raw_firm_ids = request.args.getlist("firm_ids", type=int)
            selected_firm_ids = [fid for fid in raw_firm_ids if fid in accessible_firm_ids][:10]

            filters = {
                "sector": request.args.getlist("sector"),
                "geography": request.args.getlist("geography"),
                "status": request.args.getlist("status"),
                "exit_type": request.args.getlist("exit_type"),
                "vintage": [int(v) for v in request.args.getlist("vintage") if v.isdigit()],
            }
            filters = {k: v for k, v in filters.items() if v}

            page_num = request.args.get("page", 1, type=int)
            per_page = 50

            firms_data = []
            for fid in selected_firm_ids:
                deals = build_deal_scope_query(team_id=team_id, firm_id=fid).all()
                fvl = build_fund_vintage_lookup(deals, team_id=team_id, firm_id=fid)
                firms_data.append({
                    "firm_id": fid,
                    "firm_name": firm_name_map.get(fid, "Unknown"),
                    "deals": deals,
                    "fund_vintage_lookup": fvl,
                })

            payload = compute_deal_level_comparison(firms_data, filters=filters or None)

            all_rows = payload["deal_rows"]
            total_deals = len(all_rows)
            total_pages = max(1, (total_deals + per_page - 1) // per_page)
            page_num = max(1, min(page_num, total_pages))
            start = (page_num - 1) * per_page
            paged_rows = all_rows[start:start + per_page]

            return render_template(
                "analysis_deal_comparison.html",
                page_key=page,
                page_meta=ANALYSIS_PAGES[page],
                analysis=payload,
                deal_rows=paged_rows,
                available_firms=accessible_firms,
                selected_firm_ids=selected_firm_ids,
                selected_filters=filters,
                current_page=page_num,
                total_pages=total_pages,
                total_deals=total_deals,
            )
        except SQLAlchemyError as exc:
            return _redirect_schema_failure(exc, "Deal comparison page failed")

    try:
        filter_ctx = _build_filtered_deals_context()
        membership = filter_ctx.get("active_membership")
        payload = _analysis_route_payload(
            page,
            filter_ctx["deals"],
            firm_id=filter_ctx["firm_id"],
            team_id=membership.team_id if membership is not None else None,
            benchmark_asset_class=filter_ctx.get("current_benchmark_asset_class", ""),
            metrics_by_id=filter_ctx.get("metrics_by_id"),
        )
        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        _scale_analysis_payload(page, payload, reporting["money_scale"])
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, f"Analysis page '{page}' failed")

    template_name = "analysis_page.html"
    if page == "executive-summary":
        template_name = "analysis_executive_summary.html"
    elif page == "vca-ebitda":
        template_name = "analysis_vca_ebitda.html"
    elif page == "vca-revenue":
        template_name = "analysis_vca_revenue.html"
    elif page == "vca-addons":
        template_name = "analysis_vca_ebitda.html"
    elif page == "vca-addons-revenue":
        template_name = "analysis_vca_ebitda.html"
    elif page == "benchmarking":
        template_name = "analysis_benchmarking.html"
    elif page == "nav-at-risk":
        template_name = "analysis_nav_at_risk.html"
    elif page == "public-market-comparison":
        template_name = "analysis_public_market_comparison.html"
    elif page == "lp-due-diligence-memo":
        template_name = "analysis_lp_due_diligence_memo.html"
    elif page == "organic-growth":
        template_name = "analysis_organic_growth.html"
    elif page == "data-cuts":
        template_name = "analysis_data_cuts.html"

    return render_template(
        template_name,
        page_key=page,
        page_meta=ANALYSIS_PAGES[page],
        analysis=payload,
        **filter_ctx,
    )


# ---------------------------------------------------------------------------
# Private Credit routes
# ---------------------------------------------------------------------------


@app.route("/credit/analysis/<page>")
@login_required
def credit_analysis_page(page):
    if page not in CREDIT_ANALYSIS_PAGES:
        abort(404)

    try:
        from peqa.services.credit_filtering import build_credit_analysis_context

        membership = _current_membership()
        if membership is None:
            abort(403)
        team_id = membership.team_id
        firm_id = _active_firm_id_from_session()
        benchmark_asset_classes = _benchmark_asset_classes_for_team(team_id)
        benchmark_session_key = "selected_benchmark_asset_class"
        requested_benchmark = request.args.get("benchmark_asset_class")
        if requested_benchmark is not None:
            current_benchmark_asset_class = (requested_benchmark or "").strip()
            if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
                current_benchmark_asset_class = ""
            session[benchmark_session_key] = current_benchmark_asset_class
        else:
            current_benchmark_asset_class = (session.get(benchmark_session_key, "") or "").strip()
            if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
                current_benchmark_asset_class = ""
                session[benchmark_session_key] = ""

        # Check TeamFirmAccess
        if firm_id:
            access = TeamFirmAccess.query.filter_by(team_id=team_id, firm_id=firm_id).first()
            if not access:
                abort(403)

        # Strip empty-string values (from unselected form dropdowns) so they
        # don't get applied as filter criteria that would match zero rows.
        filters = {}
        for k in request.args:
            vals = [v for v in request.args.getlist(k) if v not in ("", None)]
            if vals:
                filters[k] = vals
        ctx = build_credit_analysis_context(team_id=team_id, firm_id=firm_id, filters=filters)
        loans = ctx["loans"]
        metrics_by_id = ctx["metrics_by_id"]
        snapshots_by_loan = ctx["snapshots_by_loan"]
        fund_performance = ctx.get("fund_performance", {})

        payload = {}
        if page == "credit-track-record":
            payload = compute_credit_track_record(
                loans, metrics_by_id, fund_performance=fund_performance
            )
        elif page == "credit-benchmarking":
            payload = compute_credit_benchmarking_analysis(
                loans,
                fund_performance=fund_performance,
                benchmark_thresholds=_load_team_benchmark_thresholds(
                    team_id,
                    current_benchmark_asset_class,
                ),
                benchmark_asset_class=current_benchmark_asset_class,
            )
        elif page == "credit-concentration":
            payload = compute_credit_concentration(loans, metrics_by_id)
        elif page == "credit-fundamentals":
            payload = compute_credit_fundamentals(
                loans, metrics_by_id, snapshots_by_loan=snapshots_by_loan
            )
        elif page == "credit-pricing-trends":
            payload = compute_credit_pricing_trends(
                loans,
                metrics_by_id,
                primary_dim=request.args.get("dim", "sector"),
                time_group=request.args.get("time_group", "quarter"),
            )
        elif page == "credit-underwrite-outcome":
            payload = compute_credit_underwrite_outcome(
                loans, metrics_by_id, snapshots_by_loan=snapshots_by_loan
            )
        elif page == "credit-data-cuts":
            primary_dim = request.args.get("dim", "sector")
            secondary_dim = request.args.get("dim2", "")
            payload = compute_credit_data_cuts(
                loans, metrics_by_id,
                primary_dim=primary_dim,
                secondary_dim=secondary_dim or None,
            )

        template_map = {
            "credit-track-record": "analysis_credit_track_record.html",
            "credit-benchmarking": "analysis_credit_benchmarking.html",
            "credit-concentration": "analysis_credit_concentration.html",
            "credit-fundamentals": "analysis_credit_fundamentals.html",
            "credit-pricing-trends": "analysis_credit_pricing_trends.html",
            "credit-underwrite-outcome": "analysis_credit_underwrite_outcome.html",
            "credit-data-cuts": "analysis_credit_data_cuts.html",
        }

        return render_template(
            template_map[page],
            page_key=page,
            page_meta=CREDIT_ANALYSIS_PAGES[page],
            analysis=payload,
            loans=loans,
            metrics_by_id=metrics_by_id,
            filter_options=ctx["filter_options"],
            active_filters=filters,
            loan_count=ctx["loan_count"],
            benchmark_asset_classes=benchmark_asset_classes,
            current_benchmark_asset_class=current_benchmark_asset_class,
        )
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, f"Credit analysis page '{page}' failed")


@app.route("/credit/api/analysis/<page>/series")
@login_required
def credit_analysis_series_api(page):
    if page not in CREDIT_ANALYSIS_PAGES:
        abort(404)

    try:
        from peqa.services.credit_filtering import build_credit_analysis_context

        membership = _current_membership()
        if membership is None:
            return jsonify({"error": "Not authenticated"}), 403
        team_id = membership.team_id
        firm_id = _active_firm_id_from_session()
        benchmark_asset_classes = _benchmark_asset_classes_for_team(team_id)
        benchmark_session_key = "selected_benchmark_asset_class"
        requested_benchmark = request.args.get("benchmark_asset_class")
        if requested_benchmark is not None:
            current_benchmark_asset_class = (requested_benchmark or "").strip()
            if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
                current_benchmark_asset_class = ""
            session[benchmark_session_key] = current_benchmark_asset_class
        else:
            current_benchmark_asset_class = (session.get(benchmark_session_key, "") or "").strip()
            if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
                current_benchmark_asset_class = ""
                session[benchmark_session_key] = ""
        # Strip empty-string values (from unselected form dropdowns) so they
        # don't get applied as filter criteria that would match zero rows.
        filters = {}
        for k in request.args:
            vals = [v for v in request.args.getlist(k) if v not in ("", None)]
            if vals:
                filters[k] = vals
        ctx = build_credit_analysis_context(team_id=team_id, firm_id=firm_id, filters=filters)
        loans = ctx["loans"]
        metrics_by_id = ctx["metrics_by_id"]
        fund_performance = ctx.get("fund_performance", {})

        payload = {}
        if page == "credit-track-record":
            payload = compute_credit_track_record(
                loans, metrics_by_id, fund_performance=fund_performance
            )
        elif page == "credit-benchmarking":
            payload = compute_credit_benchmarking_analysis(
                loans,
                fund_performance=fund_performance,
                benchmark_thresholds=_load_team_benchmark_thresholds(
                    team_id,
                    current_benchmark_asset_class,
                ),
                benchmark_asset_class=current_benchmark_asset_class,
            )
        elif page == "credit-concentration":
            payload = compute_credit_concentration(loans, metrics_by_id)
        elif page == "credit-fundamentals":
            payload = compute_credit_fundamentals(
                loans, metrics_by_id, snapshots_by_loan=ctx["snapshots_by_loan"]
            )
        elif page == "credit-pricing-trends":
            payload = compute_credit_pricing_trends(
                loans,
                metrics_by_id,
                primary_dim=request.args.get("dim", "sector"),
                time_group=request.args.get("time_group", "quarter"),
            )
        elif page == "credit-underwrite-outcome":
            payload = compute_credit_underwrite_outcome(
                loans, metrics_by_id, snapshots_by_loan=ctx["snapshots_by_loan"]
            )
        elif page == "credit-data-cuts":
            primary_dim = request.args.get("dim", "sector")
            secondary_dim = request.args.get("dim2", "")
            payload = compute_credit_data_cuts(
                loans, metrics_by_id,
                primary_dim=primary_dim,
                secondary_dim=secondary_dim or None,
            )

        return jsonify({"page": page, "title": CREDIT_ANALYSIS_PAGES[page]["title"], "payload": payload})
    except SQLAlchemyError as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/upload/credit-loans", methods=["GET", "POST"])
@login_required
def upload_credit_loans():
    try:
        ensure_schema_updates()
    except Exception:
        logger.exception("Credit upload schema preparation failed")

    if request.method == "GET":
        membership = _current_membership()
        team_id = membership.team_id if membership else None
        firms = _accessible_firms_for_team(team_id) if team_id else []

        # Build per-firm credit upload summary
        credit_firms = []
        if team_id:
            try:
                from sqlalchemy import func as sa_func

                firm_stats = (
                    db.session.query(
                        CreditLoan.firm_id,
                        Firm.name,
                        sa_func.count(CreditLoan.id).label("loan_count"),
                        sa_func.count(sa_func.distinct(CreditLoan.fund_name)).label("fund_count"),
                    )
                    .join(Firm, Firm.id == CreditLoan.firm_id)
                    .filter(CreditLoan.team_id == team_id)
                    .group_by(CreditLoan.firm_id, Firm.name)
                    .order_by(Firm.name)
                    .all()
                )
                for row in firm_stats:
                    fund_names = [
                        r[0]
                        for r in db.session.query(sa_func.distinct(CreditLoan.fund_name))
                        .filter(CreditLoan.firm_id == row.firm_id, CreditLoan.team_id == team_id)
                        .all()
                    ]
                    credit_firms.append({
                        "firm_id": row.firm_id,
                        "firm_name": row.name,
                        "loan_count": row.loan_count,
                        "fund_count": row.fund_count,
                        "fund_names": sorted(fund_names),
                    })
            except Exception:
                db.session.rollback()

        return render_template(
            "upload_credit.html",
            accessible_firms=firms,
            credit_firms=credit_firms,
        )

    try:
        from services.credit_parser import parse_credit_loan_tape

        membership = _current_membership()
        if membership is None:
            abort(403)
        team_id = membership.team_id

        firm_name = (request.form.get("firm_name") or "").strip()
        if not firm_name:
            flash("Firm name is required.", "danger")
            return redirect(url_for("upload_credit_loans"))

        file = request.files.get("file")
        if not file or not file.filename:
            flash("No file selected.", "danger")
            return redirect(url_for("upload_credit_loans"))

        result = parse_credit_loan_tape(
            file_stream=file,
            firm_name=firm_name,
            team_id=team_id,
        )

        # Switch active firm to the one just uploaded so analysis pages show its data
        if result.get("firm_id"):
            _set_active_firm_scope(result["firm_id"])

        msg = f"Uploaded {result['loans']} credit loans across {len(result['funds'])} fund(s) for {firm_name}."
        if result.get("snapshots", 0) > 0:
            msg += f" {result['snapshots']} quarterly snapshots imported."
        if result.get("warnings", 0) > 0:
            msg += f" {result['warnings']} warning(s)."
        flash(msg, "success")
        return redirect(url_for("credit_analysis_page", page="credit-track-record"))
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("upload_credit_loans"))
    except SQLAlchemyError as exc:
        if _is_missing_table_error(exc):
            return _redirect_schema_failure(exc, "Credit loan upload failed", endpoint="upload_credit_loans")
        db.session.rollback()
        flash(f"Upload failed: {exc}", "danger")
        return redirect(url_for("upload_credit_loans"))
    except Exception as exc:
        db.session.rollback()
        logger.exception("Unexpected credit loan upload failure")
        flash(
            "Upload failed due to an unexpected workbook or server issue. "
            f"Detail: {type(exc).__name__}: {str(exc)[:300]}",
            "danger",
        )
        return redirect(url_for("upload_credit_loans"))


@app.route("/upload/credit-loans/template")
@login_required
def download_credit_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # --- Main Loans sheet ---
    ws = wb.active
    ws.title = "Credit Loans"

    headers = [
        # Core identification
        "Company Name", "Fund Name", "Entry Date", "Exit Date",
        "Vintage Year", "As Of Date", "Fund Size",
        # Company details
        "Sector", "Geography", "Sponsor", "Security Type",
        "Sourcing Channel", "Business Description", "Public",
        "Count of Investments",
        # Loan structure
        "Hold Size", "Committed", "Current Invested Capital",
        "Issue Size", "Instrument", "Tranche",
        # Loan economics
        "Coupon Rate", "Spread (bps)", "Floor Rate",
        "Fixed or Floating", "Reference Rate",
        "PIK", "PIK Rate",
        "OID", "Upfront Fee", "Exit Fee",
        "Maturity Date", "Loan Term", "Term (years)",
        "Amortization Type", "Payment Frequency",
        # Protections
        "Call Protection (months)", "Make-Whole Premium",
        "Prepayment Protection",
        # Credit metrics
        "Entry LTV", "Current LTV",
        "Interest Coverage Ratio", "DSCR",
        "Default Status", "Credit Rating",
        "Covenant Type", "Covenant Compliant",
        # Returns & valuation
        "Gross IRR", "Estimated IRR at Entry", "MOIC",
        "Realized Value", "Unrealized Value",
        "Unrealized Warrant/Equity Value", "Total Value",
        "Fair Value", "Yield to Maturity", "Recovery Rate",
        # Revenue & EBITDA
        "Entry Revenue", "Current Revenue",
        "Entry EBITDA", "Current EBITDA",
        # Income
        "Cumulative Interest Income", "Cumulative Fee Income",
        # Par & outstanding
        "Original Par", "Current Outstanding", "Accrued Interest",
        # Collateral & coverage
        "Entry Collateral", "Current Collateral",
        "Entry Coverage Ratio", "Current Coverage Ratio",
        "Entry Equity Cushion", "Current Equity Cushion",
        # Warrants & equity
        "Equity Investment",
        "Warrants at Entry", "Warrant Strike (entry)",
        "Warrants (current)", "Warrant Strike (current)", "Warrant Term",
        # Currency
        "Currency",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0A6B58", end_color="0A6B58", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Example row — one value per header, in exact header order
    ws.append([
        # Core identification
        "Acme Software Inc", "PCOF III", "2023-06-15", None,
        2023, "2024-12-31", 500.0,
        # Company details
        "Software", "North America", "Apollo Global", "Senior Secured",
        "Direct", "Enterprise SaaS platform for healthcare", "No",
        1,
        # Loan structure
        25.0, 30.0, 24.5,
        100.0, "Term Loan B", "First Lien",
        # Loan economics
        0.085, 425, 0.015,
        "Floating", "SOFR",
        "No", None,
        0.02, 0.005, None,
        "2028-06-15", "5 years", 5,
        "Bullet", "Quarterly",
        # Protections
        12, None,
        "12-month no-call",
        # Credit metrics
        0.55, 0.58,
        2.5, 1.8,
        "Performing", 2,
        "Maintenance", "Yes",
        # Returns & valuation
        0.12, 0.15, 1.10,
        None, 26.0,
        1.5, 27.5,
        25.5, 0.09, None,
        # Revenue & EBITDA
        50.0, 55.0,
        15.0, 17.0,
        # Income
        4.0, 0.5,
        # Par & outstanding
        25.0, 25.0, 0.2,
        # Collateral & coverage
        40.0, 42.0,
        1.6, 1.68,
        0.25, 0.28,
        # Warrants & equity
        0.5,
        50000, 10.0,
        50000, 12.5, "10 years",
        # Currency
        "USD",
    ])

    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    # --- Snapshots sheet ---
    ws2 = wb.create_sheet("Snapshots")
    snap_headers = [
        "Company Name", "Snapshot Date",
        "Current LTV", "Fair Value", "Current Revenue", "Current EBITDA",
        "Interest Coverage Ratio", "DSCR",
        "Default Status", "Credit Rating", "Covenant Compliant",
        "Current Outstanding", "Accrued Interest",
    ]
    ws2.append(snap_headers)
    for col_idx, _ in enumerate(snap_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws2.append([
        "Acme Software Inc", "2024-03-31",
        0.57, 25.2, 52.0, 16.0,
        2.4, 1.7,
        "Performing", 2, "Yes",
        25.0, 0.15,
    ])
    ws2.append([
        "Acme Software Inc", "2024-06-30",
        0.55, 25.4, 54.0, 16.5,
        2.5, 1.8,
        "Performing", 2, "Yes",
        25.0, 0.18,
    ])

    for col_idx in range(1, len(snap_headers) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 18

    # --- Fund Performance sheet ---
    ws_fp = wb.create_sheet("Fund Performance")
    fp_headers = [
        "Fund Name", "Vintage Year", "Fund Size",
        "Net IRR", "Net MOIC", "DPI", "RVPI", "TVPI",
        "Called Capital", "Distributed Capital", "NAV",
        "Report Date", "Currency",
    ]
    ws_fp.append(fp_headers)
    for col_idx, _ in enumerate(fp_headers, 1):
        cell = ws_fp.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws_fp.append([
        "PCOF III", 2021, 500.0,
        0.12, 1.35, 0.45, 0.90, 1.35,
        425.0, 190.0, 380.0,
        "2024-12-31", "USD",
    ])
    ws_fp.append([
        "PCOF IV", 2023, 750.0,
        0.08, 1.10, 0.15, 0.95, 1.10,
        300.0, 45.0, 712.0,
        "2024-12-31", "USD",
    ])

    for col_idx in range(1, len(fp_headers) + 1):
        ws_fp.column_dimensions[ws_fp.cell(row=1, column=col_idx).column_letter].width = 18

    # --- Instructions sheet ---
    ws3 = wb.create_sheet("Instructions")
    ws3.column_dimensions["A"].width = 80
    instructions = [
        "PE Portfolio Lab - Credit Loan Template",
        "",
        "CREDIT LOANS SHEET (required)",
        "  Required columns: Company Name, Fund Name, Entry Date",
        "  All other columns are optional. Include what your data has.",
        "",
        "  FIELD GUIDE",
        "  Status: AUTO-CALCULATED (Fully Realized / Partially Realized / Unrealized)",
        "    Based on realized/unrealized values. Do not fill in manually.",
        "",
        "  Dollar amounts: In millions (e.g., 25.0 = $25M)",
        "    Hold Size, Committed, Issue Size, Revenue, EBITDA, Collateral, etc.",
        "",
        "  Rates & percentages: Decimal form (e.g., 0.085 = 8.5%)",
        "    Coupon Rate, Floor Rate, Gross IRR, LTV, Equity Cushion, etc.",
        "    LTV: Values > 1 are auto-normalized (55 becomes 0.55)",
        "",
        "  Spread: In basis points (e.g., 425 = 4.25%)",
        "  Credit Rating: 1 (best) to 5 (worst)",
        "  Coverage Ratio: Multiple (e.g., 1.6 = collateral covers 1.6x the loan)",
        "  MOIC: Multiple (e.g., 1.15x)",
        "",
        "  Yes/No fields: PIK, Covenant Compliant, Public",
        "    Accepts: Yes/No, True/False, Y/N, 1/0",
        "",
        "  Text fields with specific values:",
        "    Default Status: Performing, Watch List, Default, or Restructured",
        "    Fixed or Floating: Fixed or Floating",
        "    Covenant Type: Maintenance, Incurrence, or None",
        "    Currency: ISO 3-letter code (e.g., USD, EUR, GBP)",
        "    Loan Term: text (e.g., '5 years', '60 months')",
        "",
        "  Dates: YYYY-MM-DD or Excel date format",
        "",
        "SNAPSHOTS SHEET (optional)",
        "  Quarterly updates for metrics that change over time.",
        "  Match Company Name exactly to the Loans sheet.",
        "  One row per company per quarter.",
        "",
        "FUND PERFORMANCE SHEET (optional)",
        "  Fund-level net returns shown on the Credit Track Record page.",
        "  One row per fund. Overwrites previous fund performance on re-upload.",
        "  Net IRR: decimal (0.12 = 12%). Values > 1.5 auto-divided by 100.",
        "  Net MOIC / DPI / RVPI / TVPI: multiples (e.g., 1.35x)",
        "  Fund Size / Called Capital / Distributed Capital / NAV: dollar amounts (millions)",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws3.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(name="Calibri", bold=True, size=12)
        elif line.startswith("  "):
            cell.font = Font(name="Calibri", size=10, color="666666")
        else:
            cell.font = Font(name="Calibri", bold=True, size=10)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="credit_loan_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/upload/credit-loans/<int:firm_id>/delete", methods=["POST"])
@login_required
def delete_credit_firm_data(firm_id):
    membership = _current_membership()
    if membership is None:
        abort(403)
    team_id = membership.team_id

    access = TeamFirmAccess.query.filter_by(team_id=team_id, firm_id=firm_id).first()
    if not access:
        flash("You do not have access to this firm.", "danger")
        return redirect(url_for("upload_credit_loans"))

    firm = db.session.get(Firm, firm_id)
    firm_name = firm.name if firm else f"Firm #{firm_id}"

    try:
        snap_count = CreditLoanSnapshot.query.filter(
            CreditLoanSnapshot.credit_loan_id.in_(
                db.session.query(CreditLoan.id).filter_by(firm_id=firm_id, team_id=team_id)
            )
        ).delete(synchronize_session=False)
        loan_count = CreditLoan.query.filter_by(firm_id=firm_id, team_id=team_id).delete()
        fund_perf_count = CreditFundPerformance.query.filter_by(
            firm_id=firm_id, team_id=team_id
        ).delete(synchronize_session=False)
        db.session.commit()
        flash(
            f"Deleted {loan_count} credit loan(s), {snap_count} snapshot(s), and "
            f"{fund_perf_count} fund performance row(s) for {firm_name}.",
            "success",
        )
    except SQLAlchemyError as exc:
        db.session.rollback()
        flash(f"Delete failed: {exc}", "danger")

    return redirect(url_for("upload_credit_loans"))


@app.route("/api/analysis/<page>/series")
@login_required
def analysis_series_api(page):
    if page not in ANALYSIS_PAGES:
        abort(404)

    if page == "chart-builder":
        return jsonify({"page": page, "title": ANALYSIS_PAGES[page]["title"], "payload": {}})

    try:
        filter_ctx = _build_filtered_deals_context()
        membership = filter_ctx.get("active_membership")
        payload = _analysis_route_payload(
            page,
            filter_ctx["deals"],
            firm_id=filter_ctx["firm_id"],
            team_id=membership.team_id if membership is not None else None,
            benchmark_asset_class=filter_ctx.get("current_benchmark_asset_class", ""),
            metrics_by_id=filter_ctx.get("metrics_by_id"),
        )
        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        _scale_analysis_payload(page, payload, reporting["money_scale"])
    except SQLAlchemyError as exc:
        return _json_schema_failure(exc, f"Analysis series '{page}' failed")

    return jsonify(
        {
            "page": page,
            "title": ANALYSIS_PAGES[page]["title"],
            "payload": payload,
        }
    )


@app.route("/api/data-cuts/export")
@login_required
def data_cuts_export():
    from io import BytesIO
    from openpyxl import Workbook
    from services.metrics.data_cuts import ALLOWED_METRICS

    try:
        filter_ctx = _build_filtered_deals_context()
        metrics_by_id = filter_ctx.get("metrics_by_id") or {d.id: compute_deal_metrics(d) for d in filter_ctx["deals"]}

        primary_dim = request.args.get("dim", "sector")
        secondary_dim = request.args.get("dim2", "")
        metric = request.args.get("metric", "weighted_moic")
        if metric not in ALLOWED_METRICS:
            metric = "weighted_moic"

        payload = compute_data_cuts_analytics(
            filter_ctx["deals"], metrics_by_id,
            primary_dim=primary_dim,
            secondary_dim=secondary_dim or None,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Cuts"

        # Header row with as-of date
        if payload.get("as_of_date"):
            ws.append([f"Data Cuts — {payload['primary_dim_label']}", "", f"As of {payload['as_of_date']}"])
            ws.append([])

        if payload.get("cross_tab") and payload.get("secondary_labels"):
            # Cross-tab export: rows = primary dim, columns = secondary dim, cells = selected metric
            header = [payload["primary_dim_label"]] + payload["secondary_labels"] + ["Total"]
            ws.append(header)

            for group in payload["groups"]:
                row_data = [group["label"]]
                ct_row = payload["cross_tab"].get(group["label"], {})
                for sec_label in payload["secondary_labels"]:
                    cell = ct_row.get(sec_label)
                    row_data.append(cell.get(metric) if cell else None)
                row_data.append(group.get(metric))
                ws.append(row_data)

            # Totals row
            totals_row = ["Total"]
            col_totals = payload.get("col_totals") or {}
            for sec_label in payload["secondary_labels"]:
                ct = col_totals.get(sec_label)
                totals_row.append(ct.get(metric) if ct else None)
            totals_row.append(payload["totals"].get(metric))
            ws.append(totals_row)
        else:
            # Flat grouped table export
            header = [
                payload["primary_dim_label"], "Deals", "Invested Equity ($M)",
                "Realized Value ($M)", "Unrealized Value ($M)", "Total Value ($M)",
                "Weighted MOIC", "Weighted IRR",
                "Entry TEV/EBITDA (Wtd)", "Exit TEV/EBITDA (Wtd)",
                "Entry EBITDA Margin (Wtd)", "Exit EBITDA Margin (Wtd)",
                "Loss Ratio (Count)", "Loss Ratio (Capital)",
                "Hold Period (Wtd)", "% of Invested", "% of Total Value",
            ]
            ws.append(header)

            for group in payload["groups"]:
                ws.append([
                    group["label"],
                    group["deal_count"],
                    group["invested_equity"],
                    group["realized_value"],
                    group["unrealized_value"],
                    group["total_value"],
                    group["weighted_moic"],
                    group["weighted_irr"],
                    group["weighted_entry_tev_ebitda"],
                    group["weighted_exit_tev_ebitda"],
                    group["weighted_entry_ebitda_margin"],
                    group["weighted_exit_ebitda_margin"],
                    group["loss_ratio_count"],
                    group["loss_ratio_capital"],
                    group["weighted_hold_years"],
                    group["pct_of_invested"],
                    group["pct_of_total"],
                ])

            # Totals row
            t = payload["totals"]
            ws.append([
                "Total", t["deal_count"], t["invested_equity"],
                t["realized_value"], t["unrealized_value"], t["total_value"],
                t["weighted_moic"], t["weighted_irr"],
                t["weighted_entry_tev_ebitda"], t["weighted_exit_tev_ebitda"],
                t["weighted_entry_ebitda_margin"], t["weighted_exit_ebitda_margin"],
                t["loss_ratio_count"], t["loss_ratio_capital"],
                t["weighted_hold_years"], 1.0, 1.0,
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="data-cuts-export.xlsx"'},
        )
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Data cuts export failed")


@app.route("/api/data-cuts/summary-pdf")
@login_required
def data_cuts_summary_pdf():
    """Download a single PDF showing all 9 data-cut dimensions."""
    from io import BytesIO

    from services.metrics.data_cuts import DIMENSIONS, DIMENSION_LABELS, compute_data_cuts_analytics

    try:
        filter_ctx = _build_filtered_deals_context()
        deals = filter_ctx["deals"]
        metrics_by_id = filter_ctx.get("metrics_by_id") or {
            d.id: compute_deal_metrics(d) for d in deals
        }

        all_cuts = {}
        for dim_key in DIMENSIONS:
            all_cuts[dim_key] = compute_data_cuts_analytics(
                deals, metrics_by_id, primary_dim=dim_key
            )

        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        currency_code = reporting["reporting_currency_code"]
        active_firm = filter_ctx.get("active_firm")
        firm_name = (active_firm.name if active_firm else "Portfolio") or "Portfolio"
        first_payload = next(iter(all_cuts.values()), {})
        as_of_date = first_payload.get("as_of_date")
        report_title = _report_title(firm_name, "Data Cuts Summary", as_of_date)

        pdf_bytes = _build_data_cuts_summary_pdf(
            all_cuts=all_cuts,
            report_title=report_title,
            currency_code=currency_code,
        )

        download_name = _safe_pdf_download_name(
            f"{firm_name} Data Cuts Summary", "data_cuts_summary.pdf"
        )
        return send_file(
            BytesIO(pdf_bytes),
            as_attachment=True,
            download_name=download_name,
            mimetype="application/pdf",
        )
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Data cuts summary PDF failed")


@app.route("/api/analysis/executive-summary/llm-insights")
@login_required
@limiter.limit("10/minute")
def executive_summary_llm_insights():
    """AJAX endpoint: returns LLM-generated bullet-point insights."""
    from services.metrics.executive_summary_insights import generate_executive_insights

    try:
        filter_ctx = _build_filtered_deals_context()
        membership = filter_ctx.get("active_membership")
        rank_by = (request.args.get("rank_by", "") or "moic").strip()
        if rank_by not in ("moic", "irr"):
            rank_by = "moic"

        payload = compute_executive_summary_analysis(
            filter_ctx["deals"],
            metrics_by_id=filter_ctx.get("metrics_by_id"),
            firm_id=filter_ctx["firm_id"],
            team_id=membership.team_id if membership is not None else None,
            rank_by=rank_by,
        )
        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        _scale_analysis_payload("executive-summary", payload, reporting["money_scale"])

        filter_params = {
            "firm_id": filter_ctx["firm_id"],
            "team_id": membership.team_id if membership is not None else None,
            "fund": request.args.get("fund", ""),
            "status": request.args.get("status", ""),
            "sector": request.args.get("sector", ""),
            "geography": request.args.get("geography", ""),
            "rank_by": rank_by,
        }

        result = generate_executive_insights(payload, filter_params)
        return jsonify(result)
    except Exception as exc:
        logger.exception("LLM insights endpoint failed")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route("/api/chart-builder/catalog")
@login_required
def chart_builder_catalog_api():
    membership = _require_team_scope()
    active_firm = _resolve_active_firm_for_team()
    try:
        catalog = build_chart_field_catalog(
            team_id=membership.team_id,
            firm_id=active_firm.id if active_firm is not None else None,
            global_filters=_extract_chart_builder_global_filters_from_request(),
        )
    except SQLAlchemyError as exc:
        return _json_schema_failure(exc, "Chart Builder catalog failed")
    return jsonify(catalog)


@app.route("/api/chart-builder/query", methods=["POST"])
@login_required
@limiter.limit("60/minute")
def chart_builder_query_api():
    membership = _require_team_scope()
    active_firm = _resolve_active_firm_for_team()
    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify({"error": "Request body must be a JSON object."}), 400

    global_filters = body.get("global_filters")
    if not isinstance(global_filters, dict):
        global_filters = _extract_chart_builder_global_filters_from_request()

    try:
        payload = run_chart_query(
            spec=body,
            team_id=membership.team_id,
            firm_id=active_firm.id if active_firm is not None else None,
            global_filters=global_filters,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except SQLAlchemyError as exc:
        return _json_schema_failure(exc, "Chart Builder query failed")

    return jsonify(payload)


@app.route("/api/chart-builder/templates")
@login_required
def chart_builder_templates_api():
    membership = _require_team_scope()
    try:
        rows = (
            ChartBuilderTemplate.query.filter_by(team_id=membership.team_id)
            .order_by(ChartBuilderTemplate.updated_at.desc(), ChartBuilderTemplate.id.desc())
            .all()
        )
    except SQLAlchemyError as exc:
        return _json_schema_failure(exc, "Chart Builder templates load failed")
    return jsonify({"templates": [_serialize_chart_builder_template(row) for row in rows]})


@app.route("/api/chart-builder/templates", methods=["POST"])
@login_required
@limiter.limit("60/minute")
def chart_builder_template_create_api():
    membership = _require_team_scope()
    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify({"error": "Request body must be a JSON object."}), 400

    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Template name is required."}), 400

    source = (body.get("source") or "deals").strip().lower() or "deals"
    config = body.get("config")
    if config is None:
        config = {"config_version": 1, "cards": []}
    if not isinstance(config, dict):
        return jsonify({"error": "Template config must be an object."}), 400
    if "config_version" not in config:
        config["config_version"] = 1

    row = ChartBuilderTemplate(
        team_id=membership.team_id,
        name=name,
        source=source,
        config_json=json.dumps(config, sort_keys=True),
        created_by_user_id=current_user.id,
    )
    db.session.add(row)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Template name already exists for this team."}), 400
    return jsonify(_serialize_chart_builder_template(row)), 201


@app.route("/api/chart-builder/templates/<int:template_id>", methods=["PUT"])
@login_required
@limiter.limit("60/minute")
def chart_builder_template_update_api(template_id):
    membership = _require_team_scope()
    row = ChartBuilderTemplate.query.filter_by(id=template_id, team_id=membership.team_id).first()
    if row is None:
        abort(404)

    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify({"error": "Request body must be a JSON object."}), 400

    if "name" in body:
        name = (body.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Template name cannot be empty."}), 400
        row.name = name
    if "source" in body:
        row.source = (body.get("source") or "deals").strip().lower() or "deals"
    if "config" in body:
        config = body.get("config")
        if not isinstance(config, dict):
            return jsonify({"error": "Template config must be an object."}), 400
        if "config_version" not in config:
            config["config_version"] = 1
        row.config_json = json.dumps(config, sort_keys=True)
    row.updated_at = _utc_now_naive()

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Template name already exists for this team."}), 400
    return jsonify(_serialize_chart_builder_template(row))


@app.route("/api/chart-builder/templates/<int:template_id>", methods=["DELETE"])
@login_required
@limiter.limit("60/minute")
def chart_builder_template_delete_api(template_id):
    membership = _require_team_scope()
    row = ChartBuilderTemplate.query.filter_by(id=template_id, team_id=membership.team_id).first()
    if row is None:
        abort(404)
    db.session.delete(row)
    db.session.commit()
    return jsonify({"deleted": True, "id": template_id})


@app.route("/ic-memo")
@app.route("/ic-memo/<fund_name>")
@login_required
def ic_memo(fund_name=None):
    try:
        filter_ctx = _build_filtered_deals_context(fund_override=fund_name)
        metrics_by_id = {d.id: compute_deal_metrics(d) for d in filter_ctx["deals"]}
        payload = compute_ic_memo_payload(
            filter_ctx["deals"],
            metrics_by_id=metrics_by_id,
            ranking_basis="weighted_moic",
            decile_pct=0.10,
            decile_min=1,
        )
        reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
        _scale_ic_memo_payload(payload, reporting["money_scale"])
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "IC memo page failed")

    active_fund_scope = filter_ctx["current_fund"] or "All Funds"
    payload["meta"]["fund_scope"] = active_fund_scope
    payload["meta"]["filters_applied"] = {
        "fund": filter_ctx["current_fund"],
        "status": filter_ctx["current_status"],
        "sector": filter_ctx["current_sector"],
        "geography": filter_ctx["current_geography"],
        "vintage": filter_ctx["current_vintage"],
        "exit_type": filter_ctx["current_exit_type"],
        "lead_partner": filter_ctx["current_lead_partner"],
        "deal_type": filter_ctx["current_deal_type"],
        "entry_channel": filter_ctx["current_entry_channel"],
    }

    return render_template(
        "ic_memo.html",
        memo=payload,
        fund_scope_path=fund_name,
        **filter_ctx,
    )


@app.route("/methodology")
@login_required
def methodology():
    payload = build_methodology_payload()
    return render_template("methodology.html", methodology=payload)


@app.route("/audit")
@login_required
def methodology_alias():
    return redirect(url_for("methodology"))


@app.route("/api/fund-comparison/deals")
@login_required
def api_fund_comparison_deals():
    """Lazy-load deals for a specific firm+fund in the comparison drill-down."""
    from services.metrics.deal_comparison import _deal_row, _deal_moic, _deal_hold_years

    firm_id = request.args.get("firm_id", type=int)
    fund_name = request.args.get("fund_name", "")
    if firm_id is None or not fund_name:
        return jsonify({"deals": []}), 200

    membership = _current_membership()
    if membership is None:
        abort(403)
    team_id = membership.team_id

    accessible_ids = {f.id for f in _accessible_firms_for_team(team_id)}
    if firm_id not in accessible_ids:
        abort(403)

    try:
        deals = (
            build_deal_scope_query(team_id=team_id, firm_id=firm_id)
            .filter(Deal.fund_number == fund_name)
            .all()
        )
    except SQLAlchemyError as exc:
        return _json_schema_failure(exc, "Deal comparison drill-down query failed")

    firm = db.session.get(Firm, firm_id)
    fname = firm.name if firm else "Unknown"

    rows = []
    for deal in deals:
        rows.append(_deal_row(deal, firm_id, fname, {}))

    return jsonify({"deals": rows})


@app.route("/api/deals/<int:deal_id>/bridge")
@login_required
def deal_bridge_api(deal_id):
    active_firm = _resolve_active_firm_for_team()
    firm_id = active_firm.id if active_firm is not None else None
    try:
        deal = Deal.query.filter_by(id=deal_id, firm_id=firm_id).first()
    except SQLAlchemyError as exc:
        return _json_schema_failure(exc, "Deal bridge query failed")
    if deal is None:
        abort(404)

    model = request.args.get("model", "additive").lower()
    basis = "fund"
    unit = request.args.get("unit", "moic").lower()

    if model != "additive":
        return jsonify({"error": "Only additive model is supported"}), 400
    if request.args.get("basis") and request.args.get("basis", "fund").lower() != "fund":
        return jsonify({"error": "Only fund pro-rata basis is supported"}), 400
    if unit not in {"dollar", "moic", "pct"}:
        return jsonify({"error": "unit must be dollar, moic, or pct"}), 400

    warnings = []
    bridge = compute_bridge_view(deal, model="additive", basis=basis, unit=unit, warnings=warnings)
    reporting = _reporting_currency_context(active_firm)
    money_scale = reporting["money_scale"]
    _scale_bridge_view_payload(bridge, money_scale)

    equity = _scale_money(deal.equity_invested, money_scale)
    value_created = bridge.get("value_created")
    ownership = bridge.get("ownership_pct")

    if basis == "fund":
        start_dollar = equity if equity is not None else None
    else:
        start_dollar = (equity / ownership) if (equity is not None and ownership is not None and ownership > 0) else None

    end_dollar = (
        start_dollar + value_created
        if start_dollar is not None and value_created is not None
        else None
    )

    if unit == "dollar":
        start_value, end_value = start_dollar, end_dollar
    elif unit == "moic":
        start_value = 1.0
        display_drivers = bridge.get("display_drivers") or []
        if display_drivers and all(row.get("moic") is not None for row in display_drivers):
            end_value = start_value + sum(row.get("moic") or 0.0 for row in display_drivers)
        else:
            driver_vals = bridge.get("drivers", {})
            if all(driver_vals.get(k) is not None for k in ("revenue", "margin", "multiple", "leverage", "other")):
                end_value = start_value + sum(driver_vals[k] for k in ("revenue", "margin", "multiple", "leverage", "other"))
            else:
                end_value = None
    else:  # pct
        start_value = 0.0
        end_value = 1.0 if bridge.get("ready") else None

    return jsonify(
        {
            "deal_id": deal.id,
            "company": deal.company_name,
            "model": "additive",
            "basis": basis,
            "unit": unit,
            "ready": bridge.get("ready"),
            "ownership_pct": bridge.get("ownership_pct"),
            "equity_invested": equity,
            "start_value": start_value,
            "end_value": end_value,
            "start_dollar": start_dollar,
            "end_dollar": end_dollar,
            "drivers": bridge.get("drivers"),
            "drivers_dollar": bridge.get("drivers_dollar"),
            "display_drivers": bridge.get("display_drivers") or [],
            "value_created": bridge.get("value_created"),
            "fund_value_created": bridge.get("fund_value_created"),
            "company_value_created": bridge.get("company_value_created"),
            "calculation_method": bridge.get("calculation_method"),
            "fallback_reason": bridge.get("fallback_reason"),
            "diagnostics": bridge.get("diagnostics", {}),
            "warnings": warnings,
        }
    )


@app.route("/upload")
@login_required
def upload():
    membership = _require_team_scope()
    active_firm = _resolve_active_firm_for_team()
    upload_batches = _upload_batches_for_firm(active_firm.id if active_firm is not None else None)
    benchmark_dataset = _benchmark_dataset_for_team(membership.team_id)
    return render_template(
        "upload.html",
        upload_batches=upload_batches,
        active_firm=active_firm,
        benchmark_dataset=benchmark_dataset,
    )


@app.route("/upload/batches/<batch_id>/delete", methods=["POST"])
@login_required
def delete_upload_batch(batch_id):
    _require_team_scope()
    active_firm = _resolve_active_firm_for_team()
    if active_firm is None:
        flash("No active firm selected. Choose a firm first.", "warning")
        return redirect(url_for("upload"))

    normalized_batch = (batch_id or "").strip()
    if not normalized_batch:
        flash("Upload batch id is required.", "warning")
        return redirect(url_for("upload"))

    try:
        deleted = _delete_upload_batch_for_firm(active_firm.id, normalized_batch)
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Upload batch delete failed", endpoint="upload")

    total_deleted = sum(int(v or 0) for v in deleted.values())
    if total_deleted == 0:
        flash(
            f"No records found for upload batch {normalized_batch} in {active_firm.name}.",
            "warning",
        )
        return redirect(url_for("upload"))

    flash(
        f"Deleted upload batch {normalized_batch} from {active_firm.name}.",
        "success",
    )
    flash(
        "Removed "
        f"{deleted['deals']} deals, {deleted['cashflows']} cashflow events, "
        f"{deleted['deal_quarterly']} deal quarterly rows, {deleted['fund_quarterly']} fund quarterly rows, "
        f"{deleted['underwrite']} underwrite rows, {deleted['fund_metadata']} fund metadata rows, "
        f"{deleted['fund_cashflows']} fund cashflow rows, {deleted['public_market_benchmarks']} public market rows, "
        f"and {deleted['upload_issues']} upload issue rows.",
        "info",
    )
    return redirect(url_for("upload"))


@app.route("/upload/deals/template")
@login_required
def download_deal_template():
    template_path = Path(app.root_path) / DEAL_TEMPLATE_FILENAME
    if not template_path.exists():
        logger.error("Deal template file not found at %s", template_path)
        flash("Deal template file is unavailable right now.", "danger")
        return redirect(url_for("upload"))

    return send_file(
        template_path,
        as_attachment=True,
        download_name=DEAL_TEMPLATE_FILENAME,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/upload/benchmarks/template")
@login_required
def download_benchmark_template():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmarks"
    ws.append(["Asset Class", "Vintage Year", "Metric", "Quartile", "Value"])
    ws.append(["Buyout", 2019, "Net IRR", "Median", 0.18])
    ws.append(["Buyout", 2019, "Net TVPI", "Upper Quartile", 2.1])
    ws.append(["Buyout", 2019, "Net DPI", "Top 5%", 1.7])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=BENCHMARK_TEMPLATE_FILENAME,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/upload/deals", methods=["POST"])
@login_required
@limiter.limit("20/hour")
def upload_deals():
    return _handle_upload(parse_deals, "deals")


@app.route("/upload/benchmarks", methods=["POST"])
@login_required
@limiter.limit("20/hour")
def upload_benchmarks():
    membership = _require_team_scope()

    if "file" not in request.files:
        flash("No benchmark file selected.", "danger")
        return redirect(url_for("upload"))

    file = request.files["file"]
    if file.filename == "":
        flash("No benchmark file selected.", "danger")
        return redirect(url_for("upload"))

    if not _allowed_file(file.filename):
        flash("Invalid benchmark file type. Please upload an .xlsx or .xls file.", "danger")
        return redirect(url_for("upload"))

    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(file_path)

    try:
        result = parse_benchmarks(file_path, team_id=membership.team_id, replace_mode="replace_all")
        if result.get("success"):
            flash(
                f"Loaded {result.get('rows_loaded', 0)} benchmark rows "
                f"(batch {result.get('upload_batch')}).",
                "success",
            )
            asset_classes = result.get("asset_classes") or []
            if asset_classes:
                flash("Asset classes: " + ", ".join(asset_classes), "info")
            vintage_min = result.get("vintage_min")
            vintage_max = result.get("vintage_max")
            if vintage_min is not None and vintage_max is not None:
                flash(f"Benchmark vintage coverage: {vintage_min} to {vintage_max}.", "info")
        else:
            flash("Benchmark upload failed.", "danger")

        for err in (result.get("errors") or [])[:10]:
            flash(err, "warning")
        if len(result.get("errors") or []) > 10:
            flash(f"...and {len(result['errors']) - 10} additional benchmark upload errors.", "warning")
    except Exception as exc:
        logger.exception("Error processing benchmark upload")
        flash(f"Error processing benchmark file: {str(exc)}", "danger")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

    return redirect(url_for("upload"))


@app.route("/upload/benchmarks/delete", methods=["POST"])
@login_required
def delete_benchmarks():
    membership = _require_team_scope()
    try:
        deleted = BenchmarkPoint.query.filter_by(team_id=membership.team_id).delete(synchronize_session=False)
        db.session.commit()
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Benchmark delete failed", endpoint="upload")

    if deleted == 0:
        flash("No benchmark dataset loaded for your team.", "warning")
    else:
        flash(f"Deleted benchmark dataset for your team ({deleted} rows).", "success")
    return redirect(url_for("upload"))


@app.route("/deals")
@login_required
def deals():
    try:
        filter_ctx = _build_filtered_deals_context()
        all_deals = filter_ctx["deals"]
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Deals page failed")

    return _render_deals_page(
        all_deals=all_deals,
        filter_ctx=filter_ctx,
        calculation_deals=all_deals,
        selection_enabled=False,
        page_title="Deals",
    )


def _parse_deal_id_args(*names):
    ids = set()
    for name in names:
        for raw_value in request.args.getlist(name):
            for part in str(raw_value or "").split(","):
                try:
                    deal_id = int(part.strip())
                except (TypeError, ValueError):
                    continue
                if deal_id > 0:
                    ids.add(deal_id)
    return ids


def _selected_deals_for_analysis(all_deals):
    all_deal_ids = {d.id for d in all_deals if d.id is not None}
    explicit_include_ids = _parse_deal_id_args("include_deal", "included_deal")
    if explicit_include_ids:
        included_ids = all_deal_ids & explicit_include_ids
    else:
        excluded_ids = _parse_deal_id_args("exclude_deal", "excluded_deal")
        included_ids = all_deal_ids - excluded_ids
    return [d for d in all_deals if d.id in included_ids], included_ids


def _track_record_funds_by_name(track_record):
    return {
        fund.get("fund_name") or "Unknown Fund": fund
        for fund in track_record.get("funds", [])
    }


def _render_deals_page(all_deals, filter_ctx, calculation_deals, selection_enabled=False, included_deal_ids=None, page_title="Deals"):
    deal_metrics = {d.id: compute_deal_metrics(d) for d in all_deals}
    calculation_metrics = {
        d.id: deal_metrics[d.id]
        for d in calculation_deals
        if d.id in deal_metrics
    }
    fund_vintage_lookup = _fund_vintage_lookup_for_scope(
        all_deals,
        membership=filter_ctx.get("active_membership"),
        active_firm=filter_ctx.get("active_firm"),
    )
    display_track_record = None
    if selection_enabled:
        display_track_record = compute_deal_track_record(
            all_deals,
            metrics_by_id=deal_metrics,
            fund_vintage_lookup=fund_vintage_lookup,
        )
    track_record = compute_deal_track_record(
        calculation_deals,
        metrics_by_id=calculation_metrics,
        fund_vintage_lookup=fund_vintage_lookup,
    )
    rollup_details = compute_deals_rollup_details(calculation_deals, track_record, metrics_by_id=calculation_metrics)
    reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
    scale = reporting["money_scale"]
    _scale_deal_metrics(deal_metrics, scale)
    _scale_track_record_payload(track_record, scale)
    _scale_rollup_details_payload(rollup_details, scale)
    deals_by_id = {d.id: d for d in all_deals}
    included_deal_ids = included_deal_ids if included_deal_ids is not None else {d.id for d in calculation_deals}
    return render_template(
        "deals.html",
        deals=all_deals,
        deal_metrics=deal_metrics,
        track_record=track_record,
        display_track_record=display_track_record,
        selected_track_record_funds_by_name=_track_record_funds_by_name(track_record),
        rollup_details=rollup_details,
        deals_by_id=deals_by_id,
        selection_enabled=selection_enabled,
        included_deal_ids=included_deal_ids,
        selected_deal_count=len(calculation_deals),
        total_deal_count=len(all_deals),
        deals_page_title=page_title,
    )


@app.route("/deals/analysis")
@login_required
def deals_analysis():
    try:
        filter_ctx = _build_filtered_deals_context()
        all_deals = filter_ctx["deals"]
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Deals analysis page failed")

    selected_deals, included_deal_ids = _selected_deals_for_analysis(all_deals)
    return _render_deals_page(
        all_deals=all_deals,
        filter_ctx=filter_ctx,
        calculation_deals=selected_deals,
        selection_enabled=True,
        included_deal_ids=included_deal_ids,
        page_title="Deals Analysis",
    )


@app.route("/track-record")
@login_required
def track_record():
    try:
        filter_ctx = _build_filtered_deals_context()
        all_deals = filter_ctx["deals"]
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Track record page failed")

    metrics_by_id = {d.id: compute_deal_metrics(d) for d in all_deals}
    fund_vintage_lookup = _fund_vintage_lookup_for_scope(
        all_deals,
        membership=filter_ctx.get("active_membership"),
        active_firm=filter_ctx.get("active_firm"),
    )
    record = compute_deal_track_record(all_deals, metrics_by_id=metrics_by_id, fund_vintage_lookup=fund_vintage_lookup)
    reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
    _scale_track_record_payload(record, reporting["money_scale"])
    return render_template("track_record.html", track_record=record, deals=all_deals)


@app.route("/track-record/pdf")
@login_required
def download_track_record_pdf():
    try:
        filter_ctx = _build_filtered_deals_context()
        all_deals = filter_ctx["deals"]
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Track record PDF failed")

    metrics_by_id = {d.id: compute_deal_metrics(d) for d in all_deals}
    fund_vintage_lookup = _fund_vintage_lookup_for_scope(
        all_deals,
        membership=filter_ctx.get("active_membership"),
        active_firm=filter_ctx.get("active_firm"),
    )
    record = compute_deal_track_record(all_deals, metrics_by_id=metrics_by_id, fund_vintage_lookup=fund_vintage_lookup)
    reporting = _reporting_currency_context(filter_ctx.get("active_firm"))
    _scale_track_record_payload(record, reporting["money_scale"])

    currency_code = reporting["reporting_currency_code"]
    report_title = (request.args.get("report_title", "") or "").strip() or None
    download_name = _safe_pdf_download_name(
        request.args.get("download_name", ""),
        "track_record_print_ready.pdf",
    )

    try:
        pdf_bytes = _build_track_record_pdf(
            record,
            currency_code=currency_code,
            report_title=report_title,
        )
    except ImportError:
        flash("PDF export dependency missing. Install requirements and retry.", "danger")
        return redirect(url_for("track_record"))

    return send_file(
        BytesIO(pdf_bytes),
        as_attachment=True,
        download_name=download_name,
        mimetype="application/pdf",
    )


@app.route("/credit/analysis/<page>/pdf")
@login_required
def download_credit_analysis_pdf(page):
    if page not in CREDIT_ANALYSIS_PAGES:
        abort(404)

    try:
        export_ctx = _build_credit_pdf_export_context()
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Credit PDF export failed")

    if export_ctx is None:
        flash("No active firm found for credit export.", "warning")
        return redirect(url_for("upload_credit_loans"))

    loans = export_ctx["ctx"]["loans"]
    if not loans:
        flash("No credit loans found for the active firm.", "warning")
        return redirect(url_for("upload_credit_loans"))

    active_firm = export_ctx["active_firm"]
    fund_performance = export_ctx["ctx"].get("fund_performance", {})
    as_of_date = _resolve_credit_analysis_as_of_date(loans, fund_performance)
    reporting = _reporting_currency_context(active_firm)
    report_title = _report_title(
        active_firm.name or "Unknown Firm",
        next(
            (item["analysis_name"] for item in CREDIT_PDF_EXPORT_PAGES if item["page"] == page),
            CREDIT_ANALYSIS_PAGES[page]["title"],
        ),
        as_of_date,
    )
    download_name = _safe_pdf_download_name(
        request.args.get("download_name", ""),
        f"{report_title}.pdf",
    )

    try:
        payload = _credit_pdf_payload_for_page(page, export_ctx)
        pdf_bytes = _build_credit_analysis_pdf(
            page,
            payload,
            report_title=report_title,
            currency_code=reporting["reporting_currency_code"],
            as_of_date=as_of_date,
            benchmark_asset_class=export_ctx["benchmark_asset_class"],
        )
    except ImportError:
        flash("PDF export dependency missing. Install requirements and retry.", "danger")
        return redirect(url_for("live_credit_pdf_pack"))

    return send_file(
        BytesIO(pdf_bytes),
        as_attachment=True,
        download_name=download_name,
        mimetype="application/pdf",
    )


@app.route("/reports/credit-pdf-pack/live")
@login_required
def live_credit_pdf_pack():
    try:
        export_ctx = _build_credit_pdf_export_context()
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Credit PDF link pack failed")

    if export_ctx is None:
        flash("No active firm found for credit export.", "warning")
        return redirect(url_for("upload_credit_loans"))

    active_firm = export_ctx["active_firm"]
    loans = export_ctx["ctx"]["loans"]
    if not loans:
        flash("No credit loans found for the active firm.", "warning")
        return redirect(url_for("upload_credit_loans"))

    as_of_date = _resolve_credit_analysis_as_of_date(
        loans,
        export_ctx["ctx"].get("fund_performance", {}),
    )
    firm_name = active_firm.name or "Unknown Firm"

    links = []
    for item in CREDIT_PDF_EXPORT_PAGES:
        title = _report_title(firm_name, item["analysis_name"], as_of_date)
        links.append(
            {
                "label": item["analysis_name"],
                "title": title,
                "url": url_for(
                    "download_credit_analysis_pdf",
                    page=item["page"],
                    download_name=f"{title}.pdf",
                ),
            }
        )

    return render_template(
        "reports_credit_pdf_pack_live.html",
        links=links,
        bundle_url=url_for("download_credit_pdf_pack"),
        benchmark_asset_class=export_ctx["benchmark_asset_class"],
        as_of_date=as_of_date,
    )


@app.route("/reports/credit-pdf-pack")
@login_required
def download_credit_pdf_pack():
    try:
        export_ctx = _build_credit_pdf_export_context()
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "Credit PDF pack download failed")

    if export_ctx is None:
        flash("No active firm found for credit export.", "warning")
        return redirect(url_for("upload_credit_loans"))

    active_firm = export_ctx["active_firm"]
    loans = export_ctx["ctx"]["loans"]
    if not loans:
        flash("No credit loans found for the active firm.", "warning")
        return redirect(url_for("upload_credit_loans"))

    fund_performance = export_ctx["ctx"].get("fund_performance", {})
    as_of_date = _resolve_credit_analysis_as_of_date(loans, fund_performance)
    reporting = _reporting_currency_context(active_firm)
    currency_code = reporting["reporting_currency_code"]
    firm_name = active_firm.name or "Unknown Firm"

    try:
        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, mode="w", compression=ZIP_DEFLATED) as archive:
            for item in CREDIT_PDF_EXPORT_PAGES:
                title = _report_title(firm_name, item["analysis_name"], as_of_date)
                payload = _credit_pdf_payload_for_page(item["page"], export_ctx)
                pdf_bytes = _build_credit_analysis_pdf(
                    item["page"],
                    payload,
                    report_title=title,
                    currency_code=currency_code,
                    as_of_date=as_of_date,
                    benchmark_asset_class=export_ctx["benchmark_asset_class"],
                )
                archive.writestr(f"{_sanitize_filename_component(title)}.pdf", pdf_bytes)
        zip_buffer.seek(0)
    except ImportError:
        flash("PDF export dependency missing. Install requirements and retry.", "danger")
        return redirect(url_for("live_credit_pdf_pack"))

    bundle_name = f"{_sanitize_filename_component(firm_name)} Credit Analysis PDF Pack {_as_of_ymd(as_of_date)}.zip"
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=bundle_name,
        mimetype="application/zip",
    )


@app.route("/reports/ic-pdf-pack/live")
@login_required
def live_ic_pdf_pack():
    membership = _require_team_scope()
    active_firm = _resolve_active_firm_for_team()
    if active_firm is None:
        flash("No active firm found for export.", "warning")
        return redirect(url_for("dashboard"))

    try:
        all_deals = Deal.query.filter_by(firm_id=active_firm.id).all()
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "IC PDF link pack failed")

    benchmark_asset_classes = _benchmark_asset_classes_for_team(membership.team_id)
    benchmark_session_key = "selected_benchmark_asset_class"
    current_benchmark_asset_class = (session.get(benchmark_session_key, "") or "").strip()
    if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
        current_benchmark_asset_class = ""
        session[benchmark_session_key] = ""

    as_of_date = resolve_analysis_as_of_date(all_deals)
    firm_name = active_firm.name or "Unknown Firm"

    track_title = _report_title(firm_name, "Deal Level Track Record", as_of_date)
    ebitda_title = _report_title(firm_name, "Value Creation Analysis by EBITDA", as_of_date)
    revenue_title = _report_title(firm_name, "Value Creation Analysis by Revenue", as_of_date)
    benchmarking_title = _report_title(firm_name, "Benchmarking Analysis", as_of_date)

    links = [
        {
            "label": "Track Record PDF",
            "title": track_title,
            "url": url_for(
                "download_track_record_pdf",
                report_title=track_title,
                download_name=f"{track_title}.pdf",
            ),
        },
        {
            "label": "VCA by EBITDA (Print Dialog)",
            "title": ebitda_title,
            "url": url_for(
                "analysis_page",
                page="vca-ebitda",
                autoprint="1",
                autoclose="1",
                pdf_title=ebitda_title,
            ),
        },
        {
            "label": "VCA by Revenue (Print Dialog)",
            "title": revenue_title,
            "url": url_for(
                "analysis_page",
                page="vca-revenue",
                autoprint="1",
                autoclose="1",
                pdf_title=revenue_title,
            ),
        },
        {
            "label": "Benchmarking Analysis (Print Dialog)",
            "title": benchmarking_title,
            "url": url_for(
                "analysis_page",
                page="benchmarking",
                benchmark_asset_class=current_benchmark_asset_class,
                autoprint="1",
                autoclose="1",
                pdf_title=benchmarking_title,
            ),
        },
    ]

    return render_template(
        "reports_ic_pdf_pack_live.html",
        links=links,
        benchmark_asset_class=current_benchmark_asset_class,
    )


@app.route("/reports/ic-pdf-pack")
@login_required
def download_ic_pdf_pack():
    membership = _require_team_scope()
    team_id = membership.team_id
    active_firm = _resolve_active_firm_for_team()
    if active_firm is None:
        flash("No active firm found for export.", "warning")
        return redirect(url_for("dashboard"))

    benchmark_asset_classes = _benchmark_asset_classes_for_team(team_id)
    benchmark_session_key = "selected_benchmark_asset_class"
    requested_benchmark = request.args.get("benchmark_asset_class")
    if requested_benchmark is not None:
        current_benchmark_asset_class = (requested_benchmark or "").strip()
        if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
            current_benchmark_asset_class = ""
        session[benchmark_session_key] = current_benchmark_asset_class
    else:
        current_benchmark_asset_class = (session.get(benchmark_session_key, "") or "").strip()
        if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
            current_benchmark_asset_class = ""
            session[benchmark_session_key] = ""

    try:
        all_deals = Deal.query.filter_by(firm_id=active_firm.id).all()
    except SQLAlchemyError as exc:
        return _redirect_schema_failure(exc, "IC PDF pack download failed")

    metrics_by_id = {deal.id: compute_deal_metrics(deal) for deal in all_deals}
    as_of_date = resolve_analysis_as_of_date(all_deals)
    fund_vintage_lookup = _fund_vintage_lookup_for_scope(all_deals, membership=membership, active_firm=active_firm)

    track_record_payload = compute_deal_track_record(all_deals, metrics_by_id=metrics_by_id, fund_vintage_lookup=fund_vintage_lookup)
    vca_ebitda_payload = compute_vca_ebitda_analysis(all_deals, metrics_by_id=metrics_by_id)
    vca_revenue_payload = compute_vca_revenue_analysis(all_deals, metrics_by_id=metrics_by_id)
    benchmark_thresholds = _load_team_benchmark_thresholds(team_id, current_benchmark_asset_class)
    benchmarking_payload = compute_benchmarking_analysis(
        all_deals,
        benchmark_thresholds=benchmark_thresholds,
        benchmark_asset_class=current_benchmark_asset_class,
        metrics_by_id=metrics_by_id,
        fund_vintage_lookup=fund_vintage_lookup,
    )

    reporting = _reporting_currency_context(active_firm)
    currency_code = reporting["reporting_currency_code"]
    scale = reporting["money_scale"]
    _scale_track_record_payload(track_record_payload, scale)
    _scale_analysis_payload("vca-ebitda", vca_ebitda_payload, scale)
    _scale_analysis_payload("vca-revenue", vca_revenue_payload, scale)
    _scale_analysis_payload("benchmarking", benchmarking_payload, scale)

    firm_name_raw = active_firm.name or "Unknown Firm"
    firm_name_file = _sanitize_filename_component(firm_name_raw)
    track_title = _report_title(firm_name_raw, "Deal Level Track Record", as_of_date)
    ebitda_title = _report_title(
        firm_name_raw,
        "Value Creation Analysis by EBITDA",
        (vca_ebitda_payload.get("meta") or {}).get("as_of_date"),
    )
    revenue_title = _report_title(
        firm_name_raw,
        "Value Creation Analysis by Revenue",
        (vca_revenue_payload.get("meta") or {}).get("as_of_date"),
    )
    benchmarking_title = _report_title(
        firm_name_raw,
        "Benchmarking Analysis",
        (benchmarking_payload.get("meta") or {}).get("as_of_date"),
    )

    try:
        track_pdf = _build_track_record_pdf(
            track_record_payload,
            currency_code=currency_code,
            report_title=track_title,
        )
        vca_ebitda_pdf = _build_vca_pdf(
            vca_ebitda_payload,
            report_title=ebitda_title,
            currency_code=currency_code,
            analysis_kind="ebitda",
        )
        vca_revenue_pdf = _build_vca_pdf(
            vca_revenue_payload,
            report_title=revenue_title,
            currency_code=currency_code,
            analysis_kind="revenue",
        )
        benchmarking_pdf = _build_benchmarking_pdf(
            benchmarking_payload,
            report_title=benchmarking_title,
            currency_code=currency_code,
            benchmark_asset_class=current_benchmark_asset_class,
        )
    except ImportError:
        flash("PDF export dependency missing. Install requirements and retry.", "danger")
        return redirect(url_for("dashboard"))

    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, mode="w", compression=ZIP_DEFLATED) as archive:
        archive.writestr(f"{_sanitize_filename_component(track_title)}.pdf", track_pdf)
        archive.writestr(f"{_sanitize_filename_component(ebitda_title)}.pdf", vca_ebitda_pdf)
        archive.writestr(f"{_sanitize_filename_component(revenue_title)}.pdf", vca_revenue_pdf)
        archive.writestr(f"{_sanitize_filename_component(benchmarking_title)}.pdf", benchmarking_pdf)
    zip_buffer.seek(0)

    bundle_name = f"{firm_name_file} Analysis PDF Pack {_as_of_ymd(as_of_date)}.zip"
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=bundle_name,
        mimetype="application/zip",
    )


MEMO_DOCUMENT_ROLES = {
    "prior_memo",
    "ddq",
    "ppm",
    "overview_deck",
    "supporting_material",
    "approved_generated_memo",
}
MEMO_STYLE_ROLES = {"prior_memo", "approved_generated_memo"}


def _memo_payload():
    payload = request.get_json(silent=True)
    if isinstance(payload, dict):
        return payload
    return request.form


def _memo_parse_document_ids(raw_value):
    if raw_value is None:
        return []
    if isinstance(raw_value, list):
        values = raw_value
    else:
        text_value = str(raw_value).strip()
        if not text_value:
            return []
        try:
            decoded = json.loads(text_value)
            if isinstance(decoded, list):
                values = decoded
            else:
                values = [part.strip() for part in text_value.split(",")]
        except (TypeError, ValueError):
            values = [part.strip() for part in text_value.split(",")]
    out = []
    for value in values:
        try:
            out.append(int(value))
        except (TypeError, ValueError):
            continue
    return sorted({value for value in out if value > 0})


def _memo_allowed_file(filename):
    return os.path.splitext(filename or "")[1].lower() in app.config["MEMO_ALLOWED_EXTENSIONS"]


def _memo_firm_scope(membership, requested_firm_id=None, allow_null=False):
    accessible_ids = {firm.id for firm in _accessible_firms_for_team(membership.team_id)}
    if requested_firm_id in (None, "", "null"):
        active_firm = _resolve_active_firm_for_team()
        if active_firm is not None and active_firm.id in accessible_ids:
            return active_firm.id
        return None if allow_null else abort(400)
    try:
        firm_id = int(requested_firm_id)
    except (TypeError, ValueError):
        abort(400)
    if firm_id not in accessible_ids:
        abort(403)
    return firm_id


def _memo_document_query(membership):
    return MemoDocument.query.filter(
        MemoDocument.team_id == membership.team_id,
        MemoDocument.status != "deleted",
    )


def _memo_style_profile_query(membership):
    return MemoStyleProfile.query.filter(
        MemoStyleProfile.team_id == membership.team_id,
        MemoStyleProfile.created_by_user_id == current_user.id,
        MemoStyleProfile.status != "deleted",
    )


def _memo_run_query(membership):
    return MemoGenerationRun.query.filter(MemoGenerationRun.team_id == membership.team_id)


def _memo_document_or_404(membership, document_id):
    row = _memo_document_query(membership).filter(MemoDocument.id == document_id).first()
    if row is None:
        abort(404)
    if row.document_role in MEMO_STYLE_ROLES and row.created_by_user_id != current_user.id:
        abort(403)
    return row


def _memo_style_profile_or_404(membership, profile_id):
    row = _memo_style_profile_query(membership).filter(MemoStyleProfile.id == profile_id).first()
    if row is None:
        abort(404)
    return row


def _memo_run_or_404(membership, run_id):
    row = _memo_run_query(membership).filter(MemoGenerationRun.id == run_id).first()
    if row is None:
        abort(404)
    return row


def _memo_resolve_documents(membership, document_ids):
    unique_ids = sorted({int(value) for value in (document_ids or []) if int(value) > 0})
    if not unique_ids:
        return []
    rows = (
        _memo_document_query(membership)
        .filter(MemoDocument.id.in_(unique_ids))
        .order_by(MemoDocument.created_at.asc(), MemoDocument.id.asc())
        .all()
    )
    if len(rows) != len(unique_ids):
        abort(404)
    for row in rows:
        if row.document_role in MEMO_STYLE_ROLES and row.created_by_user_id != current_user.id:
            abort(403)
    return rows


def _memo_parse_json(value, default):
    if not value:
        return default
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return default


def _memo_json_dumps(value):
    return json.dumps(value, sort_keys=True, default=str)


def _memo_status_tone(status):
    normalized = (status or "").strip().lower()
    if normalized in {"ready", "approved", "review_required", "reviewed"}:
        return "ready"
    if normalized == "canceled":
        return "attention"
    if normalized in {"failed", "blocked", "error"}:
        return "failed"
    if normalized in {"queued", "running", "uploaded", "processing", "pending"}:
        return "processing"
    return "attention"


def _memo_human_stage(progress_stage):
    stage = (progress_stage or "").strip()
    if not stage:
        return "Preparing memo generation."
    if stage == "queued":
        return "Queued. The memo engine has accepted your request and is preparing the run."
    if stage == "building_evidence":
        return "Collecting app analytics, extracted facts, and document citations."
    if stage == "drafting":
        return "Drafting sections in your style."
    if stage == "review":
        return "Draft complete. Review each section, then approve the memo when it meets your standard."
    if stage == "approved":
        return "Memo approved and ready for export."
    if stage == "canceled":
        return "Run canceled. No additional sections will be generated for this memo."
    if stage == "failed":
        return "This run did not complete. Review the error details below, correct the input issue, and rerun the affected section or restart the memo."
    if ":" in stage:
        prefix, suffix = stage.split(":", 1)
        label = suffix.replace("_", " ").strip()
        if prefix == "drafting":
            return f"Drafting {label} in your style."
        if prefix == "rerunning":
            return f"Rerunning {label} with the latest evidence."
    return stage.replace("_", " ").strip().capitalize()


def _memo_run_latest_job_label(run_status, progress_stage, latest_job):
    if latest_job and latest_job.status == "canceled":
        return "Background job canceled."
    if latest_job and latest_job.status == "failed":
        return "Background job failed."
    if latest_job and latest_job.status == "running":
        return "Background job is running."
    if latest_job and latest_job.status == "queued":
        if latest_job.error_text and "lease expired" in latest_job.error_text.lower():
            return "Retrying after worker interruption."
        return "Waiting for the memo worker."
    if (run_status or "").strip().lower() == "canceled":
        return "Memo generation canceled."
    if (run_status or "").strip().lower() == "review_required":
        return "Draft ready for review."
    if (run_status or "").strip().lower() == "approved":
        return "Memo approved."
    if (run_status or "").strip().lower() == "failed":
        return "Memo generation failed."
    return _memo_human_stage(progress_stage)


def _memo_style_profile_name(profile_id):
    if not profile_id:
        return None
    row = db.session.get(MemoStyleProfile, profile_id)
    return row.name if row is not None else None


def _memo_source_documents_for_run(row):
    document_ids = _memo_parse_json(row.document_ids_json, [])
    if not document_ids:
        return []
    rows = (
        MemoDocument.query.filter(
            MemoDocument.team_id == row.team_id,
            MemoDocument.id.in_(document_ids),
        )
        .order_by(MemoDocument.created_at.asc(), MemoDocument.id.asc())
        .all()
    )
    return [_memo_serialize_document(item) for item in rows]


def _memo_reassemble_run(run):
    style_profile = load_style_profile(run.style_profile_id)
    section_rows = (
        MemoGenerationSection.query.filter_by(run_id=run.id)
        .order_by(MemoGenerationSection.section_order.asc(), MemoGenerationSection.id.asc())
        .all()
    )
    assembled = assemble_memo(
        style_profile,
        [
            {
                "section_key": row.section_key,
                "title": row.title,
                "draft_text": row.draft_text,
                "validation": _memo_parse_json(row.validation_json, {}),
                "review_status": row.review_status,
            }
            for row in section_rows
        ],
    )
    run.final_markdown = assembled.markdown
    run.final_html = assembled.html
    run.status = "review_required"
    run.progress_stage = "review"
    run.approved_at = None
    run.export_status = "not_requested"
    db.session.add(run)
    db.session.commit()
    return run


def _memo_archive_document(membership, document, deleted_by_user_id):
    storage = get_document_storage()
    try:
        storage.delete(document.storage_key)
    except FileNotFoundError:
        pass
    except Exception:
        logger.exception("Failed to remove memo document blob for document %s", document.id)

    metadata = _memo_parse_json(document.metadata_json, {})
    metadata["deleted_at"] = datetime.now(timezone.utc).isoformat()
    metadata["deleted_by_user_id"] = deleted_by_user_id
    metadata["removed_from_studio"] = True

    document.status = "deleted"
    document.extraction_status = "deleted"
    document.error_text = "Removed from AI Memo Studio."
    document.metadata_json = _memo_json_dumps(metadata)
    db.session.add(document)

    if document.document_role in MEMO_STYLE_ROLES:
        MemoStyleExemplar.query.filter_by(document_id=document.id).update(
            {"status": "deleted"},
            synchronize_session=False,
        )
        affected_profiles = _memo_style_profile_query(membership).all()
        db.session.flush()
        for profile in affected_profiles:
            profile.status = "queued"
            db.session.add(profile)
    db.session.commit()

    if document.document_role in MEMO_STYLE_ROLES:
        for profile in _memo_style_profile_query(membership).all():
            enqueue_job(
                profile.team_id,
                "rebuild_style_profile",
                {"style_profile_id": profile.id},
            )


def _memo_archive_style_profile(profile):
    profile.status = "deleted"
    db.session.add(profile)
    MemoStyleExemplar.query.filter_by(style_profile_id=profile.id).update(
        {"status": "deleted"},
        synchronize_session=False,
    )
    db.session.commit()


def _memo_cancel_run(run, canceled_by_user_id=None):
    normalized_status = (run.status or "").strip().lower()
    if normalized_status == "canceled":
        return run
    run.status = "canceled"
    run.progress_stage = "canceled"
    run.approved_at = None
    if run.export_status not in {"ready", "completed"}:
        run.export_status = "canceled"
    db.session.add(run)

    for section in MemoGenerationSection.query.filter_by(run_id=run.id).all():
        if section.status in {"pending", "queued", "running"}:
            section.status = "canceled"
            section.review_status = "canceled"
            db.session.add(section)

    cancel_jobs_for_run(
        run.id,
        message=(
            f"Canceled by user {canceled_by_user_id}."
            if canceled_by_user_id is not None
            else "Canceled by user."
        ),
    )
    db.session.commit()
    return run


def _memo_cancel_not_allowed_response(run):
    return jsonify(
        {
            "error": "cancel_not_allowed",
            "message": "Only queued or running memo runs can be canceled.",
            "status": run.status,
        }
    ), 409


def _memo_canceled_run_response():
    return jsonify(
        {
            "error": "run_canceled",
            "message": "This memo run was canceled. Start a new run to continue generation.",
        }
    ), 409


def _memo_section_or_404(run_id, section_key):
    row = MemoGenerationSection.query.filter_by(run_id=run_id, section_key=section_key).first()
    if row is None:
        abort(404)
    return row


def _memo_persist_section_claims(run, section_row, draft, validation_result):
    MemoGenerationClaim.query.filter_by(run_id=run.id, section_id=section_row.id).delete(synchronize_session=False)
    for claim in draft.claims:
        citation_ids = [citation_id for citation_id in (claim.get("citation_ids") or []) if citation_id]
        mismatch_reason = None
        for mismatch in validation_result.numeric_mismatches:
            if mismatch.get("claim_text") == claim.get("claim_text"):
                mismatch_reason = mismatch.get("reason")
                break
        db.session.add(
            MemoGenerationClaim(
                run_id=run.id,
                section_id=section_row.id,
                claim_type=claim.get("claim_type") or "synthesis",
                claim_text=claim.get("claim_text") or "",
                provenance_type=claim.get("provenance_type"),
                provenance_id=claim.get("provenance_id"),
                citation_json=_memo_json_dumps({"citation_ids": citation_ids}),
                validation_status="ready" if mismatch_reason is None else "mismatch",
                mismatch_reason=mismatch_reason,
                status="ready" if mismatch_reason is None else "blocked",
            )
        )


def _memo_serialize_document(row):
    return {
        "id": row.id,
        "team_id": row.team_id,
        "firm_id": row.firm_id,
        "created_by_user_id": row.created_by_user_id,
        "document_role": row.document_role,
        "file_name": row.file_name,
        "mime_type": row.mime_type,
        "status": row.status,
        "status_tone": _memo_status_tone(row.status),
        "extraction_status": row.extraction_status,
        "extraction_status_tone": _memo_status_tone(row.extraction_status),
        "page_count": row.page_count,
        "error_text": row.error_text,
        "metadata": _memo_parse_json(row.metadata_json, {}),
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


def _memo_serialize_style_profile(row):
    return {
        "id": row.id,
        "team_id": row.team_id,
        "created_by_user_id": row.created_by_user_id,
        "name": row.name,
        "status": row.status,
        "status_tone": _memo_status_tone(row.status),
        "profile": _memo_parse_json(row.profile_json, {}),
        "source_document_count": row.source_document_count,
        "approved_exemplar_count": row.approved_exemplar_count,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


def _memo_serialize_section(row):
    draft = _memo_parse_json(row.draft_json, {})
    validation = _memo_parse_json(row.validation_json, {})
    metadata = draft.get("metadata") or {}
    open_questions = draft.get("open_questions") or []
    return {
        "id": row.id,
        "section_key": row.section_key,
        "section_order": row.section_order,
        "title": row.title,
        "objective": row.objective,
        "draft_text": row.draft_text,
        "draft": draft,
        "metadata": metadata,
        "citations": draft.get("citations") or [],
        "citation_count": len(draft.get("citations") or []),
        "claims": draft.get("claims") or [],
        "open_questions": open_questions,
        "open_question_count": len(open_questions),
        "editor_notes": (metadata.get("editor_notes") or "").strip(),
        "validation": validation,
        "review_status": row.review_status,
        "review_status_tone": _memo_status_tone(row.review_status),
        "status": row.status,
        "status_tone": _memo_status_tone(row.status),
    }


def _memo_serialize_run(row):
    style_profile_name = _memo_style_profile_name(row.style_profile_id)
    latest_job = (
        MemoJob.query.filter(MemoJob.run_id == row.id)
        .order_by(MemoJob.created_at.desc(), MemoJob.id.desc())
        .first()
    )
    source_documents = _memo_source_documents_for_run(row)
    section_rows = (
        MemoGenerationSection.query.filter_by(run_id=row.id)
        .order_by(MemoGenerationSection.section_order.asc(), MemoGenerationSection.id.asc())
        .all()
    )
    ready_section_count = sum(1 for section in section_rows if section.status in {"ready", "approved"})
    reviewed_section_count = sum(1 for section in section_rows if section.review_status in {"reviewed", "approved"})
    return {
        "id": row.id,
        "team_id": row.team_id,
        "firm_id": row.firm_id,
        "created_by_user_id": row.created_by_user_id,
        "style_profile_id": row.style_profile_id,
        "style_profile_name": style_profile_name,
        "memo_type": row.memo_type,
        "status": row.status,
        "status_tone": _memo_status_tone(row.status),
        "progress_stage": row.progress_stage,
        "human_stage": _memo_human_stage(row.progress_stage),
        "benchmark_asset_class": row.benchmark_asset_class,
        "filters": _memo_parse_json(row.filters_json, {}),
        "document_ids": _memo_parse_json(row.document_ids_json, []),
        "source_documents": source_documents,
        "source_document_count": len(source_documents),
        "outline": _memo_parse_json(row.outline_json, {}),
        "final_markdown": row.final_markdown,
        "final_html": row.final_html,
        "missing_data": _memo_parse_json(row.missing_data_json, []),
        "conflicts": _memo_parse_json(row.conflicts_json, []),
        "open_questions": _memo_parse_json(row.open_questions_json, []),
        "export_status": row.export_status,
        "ready_section_count": ready_section_count,
        "reviewed_section_count": reviewed_section_count,
        "latest_job": {
            "id": latest_job.id,
            "job_type": latest_job.job_type,
            "status": latest_job.status,
            "status_tone": _memo_status_tone(latest_job.status),
            "attempt_count": latest_job.attempt_count,
            "error_text": latest_job.error_text,
            "updated_at": latest_job.updated_at.isoformat() if latest_job and latest_job.updated_at else None,
        } if latest_job else None,
        "latest_job_label": _memo_run_latest_job_label(row.status, row.progress_stage, latest_job),
        "can_cancel": (row.status or "").strip().lower() in {"queued", "running"},
        "approved_at": row.approved_at.isoformat() if row.approved_at else None,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        "section_summaries": [
            {
                "id": section.id,
                "section_key": section.section_key,
                "title": section.title,
                "status": section.status,
                "status_tone": _memo_status_tone(section.status),
                "review_status": section.review_status,
                "review_status_tone": _memo_status_tone(section.review_status),
            }
            for section in section_rows
        ],
    }


def _memo_create_approved_exemplar(run):
    if not run.final_markdown:
        return None
    storage = get_document_storage()
    filename = f"investment_memo_run_{run.id}.md"
    storage_key = build_storage_key(
        "approved-memos",
        str(run.team_id),
        str(run.created_by_user_id),
        f"{secrets.token_hex(8)}-{filename}",
    )
    storage.put(run.final_markdown.encode("utf-8"), storage_key)
    document = MemoDocument(
        team_id=run.team_id,
        firm_id=run.firm_id,
        created_by_user_id=run.created_by_user_id,
        document_role="approved_generated_memo",
        file_name=filename,
        mime_type="text/markdown",
        storage_key=storage_key,
        sha256=hashlib.sha256(run.final_markdown.encode("utf-8")).hexdigest(),
        status="uploaded",
        extraction_status="pending",
    )
    db.session.add(document)
    db.session.commit()
    enqueue_job(run.team_id, "extract_document", {"document_id": document.id})
    enqueue_job(run.team_id, "rebuild_style_profile", {"style_profile_id": run.style_profile_id})
    return document


@app.cli.command("memo-worker")
@click.option("--poll-interval", type=float, default=2.0, show_default=True, help="Seconds to sleep between queue polls.")
@with_appcontext
def memo_worker_command(poll_interval):
    """Run the memo generation background worker."""
    run_memo_worker(poll_interval=poll_interval)


@app.route("/memos")
@login_required
def memo_studio():
    membership = _require_team_scope()
    recover_stale_jobs()
    active_tab = "generate"
    documents = (
        _memo_document_query(membership)
        .filter(
            or_(
                MemoDocument.document_role.notin_(sorted(MEMO_STYLE_ROLES)),
                MemoDocument.created_by_user_id == current_user.id,
            )
        )
        .order_by(MemoDocument.created_at.desc(), MemoDocument.id.desc())
        .limit(50)
        .all()
    )
    profiles = _memo_style_profile_query(membership).order_by(MemoStyleProfile.updated_at.desc(), MemoStyleProfile.id.desc()).all()
    runs = _memo_run_query(membership).order_by(MemoGenerationRun.created_at.desc(), MemoGenerationRun.id.desc()).limit(25).all()
    benchmark_asset_classes = _benchmark_asset_classes_for_team(membership.team_id)
    current_benchmark_asset_class = (session.get("selected_benchmark_asset_class", "") or "").strip()
    if current_benchmark_asset_class and current_benchmark_asset_class not in benchmark_asset_classes:
        current_benchmark_asset_class = ""
    return render_template(
        "memo_studio.html",
        active_tab=active_tab,
        memo_documents=[_memo_serialize_document(row) for row in documents],
        memo_style_profiles=[_memo_serialize_style_profile(row) for row in profiles],
        memo_runs=[_memo_serialize_run(row) for row in runs],
        benchmark_asset_classes=benchmark_asset_classes,
        current_benchmark_asset_class=current_benchmark_asset_class,
    )


@app.route("/memos/style-library")
@login_required
def memo_style_library():
    membership = _require_team_scope()
    documents = (
        _memo_document_query(membership)
        .filter(
            MemoDocument.created_by_user_id == current_user.id,
            MemoDocument.document_role.in_(sorted(MEMO_STYLE_ROLES)),
        )
        .order_by(MemoDocument.created_at.desc(), MemoDocument.id.desc())
        .all()
    )
    profiles = _memo_style_profile_query(membership).order_by(MemoStyleProfile.updated_at.desc(), MemoStyleProfile.id.desc()).all()
    return render_template(
        "memo_studio.html",
        active_tab="style",
        memo_documents=[_memo_serialize_document(row) for row in documents],
        memo_style_profiles=[_memo_serialize_style_profile(row) for row in profiles],
        memo_runs=[],
        benchmark_asset_classes=_benchmark_asset_classes_for_team(membership.team_id),
        current_benchmark_asset_class=(session.get("selected_benchmark_asset_class", "") or "").strip(),
    )


@app.route("/memos/source-library")
@login_required
def memo_source_library():
    membership = _require_team_scope()
    documents = (
        _memo_document_query(membership)
        .filter(MemoDocument.document_role.notin_(sorted(MEMO_STYLE_ROLES)))
        .order_by(MemoDocument.created_at.desc(), MemoDocument.id.desc())
        .all()
    )
    return render_template(
        "memo_studio.html",
        active_tab="sources",
        memo_documents=[_memo_serialize_document(row) for row in documents],
        memo_style_profiles=[],
        memo_runs=[],
        benchmark_asset_classes=_benchmark_asset_classes_for_team(membership.team_id),
        current_benchmark_asset_class=(session.get("selected_benchmark_asset_class", "") or "").strip(),
    )


@app.route("/memos/runs/<int:run_id>")
@login_required
def memo_run_page(run_id):
    membership = _require_team_scope()
    recover_stale_jobs()
    run = _memo_run_or_404(membership, run_id)
    sections = (
        MemoGenerationSection.query.filter_by(run_id=run.id)
        .order_by(MemoGenerationSection.section_order.asc(), MemoGenerationSection.id.asc())
        .all()
    )
    return render_template(
        "memo_run.html",
        memo_run=_memo_serialize_run(run),
        memo_sections=[_memo_serialize_section(section) for section in sections],
    )


@app.route("/api/memos/documents", methods=["GET", "POST"])
@login_required
def memo_documents_api():
    membership = _require_team_scope()
    if request.method == "GET":
        role = (request.args.get("role") or "").strip().lower()
        query = _memo_document_query(membership)
        if role:
            query = query.filter(MemoDocument.document_role == role)
            if role in MEMO_STYLE_ROLES:
                query = query.filter(MemoDocument.created_by_user_id == current_user.id)
        rows = query.order_by(MemoDocument.created_at.desc(), MemoDocument.id.desc()).all()
        return jsonify({"items": [_memo_serialize_document(row) for row in rows]})

    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "file_required"}), 400
    if not _memo_allowed_file(upload.filename):
        return jsonify({"error": "unsupported_file_type"}), 400

    payload = _memo_payload()
    document_role = (payload.get("document_role") or "supporting_material").strip().lower()
    if document_role not in MEMO_DOCUMENT_ROLES:
        return jsonify({"error": "invalid_document_role"}), 400

    file_bytes = upload.read()
    max_bytes = int(app.config["MEMO_MAX_DOCUMENT_MB"]) * 1024 * 1024
    if len(file_bytes) > max_bytes:
        return jsonify({"error": "file_too_large", "message": f"Maximum file size is {app.config['MEMO_MAX_DOCUMENT_MB']} MB."}), 400

    firm_id = _memo_firm_scope(membership, payload.get("firm_id"), allow_null=document_role in MEMO_STYLE_ROLES)
    filename = secure_filename(upload.filename)
    storage = get_document_storage()
    storage_key = build_storage_key("memo-documents", str(membership.team_id), str(current_user.id), f"{secrets.token_hex(8)}-{filename}")
    storage.put(file_bytes, storage_key)

    document = MemoDocument(
        team_id=membership.team_id,
        firm_id=firm_id,
        created_by_user_id=current_user.id,
        document_role=document_role,
        file_name=filename,
        mime_type=upload.mimetype or "application/octet-stream",
        storage_key=storage_key,
        sha256=hashlib.sha256(file_bytes).hexdigest(),
        status="uploaded",
        extraction_status="pending",
    )
    db.session.add(document)
    db.session.commit()
    enqueue_job(membership.team_id, "extract_document", {"document_id": document.id})
    document = _memo_document_or_404(membership, document.id)
    return jsonify(_memo_serialize_document(document)), 201


@app.route("/api/memos/documents/<int:document_id>", methods=["DELETE"])
@login_required
def memo_document_delete_api(document_id):
    membership = _require_team_scope()
    document = _memo_document_or_404(membership, document_id)
    _memo_archive_document(membership, document, current_user.id)
    return jsonify({"id": document_id, "status": "deleted"})


@app.route("/api/memos/style-profiles", methods=["GET"])
@login_required
def memo_style_profiles_api():
    membership = _require_team_scope()
    rows = _memo_style_profile_query(membership).order_by(MemoStyleProfile.updated_at.desc(), MemoStyleProfile.id.desc()).all()
    payload = []
    for row in rows:
        item = _memo_serialize_style_profile(row)
        item["exemplars"] = list_style_exemplars(row.id)
        payload.append(item)
    return jsonify({"items": payload})


@app.route("/api/memos/style-profiles/<int:profile_id>", methods=["DELETE"])
@login_required
def memo_style_profile_delete_api(profile_id):
    membership = _require_team_scope()
    profile = _memo_style_profile_or_404(membership, profile_id)
    _memo_archive_style_profile(profile)
    return jsonify({"id": profile_id, "status": "deleted"})


@app.route("/api/memos/style-profiles/rebuild", methods=["POST"])
@login_required
def memo_style_profile_rebuild_api():
    membership = _require_team_scope()
    payload = _memo_payload()
    document_ids = _memo_parse_document_ids(payload.get("document_ids"))
    profile_id = payload.get("style_profile_id")
    profile_name = (payload.get("name") or "My Memo Style").strip()
    if profile_id:
        profile = _memo_style_profile_or_404(membership, int(profile_id))
        if profile_name:
            profile.name = profile_name
    else:
        profile = MemoStyleProfile(
            team_id=membership.team_id,
            created_by_user_id=current_user.id,
            name=profile_name,
            status="pending",
        )
        db.session.add(profile)
        db.session.commit()
    profile.status = "queued"
    db.session.add(profile)
    db.session.commit()
    enqueue_job(
        membership.team_id,
        "rebuild_style_profile",
        {"style_profile_id": profile.id, "document_ids": document_ids},
    )
    profile = _memo_style_profile_or_404(membership, profile.id)
    data = _memo_serialize_style_profile(profile)
    data["exemplars"] = list_style_exemplars(profile.id)
    return jsonify(data), 201 if not profile_id else 200


@app.route("/api/memos/runs", methods=["GET", "POST"])
@login_required
def memo_runs_api():
    membership = _require_team_scope()
    if request.method == "GET":
        recover_stale_jobs()
        rows = _memo_run_query(membership).order_by(MemoGenerationRun.created_at.desc(), MemoGenerationRun.id.desc()).all()
        return jsonify({"items": [_memo_serialize_run(row) for row in rows]})

    payload = _memo_payload()
    style_profile_id = payload.get("style_profile_id")
    if style_profile_id in (None, ""):
        return jsonify({"error": "style_profile_id_required"}), 400
    style_profile = _memo_style_profile_or_404(membership, int(style_profile_id))
    if style_profile.status != "ready":
        status_label = style_profile.status or "pending"
        return jsonify(
            {
                "error": "style_profile_not_ready",
                "message": f"Style profile '{style_profile.name}' is still {status_label}. Wait for it to finish processing before generating a memo.",
            }
        ), 409
    document_ids = _memo_parse_document_ids(payload.get("document_ids"))
    source_documents = _memo_resolve_documents(membership, document_ids)
    blocked_documents = [
        row.file_name
        for row in source_documents
        if row.status != "ready" or row.extraction_status != "ready"
    ]
    if blocked_documents:
        return jsonify(
            {
                "error": "documents_not_ready",
                "message": "Some selected source documents are still processing. Wait for them to reach Ready before generating a memo.",
                "documents": blocked_documents,
            }
        ), 409
    firm_id = _memo_firm_scope(membership, payload.get("firm_id"), allow_null=False)
    raw_filters = payload.get("filters")
    filters = raw_filters if isinstance(raw_filters, dict) else _memo_parse_json(raw_filters, {})
    available_benchmarks = set(_benchmark_asset_classes_for_team(membership.team_id))
    dashboard_benchmark = (session.get("selected_benchmark_asset_class") or "").strip()
    if dashboard_benchmark and dashboard_benchmark not in available_benchmarks:
        dashboard_benchmark = ""
    payload_benchmark = (payload.get("benchmark_asset_class") or "").strip()
    if payload_benchmark and payload_benchmark not in available_benchmarks:
        payload_benchmark = ""
    benchmark_asset_class = dashboard_benchmark or payload_benchmark

    run = MemoGenerationRun(
        team_id=membership.team_id,
        firm_id=firm_id,
        created_by_user_id=current_user.id,
        style_profile_id=style_profile.id,
        memo_type=(payload.get("memo_type") or "fund_investment").strip(),
        filters_json=json.dumps(filters, sort_keys=True),
        benchmark_asset_class=benchmark_asset_class,
        document_ids_json=json.dumps(document_ids),
        user_notes=(payload.get("user_notes") or "").strip() or None,
        status="queued",
        progress_stage="queued",
    )
    db.session.add(run)
    db.session.commit()
    enqueue_job(membership.team_id, "generate_memo_run", {"run_id": run.id}, run_id=run.id)
    return jsonify(_memo_serialize_run(run)), 201


@app.route("/api/memos/runs/<int:run_id>", methods=["GET"])
@login_required
def memo_run_api(run_id):
    membership = _require_team_scope()
    recover_stale_jobs()
    run = _memo_run_or_404(membership, run_id)
    return jsonify(_memo_serialize_run(run))


@app.route("/api/memos/runs/<int:run_id>/sections", methods=["GET"])
@login_required
def memo_run_sections_api(run_id):
    membership = _require_team_scope()
    recover_stale_jobs()
    run = _memo_run_or_404(membership, run_id)
    rows = (
        MemoGenerationSection.query.filter_by(run_id=run.id)
        .order_by(MemoGenerationSection.section_order.asc(), MemoGenerationSection.id.asc())
        .all()
    )
    return jsonify({"items": [_memo_serialize_section(row) for row in rows]})


@app.route("/api/memos/runs/<int:run_id>/sections/<section_key>", methods=["PATCH"])
@login_required
def memo_run_section_update_api(run_id, section_key):
    membership = _require_team_scope()
    run = _memo_run_or_404(membership, run_id)
    if run.status == "canceled":
        return _memo_canceled_run_response()
    section = _memo_section_or_404(run.id, section_key)
    payload = _memo_payload()
    draft_text = (payload.get("draft_text") or "").strip()
    if not draft_text:
        return jsonify({"error": "draft_text_required"}), 400

    existing_draft = _memo_parse_json(section.draft_json, {})
    citations = existing_draft.get("citations") or []
    citation_ids = [citation.get("id") for citation in citations if citation.get("id")][:3]
    paragraphs = [paragraph.strip() for paragraph in draft_text.split("\n\n") if paragraph.strip()]
    metadata = dict(existing_draft.get("metadata") or {})
    editor_notes = (payload.get("editor_notes") or "").strip()
    if editor_notes:
        metadata["editor_notes"] = editor_notes
    else:
        metadata.pop("editor_notes", None)
    metadata["manual_edit"] = True
    metadata["last_editor_user_id"] = current_user.id

    draft = DraftSection(
        key=section.section_key,
        title=section.title,
        text=draft_text,
        citations=citations,
        claims=extract_claims(section.section_key, draft_text, citations),
        open_questions=existing_draft.get("open_questions") or [],
        paragraph_map=[
            {"text": paragraph, "citation_ids": citation_ids}
            for paragraph in paragraphs
        ],
        metadata=metadata,
    )
    evidence_bundle = build_memo_evidence_bundle(run.id)
    style_profile = load_style_profile(run.style_profile_id)
    section_spec = {
        "key": section.section_key,
        "title": section.title,
        "objective": section.objective or "",
        "required_evidence": _memo_parse_json(section.required_evidence_json, []),
    }
    retrieval_pack = retrieve_section_evidence(section_spec, evidence_bundle, style_profile)
    validation_result = validate_section(draft, evidence_bundle, style_profile=style_profile, retrieval_pack=retrieval_pack)

    section.draft_text = draft.text
    section.draft_json = _memo_json_dumps(dataclass_to_dict(draft))
    section.validation_json = _memo_json_dumps(dataclass_to_dict(validation_result))
    section.status = validation_result.status
    section.review_status = "needs_review"
    section.approved_at = None
    db.session.add(section)
    db.session.flush()

    _memo_persist_section_claims(run, section, draft, validation_result)

    if run.status == "approved":
        run.approved_at = None
    db.session.add(run)
    db.session.commit()
    run = _memo_reassemble_run(run)
    section = _memo_section_or_404(run.id, section.section_key)
    return jsonify({"run": _memo_serialize_run(run), "section": _memo_serialize_section(section)})


@app.route("/api/memos/runs/<int:run_id>/sections/<section_key>/review", methods=["POST"])
@login_required
def memo_run_section_review_api(run_id, section_key):
    membership = _require_team_scope()
    run = _memo_run_or_404(membership, run_id)
    if run.status == "canceled":
        return _memo_canceled_run_response()
    section = _memo_section_or_404(run.id, section_key)
    payload = _memo_payload()
    review_status = (payload.get("review_status") or "reviewed").strip().lower()
    if review_status not in {"reviewed", "needs_review"}:
        return jsonify({"error": "invalid_review_status"}), 400
    section.review_status = review_status
    if review_status != "reviewed":
        section.approved_at = None
    db.session.add(section)
    if run.status == "approved":
        run.status = "review_required"
        run.progress_stage = "review"
        run.approved_at = None
        db.session.add(run)
    db.session.commit()
    return jsonify({"run": _memo_serialize_run(run), "section": _memo_serialize_section(section)})


@app.route("/api/memos/runs/<int:run_id>/cancel", methods=["POST"])
@login_required
def memo_run_cancel_api(run_id):
    membership = _require_team_scope()
    run = _memo_run_or_404(membership, run_id)
    normalized_status = (run.status or "").strip().lower()
    if normalized_status == "canceled":
        return jsonify(_memo_serialize_run(run))
    if normalized_status not in {"queued", "running"}:
        return _memo_cancel_not_allowed_response(run)
    run = _memo_cancel_run(run, canceled_by_user_id=current_user.id)
    return jsonify(_memo_serialize_run(run))


@app.route("/api/memos/runs/<int:run_id>/rerun-section", methods=["POST"])
@login_required
def memo_run_rerun_section_api(run_id):
    membership = _require_team_scope()
    run = _memo_run_or_404(membership, run_id)
    if run.status == "canceled":
        return _memo_canceled_run_response()
    payload = _memo_payload()
    section_key = (payload.get("section_key") or "").strip()
    if not section_key:
        return jsonify({"error": "section_key_required"}), 400
    enqueue_job(membership.team_id, "rerun_section", {"run_id": run.id, "section_key": section_key}, run_id=run.id)
    run = _memo_run_or_404(membership, run_id)
    return jsonify(_memo_serialize_run(run))


@app.route("/api/memos/runs/<int:run_id>/approve", methods=["POST"])
@login_required
def memo_run_approve_api(run_id):
    membership = _require_team_scope()
    run = _memo_run_or_404(membership, run_id)
    if run.status == "canceled":
        return _memo_canceled_run_response()
    rows = MemoGenerationSection.query.filter_by(run_id=run.id).all()
    blocked_sections = [
        row.section_key
        for row in rows
        if row.status != "ready" or row.review_status not in {"reviewed", "approved"}
    ]
    if blocked_sections:
        return jsonify({"error": "approval_blocked", "blocked_sections": blocked_sections}), 409

    now = _utc_now_naive()
    run.status = "approved"
    run.progress_stage = "approved"
    run.approved_at = now
    for row in rows:
        row.review_status = "approved"
        row.status = "approved"
        row.approved_at = now
        db.session.add(row)
    db.session.add(run)
    db.session.commit()
    _memo_create_approved_exemplar(run)
    return jsonify(_memo_serialize_run(run))


@app.route("/api/memos/runs/<int:run_id>/export", methods=["POST"])
@login_required
def memo_run_export_api(run_id):
    membership = _require_team_scope()
    run = _memo_run_or_404(membership, run_id)
    if run.status == "canceled":
        return _memo_canceled_run_response()
    payload = _memo_payload()
    export_format = (payload.get("format") or request.args.get("format") or "markdown").strip().lower()
    if export_format not in {"markdown", "html"}:
        return jsonify({"error": "unsupported_export_format"}), 400
    enqueue_job(membership.team_id, "export_memo", {"run_id": run.id}, run_id=run.id)
    run = _memo_run_or_404(membership, run.id)
    if export_format == "html":
        return send_file(
            BytesIO((run.final_html or "").encode("utf-8")),
            as_attachment=True,
            download_name=f"investment_memo_{run.id}.html",
            mimetype="text/html",
        )
    return send_file(
        BytesIO((run.final_markdown or "").encode("utf-8")),
        as_attachment=True,
        download_name=f"investment_memo_{run.id}.md",
        mimetype="text/markdown",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5001"))
    app.run(debug=True, port=port)
