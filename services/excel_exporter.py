"""Reconstruct a multi-sheet Excel workbook from DB data for a given firm.

Values are stored in USD in the database (converted at upload time).
On export, values are reverse-converted back to their original native currency
so the downloaded file matches what was originally uploaded.
"""

from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import or_

from models import (
    Deal,
    DealCashflowEvent,
    DealQuarterSnapshot,
    DealUnderwriteBaseline,
    Firm,
    FundCashflow,
    FundMetadata,
    FundQuarterSnapshot,
    db,
)
from services.deal_parser import FIN_METRIC_CURRENCY_COLS, PERF_CURRENCY_COLS


def _reverse_convert(value, fx_rate):
    """Convert a USD value back to native currency using the stored FX rate."""
    if value is None or fx_rate is None or fx_rate == 0 or fx_rate == 1.0:
        return value
    return value / fx_rate


def _team_scope(column, team_id):
    if team_id is None:
        return column.is_(None)
    return or_(column.is_(None), column == team_id)


def export_firm_to_excel(firm_id, team_id):
    """Build an .xlsx workbook with all data for *firm_id* and return a BytesIO buffer."""

    firm = db.session.get(Firm, firm_id)
    firm_name = firm.name if firm else ""
    firm_currency = (firm.base_currency or "USD") if firm else "USD"

    wb = Workbook()

    # ── Sheet 1: Deals (always present) ──────────────────────────────
    ws = wb.active
    ws.title = "Deals"

    deal_headers = [
        "Firm Name", "Company Name", "Fund", "Sector", "Geography", "Status",
        "Exit Type", "Lead Partner", "Security Type", "Deal Type", "Entry Channel",
        "Investment Date", "Year Invested", "Exit Date", "As Of Date",
        "Equity Invested", "Ownership %", "Fund Size",
        "Entry Revenue", "Entry EBITDA", "Entry TEV", "Entry Net Debt",
        "Exit Revenue", "Exit EBITDA", "Exit TEV", "Exit Net Debt",
        "Acquired Revenue", "Acquired EBITDA", "Acquired TEV",
        "Realized Value", "Unrealized Value", "IRR", "Net IRR", "Net MOIC", "DPI",
        "Firm Currency", "Performance Currency", "Financial Metric Currency",
    ]
    ws.append(deal_headers)

    deals = (
        Deal.query.filter(Deal.firm_id == firm_id, _team_scope(Deal.team_id, team_id))
        .order_by(Deal.fund_number, Deal.company_name)
        .all()
    )
    deal_map = {d.id: d for d in deals}
    deal_ids = list(deal_map.keys())

    for d in deals:
        # Reverse-convert USD values back to original native currency
        perf_rate = getattr(d, "perf_fx_rate_to_usd", None)
        fin_rate = getattr(d, "fin_fx_rate_to_usd", None)
        perf_ccy = getattr(d, "performance_currency", None) or firm_currency
        fin_ccy = getattr(d, "financial_metric_currency", None) or perf_ccy

        ws.append([
            firm_name, d.company_name, d.fund_number, d.sector, d.geography,
            d.status, d.exit_type, d.lead_partner, d.security_type,
            d.deal_type, d.entry_channel,
            d.investment_date, d.year_invested, d.exit_date, d.as_of_date,
            _reverse_convert(d.equity_invested, perf_rate), d.ownership_pct,
            _reverse_convert(d.fund_size, perf_rate),
            _reverse_convert(d.entry_revenue, fin_rate),
            _reverse_convert(d.entry_ebitda, fin_rate),
            _reverse_convert(d.entry_enterprise_value, fin_rate),
            _reverse_convert(d.entry_net_debt, fin_rate),
            _reverse_convert(d.exit_revenue, fin_rate),
            _reverse_convert(d.exit_ebitda, fin_rate),
            _reverse_convert(d.exit_enterprise_value, fin_rate),
            _reverse_convert(d.exit_net_debt, fin_rate),
            _reverse_convert(d.acquired_revenue, fin_rate),
            _reverse_convert(d.acquired_ebitda, fin_rate),
            _reverse_convert(d.acquired_tev, fin_rate),
            _reverse_convert(d.realized_value, perf_rate),
            _reverse_convert(d.unrealized_value, perf_rate),
            d.irr, d.net_irr, d.net_moic, d.net_dpi,
            perf_ccy,  # Firm Currency = performance currency
            perf_ccy,  # Performance Currency: original currency
            fin_ccy,   # Financial Metric Currency: original currency (may vary per deal)
        ])

    # ── Sheet 2: Cashflows ───────────────────────────────────────────
    cf_rows = []
    if deal_ids:
        cf_rows = (
            DealCashflowEvent.query.filter(
                DealCashflowEvent.firm_id == firm_id,
                _team_scope(DealCashflowEvent.team_id, team_id),
                DealCashflowEvent.deal_id.in_(deal_ids),
            )
            .order_by(DealCashflowEvent.event_date)
            .all()
        )
    if cf_rows:
        ws_cf = wb.create_sheet("Cashflows")
        ws_cf.append(["Company Name", "Fund", "Event Date", "Event Type", "Amount", "Notes"])
        for r in cf_rows:
            deal = deal_map.get(r.deal_id)
            ws_cf.append([
                deal.company_name if deal else "", deal.fund_number if deal else "",
                r.event_date, r.event_type, r.amount, r.notes,
            ])

    # ── Sheet 3: Deal Quarterly ──────────────────────────────────────
    dq_rows = []
    if deal_ids:
        dq_rows = (
            DealQuarterSnapshot.query.filter(
                DealQuarterSnapshot.firm_id == firm_id,
                _team_scope(DealQuarterSnapshot.team_id, team_id),
                DealQuarterSnapshot.deal_id.in_(deal_ids),
            )
            .order_by(DealQuarterSnapshot.quarter_end)
            .all()
        )
    if dq_rows:
        ws_dq = wb.create_sheet("Deal Quarterly")
        ws_dq.append([
            "Company Name", "Fund", "Quarter End", "Revenue", "EBITDA",
            "Enterprise Value", "Net Debt", "Equity Value", "Valuation Basis", "Source",
        ])
        for r in dq_rows:
            deal = deal_map.get(r.deal_id)
            ws_dq.append([
                deal.company_name if deal else "", deal.fund_number if deal else "",
                r.quarter_end, r.revenue, r.ebitda, r.enterprise_value,
                r.net_debt, r.equity_value, r.valuation_basis, r.source,
            ])

    # ── Sheet 4: Fund Quarterly ──────────────────────────────────────
    fq_rows = (
        FundQuarterSnapshot.query.filter(
            FundQuarterSnapshot.firm_id == firm_id,
            _team_scope(FundQuarterSnapshot.team_id, team_id),
        )
        .order_by(FundQuarterSnapshot.fund_number, FundQuarterSnapshot.quarter_end)
        .all()
    )
    if fq_rows:
        ws_fq = wb.create_sheet("Fund Quarterly")
        ws_fq.append([
            "Fund", "Quarter End", "Committed Capital", "Paid In Capital",
            "Distributed Capital", "NAV", "Unfunded Commitment",
        ])
        for r in fq_rows:
            ws_fq.append([
                r.fund_number, r.quarter_end, r.committed_capital,
                r.paid_in_capital, r.distributed_capital, r.nav, r.unfunded_commitment,
            ])

    # ── Sheet 5: Fund Metadata ───────────────────────────────────────
    fm_rows = (
        FundMetadata.query.filter(
            FundMetadata.firm_id == firm_id,
            _team_scope(FundMetadata.team_id, team_id),
        )
        .order_by(FundMetadata.fund_number)
        .all()
    )
    if fm_rows:
        ws_fm = wb.create_sheet("Fund Metadata")
        ws_fm.append([
            "Fund", "Vintage Year", "Strategy", "Region Focus", "Fund Size",
            "First Close Date", "Final Close Date", "Manager Name",
            "Benchmark Peer Group", "Status",
        ])
        for r in fm_rows:
            ws_fm.append([
                r.fund_number, r.vintage_year, r.strategy, r.region_focus,
                r.fund_size, r.first_close_date, r.final_close_date,
                r.manager_name, r.benchmark_peer_group, r.status,
            ])

    # ── Sheet 6: Fund Cashflows ──────────────────────────────────────
    fc_rows = (
        FundCashflow.query.filter(
            FundCashflow.firm_id == firm_id,
            _team_scope(FundCashflow.team_id, team_id),
        )
        .order_by(FundCashflow.fund_number, FundCashflow.event_date)
        .all()
    )
    if fc_rows:
        ws_fc = wb.create_sheet("Fund Cashflows")
        ws_fc.append(["Fund", "Event Date", "Event Type", "Amount", "NAV After Event", "Currency Code"])
        for r in fc_rows:
            ws_fc.append([
                r.fund_number, r.event_date, r.event_type,
                r.amount, r.nav_after_event, r.currency_code,
            ])

    # ── Sheet 7: Underwrite ──────────────────────────────────────────
    uw_rows = []
    if deal_ids:
        uw_rows = (
            DealUnderwriteBaseline.query.filter(
                DealUnderwriteBaseline.firm_id == firm_id,
                _team_scope(DealUnderwriteBaseline.team_id, team_id),
                DealUnderwriteBaseline.deal_id.in_(deal_ids),
            )
            .all()
        )
    if uw_rows:
        ws_uw = wb.create_sheet("Underwrite")
        ws_uw.append([
            "Company Name", "Fund", "Baseline Date", "Target IRR", "Target MOIC",
            "Target Hold Years", "Target Exit Multiple", "Target Revenue CAGR",
            "Target EBITDA CAGR",
        ])
        for r in uw_rows:
            deal = deal_map.get(r.deal_id)
            ws_uw.append([
                deal.company_name if deal else "", deal.fund_number if deal else "",
                r.baseline_date, r.target_irr, r.target_moic, r.target_hold_years,
                r.target_exit_multiple, r.target_revenue_cagr, r.target_ebitda_cagr,
            ])

    # ── Write to buffer ──────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
